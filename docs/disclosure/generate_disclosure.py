#!/usr/bin/env python3
"""Generate the JHI Coding Disclosure workbook (.xlsx).

Full disclosure of all coding performed, by component (subsystem), mapped to the
skillset / job title that would perform it, with a realistic hourly value per
action and an estimated equivalent professional value.

Run:  /workspace/.venv/bin/python docs/disclosure/generate_disclosure.py
Requires: openpyxl (dev/tooling only — not a runtime dependency).

DISCLAIMER: Hours are good-faith estimates of equivalent professional effort to
build the delivered code; the platform was built founder-led with AI assistance.
Rates are representative US contractor market rates. This workbook is a planning/
disclosure artifact, not an invoice, audited record, or tax/accounting advice.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "JHI_Coding_Disclosure.xlsx"

SIGNATURE = "69M2705M"

# --- Rate card: job title -> (level note, low, mid, high) -------------------
RATES: dict[str, tuple[str, int, int, int]] = {
    "Backend Engineer": ("Mid–Senior", 100, 140, 175),
    "Frontend Engineer": ("Mid–Senior", 85, 120, 150),
    "Security Engineer": ("Senior", 150, 190, 250),
    "Integration Engineer": ("Senior", 110, 150, 180),
    "AI Engineer": ("Mid–Senior", 150, 200, 350),
    "Data Scientist / Quant": ("Mid–Senior", 150, 200, 300),
    "DevOps / Cloud Engineer": ("Mid–Senior", 100, 160, 250),
    "QA Engineer": ("Mid–Senior", 60, 95, 120),
    "UX / UI Designer": ("Mid–Senior", 80, 120, 160),
    "Tech Lead / Architect": ("Senior", 150, 180, 250),
}
RATE_SOURCE = "Representative US contractor rates; anchored to docs/COMPENSATION_AND_PRO_SERVICES_PROJECTIONS.md"

# --- Work performed: (subsystem, action, job title, level, hours, files, status)
# hours = good-faith estimate of equivalent professional effort.
ROWS: list[tuple[str, str, str, str, float, str, str]] = [
    # Platform spine
    ("Platform spine", "FastAPI app entry, router registration, CORS, /health & /ready", "Backend Engineer", "Senior", 10, "backend/app/main.py", "Built"),
    ("Platform spine", "Runtime settings + production fail-fast validation", "Backend Engineer", "Senior", 6, "backend/app/config.py", "Built"),
    ("Platform spine", "Env-gated per-IP rate-limit middleware", "Backend Engineer", "Senior", 6, "backend/app/rate_limit.py", "Built"),
    ("Platform spine", "DB engine/session, init_db + seeding", "Backend Engineer", "Senior", 10, "backend/app/database.py", "Built"),
    ("Platform spine", "SQLAlchemy ORM models (all tables)", "Backend Engineer", "Senior", 20, "backend/app/db_models.py", "Built"),
    ("Platform spine", "Auth dependencies (principal, require_admin)", "Backend Engineer", "Senior", 4, "backend/app/dependencies.py", "Built"),
    ("Platform spine", "Shared Pydantic schemas (accounting/CRM/integration)", "Backend Engineer", "Mid", 12, "backend/app/models.py", "Built"),
    # Identity / Auth / Security
    ("Identity, Auth & Security", "Crypto primitives: PBKDF2, JWT, scoped tokens, TOTP", "Security Engineer", "Senior", 18, "backend/app/security.py", "Built"),
    ("Identity, Auth & Security", "Web auth API + multi-tenant register/login/me", "Backend Engineer", "Senior", 24, "routers/auth.py, foundation_services.py", "Built"),
    ("Identity, Auth & Security", "Mobile multi-factor (TOTP + biometric) service + API", "Backend Engineer", "Senior", 28, "routers/mobile_auth.py, mobile_services.py", "Built"),
    ("Identity, Auth & Security", "Encrypt TOTP secret at rest (Fernet, backward-compat)", "Security Engineer", "Senior", 8, "backend/app/security.py", "Built (PR #17)"),
    ("Identity, Auth & Security", "Replace hand-rolled JWT with vetted PyJWT", "Security Engineer", "Senior", 6, "backend/app/security.py", "Built (PR #17)"),
    ("Identity, Auth & Security", "Real WebAuthn ES256 assertion verification", "Security Engineer", "Senior", 16, "backend/app/webauthn.py", "Built (PR #17)"),
    ("Identity, Auth & Security", "Mobile app UI: password/2FA/biometric flows", "Frontend Engineer", "Senior", 24, "src/app/mobile/page.tsx", "Built"),
    # Billing
    ("Billing & Subscriptions", "Billing API: plans, checkout, subscription, audit logs", "Backend Engineer", "Senior", 14, "routers/billing.py, foundation_services.py", "Built"),
    ("Billing & Subscriptions", "Stripe webhook signature verification + event mapping", "Security Engineer", "Senior", 8, "backend/app/billing_webhook.py", "Built (PR #17)"),
    ("Billing & Subscriptions", "Pricing + account UI", "Frontend Engineer", "Mid", 10, "src/app/{pricing,account}/page.tsx", "Built"),
    # Accounting & Reporting
    ("Accounting & Reporting", "Accounting service (DB): COA, journal entries, trial balance", "Backend Engineer", "Senior", 22, "accounting_services.py, routers/accounting.py", "Built"),
    ("Accounting & Reporting", "Reporting/dashboards service (DB): financial/audit/executive", "Backend Engineer", "Senior", 18, "reporting_services.py, routers/{reports,dashboards}.py", "Built"),
    ("Accounting & Reporting", "Reports UI", "Frontend Engineer", "Mid", 8, "src/app/reports/page.tsx", "Built"),
    # CRM
    ("CRM", "CRM service (DB): contacts, deals, activities, summary", "Backend Engineer", "Senior", 16, "crm_services.py, routers/crm.py", "Built"),
    # Market data
    ("Market Data", "Provider adapters + cache/failover (CoinGecko/Yahoo/BLS/FRED/TwelveData/Sharadar)", "Integration Engineer", "Senior", 30, "backend/app/market_services.py", "Built"),
    ("Market Data", "Market API + models", "Backend Engineer", "Mid", 8, "routers/market.py, market_models.py", "Built"),
    ("Market Data", "LiveMarket polling component + dashboard wiring", "Frontend Engineer", "Senior", 12, "src/components/live-market.tsx, src/app/dashboard/page.tsx", "Built"),
    # Research & score
    ("Research & Opportunity Score", "Multi-factor Opportunity Score engine", "Data Scientist / Quant", "Senior", 28, "backend/app/opportunity_score.py", "Built"),
    ("Research & Opportunity Score", "Validation studies: backtests, adoption, acquisition, coverage", "Data Scientist / Quant", "Senior", 30, "backend/app/research_services.py", "Built"),
    ("Research & Opportunity Score", "Research API + models", "Backend Engineer", "Mid", 10, "routers/research.py, research_models.py", "Built"),
    ("Research & Opportunity Score", "Opportunities + assistant UI", "Frontend Engineer", "Mid", 12, "src/app/{opportunities,assistant}/page.tsx", "Built"),
    # Valuations
    ("Valuations", "Modeled real-time valuation engine (live inputs)", "Data Scientist / Quant", "Senior", 16, "backend/app/valuation_services.py", "Built"),
    ("Valuations", "Valuation API + models", "Backend Engineer", "Mid", 6, "routers/valuations.py, valuation_models.py", "Built"),
    ("Valuations", "Portfolio + due-diligence UI", "Frontend Engineer", "Mid", 12, "src/app/{portfolio,due-diligence}/page.tsx", "Built"),
    # Support & agents
    ("Support & AI Agents", "Support FAQ retrieval engine", "AI Engineer", "Senior", 14, "backend/app/support_services.py", "Built"),
    ("Support & AI Agents", "Five AI agents: routing, escalation, tickets", "AI Engineer", "Senior", 22, "backend/app/agents_services.py", "Built"),
    ("Support & AI Agents", "Support + team UI", "Frontend Engineer", "Mid", 12, "src/app/{support,team}/page.tsx", "Built"),
    # Growth / leads
    ("Growth / Leads", "Leads API: capture, dedupe, count + models", "Backend Engineer", "Mid", 8, "routers/leads.py, lead_models.py", "Built"),
    ("Growth / Leads", "Waitlist form + landing/join UI", "Frontend Engineer", "Mid", 12, "src/components/waitlist-form.tsx, src/app/{join,page}.tsx", "Built"),
    # Integrations
    ("External Integrations", "Connectors, sync jobs, banking/vendor/Office endpoints", "Integration Engineer", "Senior", 24, "routers/integrations.py, services.py", "Prototype"),
    # Frontend shell & brand
    ("Frontend shell", "App shell/nav/layout + logo + brand CSS system", "Frontend Engineer", "Senior", 20, "src/components/{platform-shell,logo}.tsx, src/app/{layout,globals.css}", "Built"),
    ("Frontend shell", "UX & behavioral-science color/brand system design", "UX / UI Designer", "Senior", 12, "docs/BRAND_AND_COLOR_SYSTEM.md, globals.css", "Built"),
    ("Frontend shell", "Shared client data module", "Frontend Engineer", "Mid", 6, "src/lib/platform-data.ts", "Built"),
    # QA
    ("Quality Assurance", "Backend automated test suite (15 files, 110+ tests)", "QA Engineer", "Senior", 40, "backend/tests/*.py", "Built"),
    # DevOps
    ("Build & Deploy", "Dockerfiles + compose + hardening + health probes", "DevOps / Cloud Engineer", "Senior", 16, "Dockerfile, backend/Dockerfile, docker-compose.yml", "Built"),
    ("Build & Deploy", "Production config validation + secrets wiring", "DevOps / Cloud Engineer", "Senior", 6, "config.py, .dockerignore", "Built"),
    # Architecture
    ("Architecture", "System architecture, integration & code review", "Tech Lead / Architect", "Senior", 20, "whole platform", "Built"),
]

# --- styling helpers --------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="0B2A4A")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="0B2A4A")
SUB_FONT = Font(italic=True, size=10, color="555555")
TOTAL_FILL = PatternFill("solid", fgColor="D4AF37")
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CURRENCY = '"$"#,##0'


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def build() -> None:
    wb = Workbook()

    # ============ Sheet 1: Disclosure ============
    ws = wb.active
    ws.title = "Disclosure"
    ws["A1"] = "John Henry Investments — Coding Disclosure by Component (Subsystem)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Signature: {SIGNATURE}  |  Realistic equivalent professional value of all coding performed  |  Rates = mid-market"
    ws["A2"].font = SUB_FONT

    headers = ["Subsystem", "Work item / action", "Job title (skillset)", "Level",
               "Hourly rate (mid, $)", "Est. hours", "Value ($)", "Key files", "Status"]
    hrow = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    _style_header(ws, hrow, len(headers))
    ws.freeze_panes = f"A{hrow + 1}"

    r = hrow + 1
    for (sub, action, title, level, hours, files, status) in ROWS:
        mid = RATES[title][2]
        ws.cell(row=r, column=1, value=sub)
        ws.cell(row=r, column=2, value=action)
        ws.cell(row=r, column=3, value=title)
        ws.cell(row=r, column=4, value=level)
        ws.cell(row=r, column=5, value=mid).number_format = CURRENCY
        ws.cell(row=r, column=6, value=hours)
        vcell = ws.cell(row=r, column=7, value=f"=E{r}*F{r}")
        vcell.number_format = CURRENCY
        ws.cell(row=r, column=8, value=files)
        ws.cell(row=r, column=9, value=status)
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP
        r += 1

    # totals
    last = r - 1
    ws.cell(row=r, column=4, value="TOTAL")
    ws.cell(row=r, column=6, value=f"=SUM(F{hrow + 1}:F{last})")
    tv = ws.cell(row=r, column=7, value=f"=SUM(G{hrow + 1}:G{last})")
    tv.number_format = CURRENCY
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER

    widths = [26, 52, 24, 12, 16, 11, 14, 46, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ============ Sheet 2: Rate Card ============
    rc = wb.create_sheet("Rate Card")
    rc["A1"] = "Realistic Hourly Rate Card by Job Title"
    rc["A1"].font = TITLE_FONT
    rc["A2"] = RATE_SOURCE
    rc["A2"].font = SUB_FONT
    rc_headers = ["Job title (skillset)", "Typical level", "Low ($/hr)", "Mid ($/hr)", "High ($/hr)"]
    for i, h in enumerate(rc_headers, start=1):
        rc.cell(row=4, column=i, value=h)
    _style_header(rc, 4, len(rc_headers))
    rr = 5
    for title, (level, low, mid, high) in RATES.items():
        rc.cell(row=rr, column=1, value=title)
        rc.cell(row=rr, column=2, value=level)
        rc.cell(row=rr, column=3, value=low).number_format = CURRENCY
        rc.cell(row=rr, column=4, value=mid).number_format = CURRENCY
        rc.cell(row=rr, column=5, value=high).number_format = CURRENCY
        for c in range(1, 6):
            rc.cell(row=rr, column=c).border = BORDER
        rr += 1
    for i, w in enumerate([26, 16, 14, 14, 14], start=1):
        rc.column_dimensions[get_column_letter(i)].width = w

    # ============ Sheet 3: Subsystem Summary ============
    ss = wb.create_sheet("Subsystem Summary")
    ss["A1"] = "Summary by Subsystem"
    ss["A1"].font = TITLE_FONT
    ss_headers = ["Subsystem", "Est. hours", "Value ($)"]
    for i, h in enumerate(ss_headers, start=1):
        ss.cell(row=3, column=i, value=h)
    _style_header(ss, 3, len(ss_headers))
    subs: list[str] = []
    for (sub, *_rest) in ROWS:
        if sub not in subs:
            subs.append(sub)
    sr = 4
    drange_sub = f"Disclosure!$A${hrow + 1}:$A${last}"
    drange_hrs = f"Disclosure!$F${hrow + 1}:$F${last}"
    drange_val = f"Disclosure!$G${hrow + 1}:$G${last}"
    for sub in subs:
        ss.cell(row=sr, column=1, value=sub)
        ss.cell(row=sr, column=2, value=f'=SUMIF({drange_sub},A{sr},{drange_hrs})')
        ss.cell(row=sr, column=3, value=f'=SUMIF({drange_sub},A{sr},{drange_val})').number_format = CURRENCY
        for c in range(1, 4):
            ss.cell(row=sr, column=c).border = BORDER
        sr += 1
    ss.cell(row=sr, column=1, value="TOTAL")
    ss.cell(row=sr, column=2, value=f"=SUM(B4:B{sr - 1})")
    ss.cell(row=sr, column=3, value=f"=SUM(C4:C{sr - 1})").number_format = CURRENCY
    for c in range(1, 4):
        cell = ss.cell(row=sr, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
    for i, w in enumerate([30, 14, 16], start=1):
        ss.column_dimensions[get_column_letter(i)].width = w

    # ============ Sheet 4: Notes / Disclosure ============
    nt = wb.create_sheet("Notes & Disclosure")
    nt["A1"] = "Methodology & Full Disclosure"
    nt["A1"].font = TITLE_FONT
    notes = [
        "",
        "Purpose: Full disclosure of all coding performed on the JHI platform, organized by",
        "component (subsystem), mapped to the skillset / job title that would perform the work,",
        "with a realistic hourly value per action and the estimated equivalent professional value.",
        "",
        "Methodology:",
        "- 'Est. hours' = good-faith estimate of equivalent professional effort to build the",
        "  delivered code (not stopwatch time). The platform was built founder-led with AI",
        "  assistance, which compresses calendar time; this workbook expresses the market-rate",
        "  value of the equivalent professional work product.",
        "- 'Hourly rate (mid)' = mid-market US contractor rate for that job title (see Rate Card).",
        "- 'Value' = rate x hours (computed by formula).",
        "- Job titles map to the skillset most appropriate to each action (see",
        "  docs/JOB_DESCRIPTIONS_AND_STAFFING_REQUIREMENTS.md).",
        "",
        "Rate anchors: docs/COMPENSATION_AND_PRO_SERVICES_PROJECTIONS.md",
        "  (e.g., DevOps $100-$250/hr; AI/data $150-$350/hr; counsel $300-$1,200/hr).",
        "",
        "Scope: application source code across all subsystems plus tests and DevOps. Items marked",
        "'(PR #17)' are the security-hardening additions; 'Prototype' = partial/not production-grade.",
        "",
        "DISCLAIMER: This is a planning / disclosure artifact. It is NOT an invoice, an audited",
        "record, a payroll document, or tax/accounting/legal advice. Actual cost-to-build or",
        "fair value depends on geography, hiring market, scope, and engagement model. Confirm",
        "with a CPA/controller before using for capitalization, valuation, or tax purposes.",
        "",
        f"Founder signature: {SIGNATURE}  |  John Henry Investments (proprietary)",
    ]
    for i, line in enumerate(notes, start=2):
        nt.cell(row=i, column=1, value=line)
    nt.column_dimensions["A"].width = 100

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
