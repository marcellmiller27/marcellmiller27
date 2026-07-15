"""Generate the JHI Data-Sources Comparison workbook.

A board/ops reference comparing JHI's market & economic data sources across
coverage, cost, and — critically — redistribution rights to our subscriber base.

Run:  /workspace/.venv/bin/python scripts/data_sources_comparison_xlsx.py
Output: public/downloads/JHI_Data_Sources_Comparison.xlsx

Notes are internal planning estimates; commercial terms should be confirmed with
each vendor and licensing details with counsel. Public-source redistribution is
generally permitted with attribution (confirm per-dataset for IMF/OECD/World Bank).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "0C1F33"
GOLD = "9A6B12"
LIGHT = "EAEFF6"
GREEN = "0F9D63"
RED = "B23B3B"
AMBER = "9A6B12"
WHITE = "FFFFFF"

HEADER_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=15)
SUB_FONT = Font(italic=True, color="5A6B7D", size=9)
CELL_FONT = Font(size=10, color="0C1F33")
BOLD = Font(bold=True, size=10, color="0C1F33")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="C9D3DF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _write_table(ws, headers, rows, widths, start_row: int) -> None:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, start_row, len(headers))
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if c == 1:
                cell.font = BOLD
        if (r - start_row) % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ALT_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start_row + 1, column=2)


def _title(ws, title: str, subtitle: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    ws.row_dimensions[1].height = 22


# --------------------------------------------------------------------------- #
# Sheet 1 — Source comparison
# --------------------------------------------------------------------------- #
SRC_HEADERS = [
    "Source", "Type", "Key data / coverage", "Geographic scope", "Cost",
    "Redistribute to subscribers?", "API & format", "Update cadence",
    "Best JHI use", "Caveats",
]
SRC_WIDTHS = [22, 20, 34, 18, 20, 30, 24, 16, 34, 30]
SRC_ROWS = [
    [
        "Nasdaq Data Link (Sharadar SF1)", "Commercial vendor",
        "Point-in-time US equity fundamentals (SF1): as-reported financials, ratios; ~14k+ tickers, 20+ yrs",
        "US equities", "$18,000/yr (SF1 + platform); up to 1,000 users + overage",
        "YES — external DISTRIBUTION of Derived Data to subscribers (per Additional Terms); per-user, capped",
        "REST datatables (JSON); key NASDAQ_DATA_LINK_API_KEY", "Daily (as filed)",
        "Opportunity Score fundamentals, Deal X-Ray/QoE comps, valuation, screeners; powers H5",
        "Per-user cap + overage; third-party Sharadar terms; delete-on-termination",
    ],
    [
        "Twelve Data (Venture)", "Commercial vendor",
        "Real-time & historical prices: US equities/ETFs, FX, crypto, commodities, indices; technicals; some fundamentals",
        "US + global (Venture: US/FX/crypto/commodities)", "$499/mo (~$5k/yr; ~$414/mo annual). Flat",
        "DISPLAY only (in-app) + UNRESTRICTED derived; NO raw redistribution via own API/feed",
        "REST + WebSocket (JSON); key TWELVEDATA_API_KEY", "Real-time / intraday / EOD",
        "Live ticker & dashboards, cross-asset analytics, derived scores, real-time context",
        "No client raw-data API; attribution; flat (not per-user)",
    ],
    [
        "FRED (St. Louis Fed)", "Public (govt-adjacent)",
        "~800k US macro series: GDP, CPI, rates, M2, credit, delinquencies, retail, sentiment, housing, yields (aggregates BLS/BEA/Census/Fed/Treasury)",
        "US (some intl)", "FREE (free API key)",
        "YES — free to display & redistribute (attribution; few copyrighted series excepted)",
        "REST (JSON); key FRED_API_KEY (active)", "Daily–quarterly",
        "Macro dashboards, newsletters, context; RAW exportable in Excel (public)",
        "Attribution; a few series carry source copyright",
    ],
    [
        "SEC EDGAR", "Public (US SEC)",
        "Filings: 10-K/10-Q/8-K/S-1/proxy, insider (3/4/5), 13F; XBRL financial-statement datasets; full-text",
        "US public cos (+ foreign filers)", "FREE",
        "YES — US govt public domain; freely usable & redistributable",
        "REST (data.sec.gov, XBRL frames); full-text; bulk", "Real-time as filed",
        "Deal X-Ray/QoE financials from filings, fundamentals cross-check, insider/13F signals, red-flag scans",
        "10 req/s fair-access limit; declared User-Agent required",
    ],
    [
        "Bureau of Labor Statistics (BLS)", "Public (US DOL)",
        "CPI, PPI, employment (payrolls, JOLTS), wages, productivity, unemployment",
        "US", "FREE (optional key for higher limits)",
        "YES — public domain (attribution)", "REST v2 (JSON)", "Monthly / quarterly",
        "Inflation (CPI — in use), labor-market context, newsletters",
        "Rate limits without a key",
    ],
    [
        "BEA (Bureau of Economic Analysis)", "Public (US DOC)",
        "GDP (detailed), personal income, PCE, trade, industry & regional accounts",
        "US (national/regional/intl transactions)", "FREE (free key)",
        "YES — public domain (attribution)", "REST (JSON/XML)", "Monthly / quarterly",
        "GDP detail, PCE inflation, personal income for macro analytics/newsletters",
        "API key; frequent revisions",
    ],
    [
        "U.S. Treasury (Fiscal Data)", "Public (US Treasury)",
        "Federal debt & deficit, daily Treasury yield curve, auctions, interest rates, fiscal statements",
        "US", "FREE (no key)",
        "YES — public domain", "REST (fiscaldata.treasury.gov, JSON)", "Daily / monthly",
        "Yield curve, risk-free rate for valuation, debt/deficit context, newsletters",
        "None major",
    ],
    [
        "Federal Reserve (Board of Governors)", "Public (US Fed)",
        "Rates (H.15), bank assets/liabilities (H.8), flow of funds (Z.1), consumer credit (G.19), FX (H.10), industrial production",
        "US", "FREE",
        "YES — public domain", "Fed DDP (Data Download); much also via FRED", "Daily–quarterly",
        "Rates, bank health, liquidity — often simplest via FRED",
        "Overlaps FRED; DDP format",
    ],
    [
        "IMF", "Public (international org)",
        "World Economic Outlook, Intl Financial Statistics (IFS), Balance of Payments, Direction of Trade, govt finance, FX",
        "Global (190+ countries)", "FREE",
        "Generally YES with attribution (confirm per dataset)", "SDMX / REST (IMF Data Portal)", "Monthly–annual",
        "Global macro, country risk, FX for global/enterprise clients, newsletters",
        "SDMX complexity; per-dataset terms",
    ],
    [
        "OECD", "Public (international org)",
        "Composite leading indicators (CLI), PMIs, labor, trade, productivity, tax, inequality",
        "Global (OECD + partners)", "FREE",
        "Mostly YES with attribution (much CC-BY; some restricted)", "SDMX / REST (OECD Data API)", "Monthly–annual",
        "Global leading indicators, cross-country macro, newsletters",
        "SDMX complexity; check per-dataset terms",
    ],
    [
        "World Bank", "Public (international org)",
        "World Development Indicators (~1,400): GDP, population, poverty, trade, finance, governance",
        "Global (~200 countries)", "FREE",
        "Mostly CC-BY 4.0 (open, attribution); some restricted", "REST (JSON)", "Mostly annual",
        "Long-run global/country context, emerging markets, newsletters",
        "Annual lag; some indicators restricted",
    ],
]

# --------------------------------------------------------------------------- #
# Sheet 2 — FRED datasets
# --------------------------------------------------------------------------- #
FRED_HEADERS = ["Series ID", "Name", "Category", "Cadence", "JHI use", "In platform now?"]
FRED_WIDTHS = [16, 34, 16, 12, 34, 16]
FRED_ROWS = [
    ["GDP", "US Gross Domestic Product", "Output", "Quarterly", "Macro dashboard / newsletter", "Yes"],
    ["UNRATE", "Unemployment Rate", "Labor", "Monthly", "Macro read", "Yes"],
    ["FEDFUNDS", "Federal Funds Rate", "Rates", "Monthly", "Rates / policy", "Yes"],
    ["M2SL", "M2 Money Supply", "Money", "Monthly", "Liquidity", "Yes"],
    ["TOTALSL", "Consumer Credit (total)", "Credit", "Monthly", "Consumer debt", "Yes"],
    ["REVOLSL", "Revolving Consumer Credit", "Credit", "Monthly", "Credit-card debt", "Yes"],
    ["HDTGPDUSQ163N", "Household Debt to GDP", "Credit", "Quarterly", "Leverage", "Yes"],
    ["DRCCLACBS", "Credit Card Delinquency Rate", "Credit risk", "Quarterly", "Delinquency signal", "Yes"],
    ["DRALACBN", "All Bank Loans Delinquency Rate", "Credit risk", "Quarterly", "Delinquency signal", "Yes"],
    ["DRSFRMACBS", "Mortgage Delinquency Rate", "Credit risk", "Quarterly", "Delinquency signal", "Yes"],
    ["RSAFS", "Retail Sales", "Consumer", "Monthly", "Demand", "Yes"],
    ["UMCSENT", "Consumer Sentiment (UMich)", "Sentiment", "Monthly", "Sentiment", "Yes"],
    ["INDPRO", "Industrial Production", "Output", "Monthly", "Production", "Yes"],
    ["CPIAUCSL", "CPI (All Urban, SA)", "Inflation", "Monthly", "Inflation (FRED alt to BLS)", "Via BLS"],
    ["DGS10", "10-Yr Treasury Yield", "Rates", "Daily", "Risk-free / curve", "Via Yahoo"],
    ["PCEPI", "PCE Price Index", "Inflation", "Monthly", "Fed's preferred inflation", "Candidate"],
    ["HOUST", "Housing Starts", "Housing", "Monthly", "Housing cycle", "Candidate"],
    ["CSUSHPINSA", "Case-Shiller Home Price Index", "Housing", "Monthly", "Home prices", "Candidate"],
]

# --------------------------------------------------------------------------- #
# Sheet 3 — Redistribution rights matrix (the key licensing view)
# --------------------------------------------------------------------------- #
REDIST_HEADERS = [
    "Source", "Serve DERIVED to subscribers?", "DISPLAY raw in-app?",
    "Export / redistribute RAW to clients?", "Cost basis", "Notes",
]
REDIST_WIDTHS = [30, 26, 20, 30, 16, 34]
REDIST_ROWS = [
    ["Nasdaq Data Link (SF1)", "YES (licensed distributor)", "YES (licensed)",
     "Within distributor terms (per-user cap)", "Per external user", "Formal external-distribution license (Additional Terms)"],
    ["Twelve Data (Venture)", "YES (unrestricted derived)", "YES", "NO (no raw API/feed to clients)",
     "Flat", "Display-only for raw; upgrade to Enterprise/distribution for raw feeds"],
    ["FRED", "YES", "YES", "YES (public)", "Free", "Attribution; few copyrighted series excepted"],
    ["SEC EDGAR", "YES", "YES", "YES (public domain)", "Free", "User-Agent + 10 req/s limit"],
    ["BLS", "YES", "YES", "YES (public)", "Free", "Attribution"],
    ["BEA", "YES", "YES", "YES (public)", "Free", "Attribution"],
    ["U.S. Treasury", "YES", "YES", "YES (public)", "Free", "Public domain"],
    ["Federal Reserve", "YES", "YES", "YES (public)", "Free", "Overlaps FRED"],
    ["IMF", "YES", "YES", "Mostly (attribution)", "Free", "Confirm per-dataset terms"],
    ["OECD", "YES", "YES", "Mostly (CC-BY)", "Free", "Confirm per-dataset terms"],
    ["World Bank", "YES", "YES", "Mostly (CC-BY 4.0)", "Free", "Attribution; some restricted"],
]


def main() -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Source Comparison"
    _title(ws1, "JHI Data-Sources Comparison",
           "Market & economic data sources by coverage, cost, and redistribution rights — internal planning reference.",
           len(SRC_HEADERS))
    _write_table(ws1, SRC_HEADERS, SRC_ROWS, SRC_WIDTHS, start_row=4)

    ws2 = wb.create_sheet("FRED Datasets")
    _title(ws2, "FRED Datasets (key series)",
           "Public, free, broadly redistributable US macro series (attribution).", len(FRED_HEADERS))
    _write_table(ws2, FRED_HEADERS, FRED_ROWS, FRED_WIDTHS, start_row=4)

    ws3 = wb.create_sheet("Redistribution Rights")
    _title(ws3, "Redistribution Rights to Subscribers",
           "The decisive licensing view: commercial vendors are restricted; public sources are broadly redistributable.",
           len(REDIST_HEADERS))
    _write_table(ws3, REDIST_HEADERS, REDIST_ROWS, REDIST_WIDTHS, start_row=4)

    out = Path("public/downloads/JHI_Data_Sources_Comparison.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}  ({out.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
