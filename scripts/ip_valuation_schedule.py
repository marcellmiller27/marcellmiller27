# JHI-SIG: 69M2705M | IP Valuation Schedule | John Henry Investments (proprietary)
"""Generate the interactive IP Valuation Schedule (Excel).

Supports the value of the platform/mobile IP contributed to JHI Research &
Analytics Firm, Inc. in exchange for 10,000,000 founder shares. Values completed
engineering + professional work plus projected work-to-launch by
workstream x skillset x market rate x hours, and reconciles to the ~$400,000 IP
contribution value (posted as $1,000 Common Stock at $0.0001 par + $399,000 APIC).

NOT a formal appraisal. Par value != fair value — the official Section 351 basis
is the CPA/appraiser's determination; this schedule is defensible support.

Run:  python scripts/ip_valuation_schedule.py [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SIG = "JHI-SIG: 69M2705M"
ENTITY = "JHI Research & Analytics Firm, Inc."
_ENTITY_HF = ENTITY.replace("&", "&&")

_NAVY_FILL = PatternFill("solid", fgColor="0C1F33")
_SUB_FILL = PatternFill("solid", fgColor="EAEFF6")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")
_TOTAL_FILL = PatternFill("solid", fgColor="E6EEFA")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)
_NAVY = "0C1F33"
_MONEY = '"$"#,##0'
_RATE = '"$"#,##0"/hr"'

# (workstream, category, skillset, rate, hours)
COMPLETED = [
    ("Backend platform & 17 API modules", "Completed", "Sr. Backend", 140, 500),
    ("Frontend + mobile app (Next.js/React)", "Completed", "Sr. Frontend", 130, 360),
    ("Research / quant (Opportunity Score, validation, fundamentals)", "Completed", "Data/Quant", 180, 210),
    ("Deal X-Ray / QoE / financial-diligence engines", "Completed", "Sr. Backend + domain", 160, 230),
    ("Interactive Excel + PDF export engines", "Completed", "Sr. Backend", 140, 95),
    ("Deal Pipeline + Postgres persistence", "Completed", "Sr. Backend", 140, 85),
    ("DevOps / Docker / infrastructure", "Completed", "DevOps", 150, 80),
    ("Security (auth, 2FA, WebAuthn, encryption)", "Completed", "Security", 160, 95),
    ("Architecture / product / competitive strategy", "Completed", "Product/Strategy", 175, 175),
    ("Finance & legal-adjacent docs (valuation, NASDAQ, posture, board)", "Completed", "Finance/Legal-adj.", 150, 160),
    ("QA / automated + manual testing", "Completed", "QA", 110, 130),
]
PROJECTED = [
    ("Homepage Phase B + About polish", "Projected", "Sr. Frontend", 130, 55),
    ("Stripe checkout + trial/paywall + Stripe Tax", "Projected", "Sr. Backend", 140, 90),
    ("CI/CD + observability + backups", "Projected", "DevOps", 150, 75),
    ("Production security hardening (RBAC, secrets, headers)", "Projected", "Security", 160, 65),
    ("SF1 fundamentals integration + H5 validation", "Projected", "Data/Quant", 180, 110),
    ("Module UI wiring (accounting / reports / CRM)", "Projected", "Sr. Frontend", 130, 80),
    ("Terms / Privacy / compliance + counsel coordination", "Projected", "Legal-adj.", 150, 35),
    ("E2E testing + launch hardening", "Projected", "QA", 110, 90),
]

SHARES = 10_000_000
PAR = 0.0001  # -> $1,000 Common Stock (par)
CONTRIBUTION = 400_000  # IP contribution value -> Common Stock + APIC


def _watermark(ws: Worksheet) -> None:
    ws.oddFooter.left.text = f"&8© {_ENTITY_HF}  ·  {SIG}"
    ws.oddFooter.center.text = "&8Confidential — not a formal appraisal"
    ws.oddFooter.right.text = "&8Page &P of &N"


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Valuation Schedule"
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 16

    ws.merge_cells("A1:F1")
    ws["A1"] = ENTITY
    ws["A1"].font = _WHITE
    ws["A1"].fill = _NAVY_FILL
    ws.merge_cells("A2:F2")
    ws["A2"] = "IP Valuation Schedule — platform & mobile software (contributed for 10,000,000 shares)"
    ws["A2"].font = Font(bold=True, size=13, color=_NAVY)
    ws.merge_cells("A3:F3")
    ws["A3"] = f"generated {date.today().isoformat()}  ·  {SIG}  ·  edit the gold Rate/Hours cells"
    ws["A3"].font = _MUTED

    headers = ["Workstream", "Category", "Skillset", "Rate", "Hours", "Value"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL

    row = 6

    def _section(title: str, rows: list) -> tuple[int, int]:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sc = ws.cell(row=row, column=1, value=title)
        sc.font = _BOLD
        sc.fill = _SUB_FILL
        row += 1
        first = row
        for ws_name, cat, skill, rate, hours in rows:
            ws.cell(row=row, column=1, value=ws_name)
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=skill)
            rc = ws.cell(row=row, column=4, value=rate)
            rc.fill = _INPUT_FILL
            rc.number_format = _RATE
            hc = ws.cell(row=row, column=5, value=hours)
            hc.fill = _INPUT_FILL
            vc = ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
            vc.number_format = _MONEY
            row += 1
        return first, row - 1

    c_first, c_last = _section("COMPLETED WORK (to date)", COMPLETED)
    ws.cell(row=row, column=1, value="Completed subtotal").font = _BOLD
    ws.cell(row=row, column=5, value=f"=SUM(E{c_first}:E{c_last})").font = _BOLD
    ws.cell(row=row, column=6, value=f"=SUM(F{c_first}:F{c_last})").number_format = _MONEY
    ws.cell(row=row, column=6).font = _BOLD
    for cc in range(1, 7):
        ws.cell(row=row, column=cc).fill = _TOTAL_FILL
    row += 2

    p_first, p_last = _section("PROJECTED WORK (to launch)", PROJECTED)
    ws.cell(row=row, column=1, value="Projected subtotal").font = _BOLD
    ws.cell(row=row, column=5, value=f"=SUM(E{p_first}:E{p_last})").font = _BOLD
    ws.cell(row=row, column=6, value=f"=SUM(F{p_first}:F{p_last})").number_format = _MONEY
    ws.cell(row=row, column=6).font = _BOLD
    for cc in range(1, 7):
        ws.cell(row=row, column=cc).fill = _TOTAL_FILL
    row += 2

    grand = row
    ws.cell(row=grand, column=1, value="GRAND TOTAL (completed + projected)").font = Font(bold=True, size=12, color=_NAVY)
    ws.cell(row=grand, column=5, value=f"=SUM(E{c_first}:E{c_last})+SUM(E{p_first}:E{p_last})").font = _BOLD
    ws.cell(row=grand, column=6, value=f"=SUM(F{c_first}:F{c_last})+SUM(F{p_first}:F{p_last})").number_format = _MONEY
    ws.cell(row=grand, column=6).font = Font(bold=True, size=12, color=_NAVY)
    for cc in range(1, 7):
        ws.cell(row=grand, column=cc).fill = _TOTAL_FILL
    row += 2

    # Equity / contribution reconciliation
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="EQUITY / CONTRIBUTION RECONCILIATION").font = _BOLD
    ws.cell(row=row, column=1).fill = _SUB_FILL
    row += 1
    r0 = row
    ws.cell(row=r0, column=1, value="Shares issued").font = _BOLD
    ws.cell(row=r0, column=6, value=SHARES).number_format = "#,##0"
    ws.cell(row=r0 + 1, column=1, value="Par value per share").font = _BOLD
    ws.cell(row=r0 + 1, column=6, value=PAR).number_format = '"$"0.0000'
    ws.cell(row=r0 + 2, column=1, value="Common Stock (shares × par)").font = _BOLD
    ws.cell(row=r0 + 2, column=6, value=f"=F{r0}*F{r0 + 1}").number_format = _MONEY
    ws.cell(row=r0 + 3, column=1, value="IP contribution value (target)").font = _BOLD
    ic = ws.cell(row=r0 + 3, column=6, value=CONTRIBUTION)
    ic.number_format = _MONEY
    ic.fill = _INPUT_FILL
    ws.cell(row=r0 + 4, column=1, value="Additional Paid-In Capital (contribution − common stock)").font = _BOLD
    ws.cell(row=r0 + 4, column=6, value=f"=F{r0 + 3}-F{r0 + 2}").number_format = _MONEY
    ws.cell(row=r0 + 5, column=1, value="Schedule grand total (IP support)").font = _BOLD
    ws.cell(row=r0 + 5, column=6, value=f"=F{grand}").number_format = _MONEY
    ws.cell(row=r0 + 6, column=1, value="Variance (grand total − contribution)").font = _BOLD
    ws.cell(row=r0 + 6, column=6, value=f"=F{grand}-F{r0 + 3}").number_format = _MONEY
    row = r0 + 8

    for note in [
        "Par value $0.0001 × 10,000,000 = $1,000 Common Stock; the ~$400,000 IP contribution posts",
        "as $1,000 Common Stock + $399,000 APIC. Par value does not change as company value grows.",
        "Under U.S. GAAP, internally generated increases in enterprise value are not recognized.",
        "NOT a formal appraisal — the official IRC §351 fair-value basis is the determination",
        "of a licensed CPA / qualified appraiser. Rates and hours are editable inputs.",
        f"© {date.today().year} {ENTITY}. Confidential. {SIG}.",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=note)
        c.font = _MUTED
        c.alignment = Alignment(wrap_text=True)
        row += 1

    _watermark(ws)
    return wb


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "JHI_IP_Valuation_Schedule.xlsx"
    build().save(out)
    print(f"wrote {out}")
