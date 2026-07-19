# JHI-SIG: 69M2705M | 5-Year Consolidated Projections | JHI Research & Analytics Firm, Inc. (proprietary)
"""Generate the 5-year consolidated, audited-realistic projections workbook.

Fuses our prior forecasts (prepaid-MSA sales commission, per-seat unit economics, cloud/
system cost curve, staged staffing) into ONE model:
  - Sales scaling 1 -> 30 reps toward the 100/mo (1,200/yr) goal.
  - Prepaid annual MSAs: 15% upfront commission + year-end MSA-completion bonus.
  - Cloud/system COGS that scale with active subscribers.
  - EBITDA-gated staged staffing (admin first; senior SWE when EBITDA supports the salary).
  - 5-year detail, MONTHLY, one spreadsheet per year (Y1..Y5), + summary + dashboard.

All inputs live on the Assumptions sheet (this is a working model — tweak & regenerate).

Run:  python scripts/consolidated_projections_model.py [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SIG = "JHI-SIG: 69M2705M"
ENTITY = "JHI Research & Analytics Firm, Inc."
_ENTITY_HF = ENTITY.replace("&", "&&")
_NAVY_FILL = PatternFill("solid", fgColor="0C1F33")
_SUB_FILL = PatternFill("solid", fgColor="EAEFF6")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")
_HILITE = PatternFill("solid", fgColor="E4F3E9")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)
_MONEY = '"$"#,##0'
_PCT = "0.0%"
_NAVY = "0C1F33"

# ── Assumptions (defaults; editable in the workbook / here) ──────────────────
A = {
    "reps": 30,                 # sales force from the start (target/ideal)
    "closes_per_rep_mo": 3.5,   # premium closes / rep / month at full ramp (30*3.5=105/mo)
    "ramp": [0.25, 0.5, 0.75],  # new-product productivity ramp, months 1-3 (then 1.0)
    "t1_price": 1500.0, "t2_price": 299.0, "t1_mix": 0.10,  # rep-sold premium (T1/T2)
    "prepay_disc": 0.10,        # annual prepay discount
    "commission": 0.15,         # upfront, % of net booking
    "msa_bonus": 0.05,          # year-end completion bonus, % of completed-contract value
    "annual_renewal": 0.85,     # prepaid annual renewal rate
    "processing": 0.03,         # payment processing % of revenue
    "cloud_per_user_mo": 1.50,  # AWS/GCP per active user / month
    "data_annual": 18000.0,     # NASDAQ/SF1 flat to cap
    "data_cap_users": 1000,
    "data_overage_user_mo": 3.0,
    "marketing_mo": 8000.0,
    "legal_mo": 1500.0,
    "software_mo": 1500.0,
    "founder_annual": 1.0,      # symbolic $1/yr (+ discretionary, not modeled)
    "months": 60,
}
# Staged staffing: (role, annual salary, EBITDA gate $/mo) — hired in priority order when
# monthly EBITDA (with roles already hired) crosses the gate.
STAFF = [
    ("Admin / Ops Coordinator", 55000, 20000),
    ("Customer Success Lead", 75000, 80000),
    ("Marketing Manager", 95000, 150000),
    ("Senior Software Engineer", 175000, 220000),
    ("Controller (FT)", 110000, 320000),
    ("Second Senior Software Engineer", 170000, 450000),
]


def compute(a=A):
    avg_annual = ((a["t1_price"] * a["t1_mix"]) + (a["t2_price"] * (1 - a["t1_mix"]))) * 12 * (1 - a["prepay_disc"])
    r = a["annual_renewal"]
    M = a["months"]
    # 1) New acquisitions (rep-driven, with new-product ramp)
    new = [a["reps"] * a["closes_per_rep_mo"] * (a["ramp"][m] if m < len(a["ramp"]) else 1.0) for m in range(M)]
    # 2) Cohort survival (annual renewal) → active base; renewal payments (annual re-billing)
    active = [0.0] * M; renew_cnt = [0.0] * M
    for m in range(M):
        tot = 0.0; rc = 0.0
        for k in range(m + 1):
            yrs = (m - k) // 12
            tot += new[k] * (r ** yrs)                       # survivors still active
            if (m - k) > 0 and (m - k) % 12 == 0:
                rc += new[k] * (r ** yrs)                    # this cohort re-bills this month
        active[m] = tot; renew_cnt[m] = rc
    rec = [active[m] * avg_annual / 12 for m in range(M)]     # recognized ratably (accrual)
    cash = [(new[m] + renew_cnt[m]) * avg_annual for m in range(M)]  # prepaid cash: new + renewals
    # 3) Cost / EBITDA / cash-flow / balance-sheet loop
    cogs = [0.0] * M; gp = [0.0] * M; comm = [0.0] * M; bonus = [0.0] * M
    mktg = [0.0] * M; fixed = [0.0] * M; staff = [0.0] * M; ebitda = [0.0] * M
    headcount = [0] * M
    cash_out = [0.0] * M; net_cash = [0.0] * M; cum_cash = [0.0] * M
    deferred = [0.0] * M; retained = [0.0] * M
    hired = []; hire_month = {}
    for m in range(M):
        proc = rec[m] * a["processing"]
        cloud = active[m] * a["cloud_per_user_mo"]
        data = a["data_annual"] / 12 + max(0.0, active[m] - a["data_cap_users"]) * a["data_overage_user_mo"]
        cogs[m] = proc + cloud + data
        gp[m] = rec[m] - cogs[m]
        comm[m] = new[m] * avg_annual * a["commission"]  # 15% upfront on NEW acquisition bookings
        bonus[m] = (new[m - 12] * r * avg_annual * a["msa_bonus"]) if m >= 12 else 0.0
        mktg[m] = a["marketing_mo"]
        fixed[m] = a["legal_mo"] + a["software_mo"] + a["founder_annual"] / 12
        staff[m] = sum(s[1] for s in hired) / 12
        headcount[m] = len(hired)
        ebitda[m] = gp[m] - comm[m] - bonus[m] - mktg[m] - fixed[m] - staff[m]
        cash_out[m] = comm[m] + cogs[m] + mktg[m] + fixed[m] + staff[m] + bonus[m]
        net_cash[m] = cash[m] - cash_out[m]              # cash[m] = prepaid bookings + renewals
        cum_cash[m] = (cum_cash[m - 1] if m > 0 else 0.0) + net_cash[m]
        deferred[m] = (deferred[m - 1] if m > 0 else 0.0) + cash[m] - rec[m]  # unearned/accrued
        retained[m] = (retained[m - 1] if m > 0 else 0.0) + ebitda[m]         # net income proxy
        for role in STAFF:
            if role in hired:
                continue
            if ebitda[m] >= role[2]:
                hired.append(role)
                hire_month[role[0]] = m + 1
            break
    return {
        "avg_annual": avg_annual, "new": new, "active": active, "rec": rec, "cash": cash,
        "cogs": cogs, "gp": gp, "comm": comm, "bonus": bonus, "mktg": mktg, "fixed": fixed,
        "staff": staff, "ebitda": ebitda, "headcount": headcount, "hire_month": hire_month,
        "cash_out": cash_out, "net_cash": net_cash, "cum_cash": cum_cash,
        "deferred": deferred, "retained": retained,
    }


def _watermark(ws: Worksheet) -> None:
    ws.oddFooter.left.text = f"&8© {_ENTITY_HF}  ·  {SIG}"
    ws.oddFooter.center.text = "&8Confidential — not for redistribution"
    ws.oddFooter.right.text = "&8Page &P of &N"


def _title(ws, subtitle, span="N"):
    ws.merge_cells(f"A1:{span}1"); ws["A1"] = ENTITY
    ws["A1"].font = _WHITE; ws["A1"].fill = _NAVY_FILL
    ws.merge_cells(f"A2:{span}2"); ws["A2"] = subtitle
    ws["A2"].font = Font(bold=True, size=13, color=_NAVY)
    ws.merge_cells(f"A3:{span}3"); ws["A3"] = f"generated {date.today().isoformat()}  ·  {SIG}"
    ws["A3"].font = _MUTED


def _assumptions(wb, R):
    a = wb.active; a.title = "Assumptions"
    a.column_dimensions["A"].width = 40; a.column_dimensions["B"].width = 14; a.column_dimensions["C"].width = 44
    _title(a, "5-Year Consolidated Projections — assumptions", "C")
    rows = [
        ("Sales reps (from start)", A["reps"], "Target/ideal sales force"),
        ("Closes / rep / month (full ramp)", A["closes_per_rep_mo"], "30 x 3.5 = 105/mo ≈ 1,260/yr (goal 100/mo)"),
        ("Tier 1 price ($/mo)", A["t1_price"], "Enterprise (rep-sold, prepaid annual)"),
        ("Tier 2 price ($/mo)", A["t2_price"], "Professional (rep-sold, prepaid annual)"),
        ("Tier 1 mix (%)", A["t1_mix"] * 100, "Remainder Tier 2"),
        ("Annual prepay discount (%)", A["prepay_disc"] * 100, "For paying 12 mo up front"),
        ("Upfront commission (%)", A["commission"] * 100, "1099 contractor, paid at signing"),
        ("Year-end MSA bonus (%)", A["msa_bonus"] * 100, "On completed-contract value"),
        ("Annual renewal (%)", A["annual_renewal"] * 100, "Prepaid annual retention"),
        ("Payment processing (%)", A["processing"] * 100, "of revenue (COGS)"),
        ("Cloud / active user / mo ($)", A["cloud_per_user_mo"], "AWS/GCP (COGS)"),
        ("Data license ($/yr, to cap)", A["data_annual"], "NASDAQ/SF1 flat to 1,000 users"),
        ("Data overage ($/user/mo)", A["data_overage_user_mo"], "above 1,000 users"),
        ("Marketing ($/mo)", A["marketing_mo"], "editable"),
        ("Legal & accounting ($/mo)", A["legal_mo"], "editable"),
        ("Software & tools ($/mo)", A["software_mo"], "editable"),
        ("Founder comp ($/yr)", A["founder_annual"], "symbolic $1 + discretionary bonuses"),
    ]
    r = 5
    for label, val, note in rows:
        a.cell(row=r, column=1, value=label).font = _BOLD
        c = a.cell(row=r, column=2, value=val); c.fill = _INPUT_FILL
        c.number_format = _MONEY if val >= 1000 else "0.0"
        a.cell(row=r, column=3, value=note).font = _MUTED
        r += 1
    a.cell(row=r + 1, column=1, value=f"Avg annual price / sub (net): ${R['avg_annual']:,.0f}").font = _BOLD
    a.cell(row=r + 2, column=1,
           value="Basis: accrual (revenue recognized ratably); commission expensed upfront on "
                 "bookings (conservative). Cash view on the Cash Flow lens. Working model — "
                 "tweak inputs and regenerate.").font = _MUTED


def _year_sheet(wb, R, year):
    ws = wb.create_sheet(f"Year {year} (monthly)")
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGHIJKLM":
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["N"].width = 14
    _title(ws, f"Year {year} — monthly P&L, Cash Flow & Balance Sheet")
    hdr = 5
    ws.cell(row=hdr, column=1, value="$ / month").font = _WHITE
    ws.cell(row=hdr, column=1).fill = _NAVY_FILL
    base_m = (year - 1) * 12
    for i in range(12):
        c = ws.cell(row=hdr, column=2 + i, value=f"Mo {base_m + i + 1}")
        c.font = _WHITE; c.fill = _NAVY_FILL; c.alignment = Alignment(horizontal="right")
    tc = ws.cell(row=hdr, column=14, value=f"Year {year}")
    tc.font = _WHITE; tc.fill = _NAVY_FILL; tc.alignment = Alignment(horizontal="right")

    def band(r, title):
        c = ws.cell(row=r, column=1, value=title); c.font = _WHITE; c.fill = _NAVY_FILL
        for i in range(13):
            ws.cell(row=r, column=2 + i).fill = _NAVY_FILL
        return r + 1

    def row(r, label, key, fmt, bold=False, mode="sum", hilite=False):
        lc = ws.cell(row=r, column=1, value=label)
        if bold:
            lc.font = _BOLD
        for i in range(12):
            v = R[key][base_m + i]
            cell = ws.cell(row=r, column=2 + i, value=round(v)); cell.number_format = fmt
            if bold:
                cell.font = _BOLD
            if hilite:
                cell.fill = _HILITE
        tot = R[key][base_m + 11] if mode == "end" else sum(R[key][base_m:base_m + 12])
        tcell = ws.cell(row=r, column=14, value=round(tot)); tcell.number_format = fmt
        if bold:
            tcell.font = _BOLD
        if hilite:
            tcell.fill = _HILITE
        return r + 1

    def margin_row(r):
        ws.cell(row=r, column=1, value="EBITDA margin").font = _BOLD
        for i in range(12):
            rv = R["rec"][base_m + i]; eb = R["ebitda"][base_m + i]
            ws.cell(row=r, column=2 + i, value=(eb / rv if rv else 0)).number_format = _PCT
        yrv = sum(R["rec"][base_m:base_m + 12]); yeb = sum(R["ebitda"][base_m:base_m + 12])
        ws.cell(row=r, column=14, value=(yeb / yrv if yrv else 0)).number_format = _PCT
        return r + 1

    # ── P&L (accrual) ──
    r = 6
    r = row(r, "New subscriptions", "new", "#,##0")
    r = row(r, "Active subscriptions", "active", "#,##0", mode="end")
    r = row(r, "Recognized revenue (accrual)", "rec", _MONEY, bold=True)
    r = row(r, "Total COGS", "cogs", _MONEY)
    r = row(r, "Gross profit", "gp", _MONEY, bold=True)
    r = row(r, "Upfront commission (15%)", "comm", _MONEY)
    r = row(r, "Year-end MSA bonus", "bonus", _MONEY)
    r = row(r, "Marketing", "mktg", _MONEY)
    r = row(r, "Staffing (staged)", "staff", _MONEY)
    r = row(r, "Legal/software/founder", "fixed", _MONEY)
    r = row(r, "EBITDA", "ebitda", _MONEY, bold=True, hilite=True)
    r = margin_row(r)

    # ── Cash Flow (based on prepaid / unrealized revenue) ──
    r += 1
    r = band(r, "CASH FLOW — based on prepaid (unrealized) revenue")
    r = row(r, "Bookings — prepaid cash in", "cash", _MONEY, bold=True)
    r = row(r, "Total cash outflows", "cash_out", _MONEY)
    r = row(r, "Net operating cash", "net_cash", _MONEY, bold=True)
    r = row(r, "Cumulative cash (end)", "cum_cash", _MONEY, bold=True, mode="end", hilite=True)

    # ── Balance Sheet (period end) ──
    r += 1
    r = band(r, "BALANCE SHEET — period end (Cash = Deferred Rev + Retained Earnings)")
    r = row(r, "Assets: Cash", "cum_cash", _MONEY, bold=True, mode="end")
    r = row(r, "Liabilities: Deferred (unearned/accrued) revenue", "deferred", _MONEY, mode="end")
    r = row(r, "Equity: Retained earnings", "retained", _MONEY, mode="end")
    # balance check = Cash - Deferred - Retained (should be ~0)
    lc = ws.cell(row=r, column=1, value="Balance check (Assets − Liab − Equity)"); lc.font = _MUTED
    for i in range(12):
        chk = R["cum_cash"][base_m + i] - R["deferred"][base_m + i] - R["retained"][base_m + i]
        ws.cell(row=r, column=2 + i, value=round(chk)).number_format = _MONEY
    ws.cell(row=r, column=14,
            value=round(R["cum_cash"][base_m + 11] - R["deferred"][base_m + 11] - R["retained"][base_m + 11])
            ).number_format = _MONEY


def _sales_scaling(wb, R):
    ws = wb.create_sheet("Sales Scaling")
    for col, w in zip("ABCDEF", (26, 16, 16, 18, 18, 18)):
        ws.column_dimensions[col].width = w
    _title(ws, "Sales scaling — 1 → 30 reps vs the 100/mo (1,200/yr) goal", "F")
    ws.cell(row=4, column=1,
            value=f"At {A['closes_per_rep_mo']} closes/rep/mo (full ramp). Goal line = 100/mo = "
                  f"1,200/yr. Revenue/commission shown at full-ramp run-rate (avg price "
                  f"${R['avg_annual']:,.0f}/sub/yr).").font = _MUTED
    hdr = ["Reps", "Closes/mo", "Annual subs", "Bookings/yr (cash)", "Upfront comm/yr (15%)", "vs 100/mo goal"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=5, column=c, value=h); cell.font = _WHITE; cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")
    r = 6
    for reps in (1, 5, 10, 15, 20, 25, 30):
        cmo = reps * A["closes_per_rep_mo"]
        subs = cmo * 12
        book = subs * R["avg_annual"]
        comm = book * A["commission"]
        ws.cell(row=r, column=1, value=reps).font = _BOLD if reps == 30 else Font()
        ws.cell(row=r, column=2, value=round(cmo)).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(subs)).number_format = "#,##0"
        ws.cell(row=r, column=4, value=round(book)).number_format = _MONEY
        ws.cell(row=r, column=5, value=round(comm)).number_format = _MONEY
        ws.cell(row=r, column=6, value=f"{cmo/100*100:.0f}% of goal").alignment = Alignment(horizontal="right")
        if reps == 30:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = _HILITE
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Reps from the start is the unknown variable (qualified 1099 hires). Target = "
                  "30. This table shows output at each attainable headcount.").font = _MUTED


def _staffing(wb, R):
    ws = wb.create_sheet("Staffing Plan")
    for col, w in zip("ABCD", (36, 16, 20, 16)):
        ws.column_dimensions[col].width = w
    _title(ws, "Staged staffing — EBITDA-gated hiring", "D")
    ws.cell(row=4, column=1,
            value="Hire in priority order when monthly EBITDA (with roles already on payroll) "
                  "crosses the gate — so a salary is only added once the business supports it.").font = _MUTED
    hdr = ["Role", "Salary ($/yr)", "EBITDA gate ($/mo)", "Hired (month #)"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=5, column=c, value=h); cell.font = _WHITE; cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")
    r = 6
    for role, sal, gate in STAFF:
        hm = R["hire_month"].get(role)
        ws.cell(row=r, column=1, value=role).font = _BOLD
        ws.cell(row=r, column=2, value=sal).number_format = _MONEY
        ws.cell(row=r, column=3, value=gate).number_format = _MONEY
        ws.cell(row=r, column=4,
                value=(f"Mo {hm}" if hm else "not within 5 yr")).alignment = Alignment(horizontal="right")
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Admin/lean first; senior software engineer(s) added only when EBITDA supports "
                  "the salary. Editable gates — this is the working staging rule.").font = _MUTED


def _summary(wb, R):
    ws = wb.create_sheet("5-Year Summary")
    for col, w in zip("ABCDEF", (28, 15, 15, 15, 15, 15)):
        ws.column_dimensions[col].width = w
    _title(ws, "5-Year summary (annual)", "F")
    hdr = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=5, column=c, value=h); cell.font = _WHITE; cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")
    def ysum(key, y):
        return sum(R[key][(y - 1) * 12:(y - 1) * 12 + 12])
    metrics = [
        ("New subscriptions", "new", "#,##0", "sum"),
        ("Ending active subs", "active", "#,##0", "end"),
        ("Recognized revenue", "rec", _MONEY, "sum"),
        ("Bookings (cash in)", "cash", _MONEY, "sum"),
        ("Gross profit", "gp", _MONEY, "sum"),
        ("Upfront commission", "comm", _MONEY, "sum"),
        ("Staffing", "staff", _MONEY, "sum"),
        ("EBITDA", "ebitda", _MONEY, "sum"),
        ("Deferred (unearned) revenue (end)", "deferred", _MONEY, "end"),
        ("Cumulative cash (end)", "cum_cash", _MONEY, "end"),
    ]
    r = 6
    for label, key, fmt, mode in metrics:
        lc = ws.cell(row=r, column=1, value=label)
        if key in ("rec", "ebitda"):
            lc.font = _BOLD
        for y in range(1, 6):
            val = R[key][(y - 1) * 12 + 11] if mode == "end" else ysum(key, y)
            cell = ws.cell(row=r, column=1 + y, value=round(val)); cell.number_format = fmt
            if key in ("rec", "ebitda"):
                cell.font = _BOLD
            if key == "ebitda":
                cell.fill = _HILITE
        r += 1
    # margins
    ws.cell(row=r, column=1, value="EBITDA margin").font = _BOLD
    for y in range(1, 6):
        rv = ysum("rec", y); eb = ysum("ebitda", y)
        ws.cell(row=r, column=1 + y, value=(eb / rv if rv else 0)).number_format = _PCT


def _dashboard(wb, R):
    ws = wb.create_sheet("Exec Dashboard", 0)
    for col, w in zip("AB", (40, 22)):
        ws.column_dimensions[col].width = w
    _title(ws, "Executive dashboard — audited realistic 5-yr projection", "B")
    end_active = R["active"][-1]
    arr = end_active * R["avg_annual"]
    tot_rev = sum(R["rec"]); tot_eb = sum(R["ebitda"]); tot_cash = sum(R["cash"])
    # first EBITDA-positive month
    bem = next((i + 1 for i, e in enumerate(R["ebitda"]) if e > 0), None)
    y5_eb = sum(R["ebitda"][48:60]); y5_rev = sum(R["rec"][48:60])
    kpis = [
        ("Target sales force", f"{A['reps']} reps"),
        ("Ending active subscribers (Yr 5)", f"{end_active:,.0f}"),
        ("Ending ARR (Yr 5)", f"${arr:,.0f}"),
        ("EBITDA-positive from", f"Month {bem}" if bem else "n/a"),
        ("Year 5 revenue (recognized)", f"${y5_rev:,.0f}"),
        ("Year 5 EBITDA", f"${y5_eb:,.0f}"),
        ("Year 5 EBITDA margin", f"{(y5_eb/y5_rev*100 if y5_rev else 0):.1f}%"),
        ("5-yr recognized revenue", f"${tot_rev:,.0f}"),
        ("5-yr EBITDA", f"${tot_eb:,.0f}"),
        ("5-yr cash collected (prepaid)", f"${tot_cash:,.0f}"),
        ("Ending cash (Yr 5)", f"${R['cum_cash'][-1]:,.0f}"),
        ("Ending deferred (unearned) revenue", f"${R['deferred'][-1]:,.0f}"),
        ("Headcount at Yr 5", f"{R['headcount'][-1]} staged hires"),
    ]
    r = 5
    for label, val in kpis:
        ws.cell(row=r, column=1, value=label).font = _BOLD
        c = ws.cell(row=r, column=2, value=val); c.alignment = Alignment(horizontal="right")
        c.fill = _HILITE if "EBITDA" in label or "ARR" in label else _SUB_FILL
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Audited-realistic basis: conservative rep ramp, 85% annual renewal, accrual "
                  "revenue, commission expensed upfront, EBITDA-gated hiring. Working model — "
                  "see Assumptions to tweak.").font = _MUTED


def _legal(wb):
    lg = wb.create_sheet("Legal & Provenance")
    lg.column_dimensions["A"].width = 100
    lg.cell(row=1, column=1, value=ENTITY).font = Font(bold=True, size=13, color=_NAVY)
    for i, line in enumerate([
        "Internal planning model consolidating prior JHI forecasts from user-supplied "
        "assumptions — illustrative only, not a forecast, guarantee, or offer.",
        "Sales headcount (qualified 1099 reps) is the key unknown variable; the Sales Scaling "
        "tab shows output at each attainable headcount toward the 100/mo (1,200/yr) goal.",
        "Staged staffing is EBITDA-gated so salaries are added only when the business supports them.",
        f"Founder signature of provenance: {SIG}",
        f"© {date.today().year} {ENTITY}. All rights reserved. Confidential.",
    ], start=3):
        c = lg.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def build() -> Workbook:
    R = compute()
    wb = Workbook()
    _assumptions(wb, R)
    _dashboard(wb, R)
    _sales_scaling(wb, R)
    _staffing(wb, R)
    for y in range(1, 6):
        _year_sheet(wb, R, y)
    _summary(wb, R)
    _legal(wb)
    for ws in wb.worksheets:
        _watermark(ws)
    return wb


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "JHI_5yr_Consolidated_Projections.xlsx"
    build().save(out)
    print(f"wrote {out}")
