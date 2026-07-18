# JHI-SIG: 69M2705M | Prepaid-MSA Sales Commission Model | John Henry Investments (proprietary)
"""Generate the Prepaid-MSA Sales Commission Excel model.

A rep sells **full-paid (prepaid annual, 12-month MSA) subscriptions** and earns:
  1) an **upfront commission = 15% of the full-paid (net) contract value** at signing, and
  2) a **year-end MSA-completion bonus** on contracts that complete the full 12-month term.

Forecast basis: 100 closes/month → 1,200 subscriptions/year, with ALL expenditures
(COGS + operating costs) included in a full P&L. Every input is editable; formulas
recompute live. No macros.

Sheets:
  1. Assumptions               — editable inputs (prices, mix, prepay, comp rates, opex)
  2. Upfront Commission (15%)  — full-paid contract values + 15% upfront commission
  3. Year-End MSA Bonus        — completion-based year-end bonus
  4. Full Forecast P&L         — revenue → COGS → opex (incl. comp) → EBITDA
  5. Salesperson Total Comp    — upfront + year-end, with mix sensitivity
  6. Legal & Provenance

Run:  python scripts/sales_commission_prepaid_model.py [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SIG = "JHI-SIG: 69M2705M"
ENTITY = "JHI Research & Analytics Firm, Inc."
_ENTITY_HF = ENTITY.replace("&", "&&")

_NAVY_FILL = PatternFill("solid", fgColor="0C1F33")
_SUB_FILL = PatternFill("solid", fgColor="EAEFF6")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")
_HILITE = PatternFill("solid", fgColor="E4F3E9")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)
_MONEY = '"$"#,##0'
_MONEY2 = '"$"#,##0.00'
_PCT = '0.0%'
_NAVY = "0C1F33"


def _watermark(ws: Worksheet) -> None:
    ws.oddFooter.left.text = f"&8© {_ENTITY_HF}  ·  {SIG}"
    ws.oddFooter.center.text = "&8Confidential — not for redistribution"
    ws.oddFooter.right.text = "&8Page &P of &N"


def _title(ws: Worksheet, subtitle: str, span: str = "D") -> None:
    ws.merge_cells(f"A1:{span}1")
    ws["A1"] = ENTITY
    ws["A1"].font = _WHITE
    ws["A1"].fill = _NAVY_FILL
    ws.merge_cells(f"A2:{span}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(bold=True, size=13, color=_NAVY)
    ws.merge_cells(f"A3:{span}3")
    ws["A3"] = f"generated {date.today().isoformat()}  ·  {SIG}"
    ws["A3"].font = _MUTED


def _assumptions(wb: Workbook) -> Worksheet:
    a = wb.active
    a.title = "Assumptions"
    a.column_dimensions["A"].width = 40
    a.column_dimensions["B"].width = 14
    a.column_dimensions["C"].width = 46
    _title(a, "Prepaid-MSA Sales Commission — assumptions", "C")

    inputs = [
        ("Tier 1 price ($/mo)", 1500, "Enterprise / Family Office"),
        ("Tier 2 price ($/mo)", 299, "Professional"),
        ("Tier 1 mix (% of closes)", 10, "Share of closes that are Tier 1"),
        ("Closes per month", 100, "New full-paid subs signed each month"),
        ("Annual prepay discount (%)", 10, "Discount for paying 12 months up front"),
        ("Upfront commission rate (%)", 15, "Rep % of full-paid (net) contract value, at signing"),
        ("MSA completion / renewal rate (%)", 90, "Share of contracts that complete the full 12-mo MSA"),
        ("Year-end MSA-completion bonus rate (%)", 5, "Rep bonus % of completed-contract value, paid at year-end"),
    ]
    r = 5
    for label, val, note in inputs:
        a.cell(row=r, column=1, value=label).font = _BOLD
        c = a.cell(row=r, column=2, value=val)
        c.fill = _INPUT_FILL
        c.number_format = "#,##0" if val >= 100 else "0.0"
        a.cell(row=r, column=3, value=note).font = _MUTED
        r += 1
    # refs: B5 p1, B6 p2, B7 mix, B8 closes, B9 disc, B10 commrate, B11 complrate, B12 bonusrate

    # Derived
    a.cell(row=14, column=1, value="DERIVED").font = _BOLD
    a.cell(row=14, column=1).fill = _SUB_FILL
    a.cell(row=15, column=1, value="Subscriptions per year").font = _BOLD
    a.cell(row=15, column=2, value="=B8*12").number_format = "#,##0"
    a.cell(row=15, column=3, value="Closes/mo × 12 (e.g. 100 → 1,200)").font = _MUTED
    a.cell(row=16, column=1, value="Avg full-paid price / sub (net, annual)").font = _BOLD
    a.cell(row=16, column=2,
           value="=((B5*B7/100)+(B6*(1-B7/100)))*12*(1-B9/100)").number_format = _MONEY
    a.cell(row=16, column=3, value="Blended monthly price × 12 × (1 − prepay discount)").font = _MUTED
    a.cell(row=17, column=1, value="Total prepaid annual revenue (net)").font = _BOLD
    a.cell(row=17, column=2, value="=B15*B16").number_format = _MONEY
    a.cell(row=17, column=2).fill = _HILITE
    # refs: B15 subs/yr, B16 avg price/sub, B17 total revenue

    a.cell(row=19, column=1, value="OPERATING ASSUMPTIONS (editable)").font = _BOLD
    a.cell(row=19, column=1).fill = _SUB_FILL
    opex = [
        ("Payment processing (% of revenue)", 3),
        ("Infra / support (% of revenue)", 2),
        ("Data license — SF1 ($/yr, fixed)", 18000),
        ("Marketing ($/yr)", 36000),
        ("Legal & accounting ($/yr)", 18000),
        ("Software & AI tools ($/yr)", 12000),
        ("Founder / ops comp ($/yr)", 60000),
    ]
    rr = 20
    for label, val in opex:
        a.cell(row=rr, column=1, value=label).font = _BOLD
        cc = a.cell(row=rr, column=2, value=val)
        cc.fill = _INPUT_FILL
        cc.number_format = _MONEY if val >= 1000 else "0.0"
        rr += 1
    # refs: proc=B20 infra=B21 sf1=B22 mktg=B23 legal=B24 sw=B25 founder=B26
    return a


def _upfront(wb: Workbook) -> None:
    u = wb.create_sheet("Upfront Commission (15%)")
    for col, w in zip("ABCD", (40, 16, 16, 16)):
        u.column_dimensions[col].width = w
    _title(u, "Upfront commission — 15% of full-paid (net) contract value")
    u.cell(row=4, column=1,
           value="Contracts are prepaid annual (12-mo MSA). Rep earns the upfront commission at "
                 "signing on the net (post-discount) contract value. Pay on collected cash.").font = _MUTED

    hdr = ["", "Tier 1", "Tier 2", "Total"]
    for c, h in enumerate(hdr, start=1):
        cell = u.cell(row=5, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")

    rows = [
        ("Subscriptions / year", "=Assumptions!$B$15*Assumptions!$B$7/100",
         "=Assumptions!$B$15*(1-Assumptions!$B$7/100)", "=B6+C6", "#,##0"),
        ("Full-paid price / sub (net, annual)", "=Assumptions!$B$5*12*(1-Assumptions!$B$9/100)",
         "=Assumptions!$B$6*12*(1-Assumptions!$B$9/100)", "", _MONEY),
        ("Prepaid annual revenue (net)", "=B6*B7", "=C6*C7", "=B8+C8", _MONEY),
        ("Upfront commission / sub (15%)", "=B7*Assumptions!$B$10/100",
         "=C7*Assumptions!$B$10/100", "", _MONEY),
        ("Upfront commission (total)", "=B8*Assumptions!$B$10/100",
         "=C8*Assumptions!$B$10/100", "=B10+C10", _MONEY),
    ]
    r = 6
    for label, b, c, d, fmt in rows:
        lc = u.cell(row=r, column=1, value=label)
        if label.startswith("Upfront commission (total)") or label.startswith("Prepaid annual revenue"):
            lc.font = _BOLD
        for col, formula in zip("BCD", (b, c, d)):
            if formula == "":
                continue
            cell = u[f"{col}{r}"]
            cell.value = formula
            cell.number_format = fmt
            if label.startswith("Upfront commission (total)"):
                cell.font = _BOLD
                cell.fill = _HILITE
        r += 1

    u.cell(row=13, column=1,
           value="Cash-flow note: 15% paid UP FRONT on the full annual contract is front-loaded "
                 "vs. a monthly residual — model the cash outlay at signing.").font = _MUTED


def _bonus(wb: Workbook) -> None:
    b = wb.create_sheet("Year-End MSA Bonus")
    for col, w in zip("ABC", (48, 18, 44)):
        b.column_dimensions[col].width = w
    _title(b, "Year-end MSA-completion bonus", "C")
    b.cell(row=4, column=1,
           value="Additional bonus rewarding contracts that COMPLETE the full 12-month MSA term "
                 "(retention). Paid at year-end on completed-contract value.").font = _MUTED

    rows = [
        ("Subscriptions / year (signed)", "=Assumptions!$B$15", "#,##0",
         "From Assumptions (closes/mo × 12)"),
        ("MSA completion / renewal rate", "=Assumptions!$B$11/100", _PCT,
         "Share completing the full 12-mo term"),
        ("Contracts completing full 12-mo MSA", "=B5*B6", "#,##0",
         "Signed × completion rate"),
        ("Completed-contract value (net)", "=Assumptions!$B$17*B6", _MONEY,
         "Total prepaid revenue × completion rate"),
        ("Year-end bonus rate", "=Assumptions!$B$12/100", _PCT,
         "Rep % of completed-contract value"),
        ("Year-end MSA-completion bonus", "=B8*B9", _MONEY,
         "Completed value × bonus rate"),
    ]
    r = 5
    for label, formula, fmt, note in rows:
        lc = b.cell(row=r, column=1, value=label)
        cc = b.cell(row=r, column=2, value=formula)
        cc.number_format = fmt
        b.cell(row=r, column=3, value=note).font = _MUTED
        if label.startswith("Year-end MSA-completion bonus"):
            lc.font = _BOLD
            cc.font = _BOLD
            cc.fill = _HILITE
        r += 1


def _pnl(wb: Workbook) -> None:
    p = wb.create_sheet("Full Forecast P&L")
    for col, w in zip("ABC", (40, 18, 44)):
        p.column_dimensions[col].width = w
    _title(p, "Full-year forecast P&L — all expenditures included", "C")
    p.cell(row=4, column=1,
           value="1,200 prepaid annual subs (editable). Includes COGS + upfront 15% commission + "
                 "year-end MSA bonus + all operating costs. Illustrative internal model.").font = _MUTED

    rows = [
        ("Revenue (prepaid annual, net)", "=Assumptions!$B$17", _MONEY, True, None),
        ("  Payment processing", "=-B5*Assumptions!$B$20/100", _MONEY, False, None),
        ("  Infrastructure & support", "=-B5*Assumptions!$B$21/100", _MONEY, False, None),
        ("  Data license (SF1)", "=-Assumptions!$B$22", _MONEY, False, None),
        ("Gross profit", "=B5+SUM(B6:B8)", _MONEY, True, _SUB_FILL),
        ("  Upfront sales commission (15%)", "=-'Upfront Commission (15%)'!$D$10", _MONEY, False, None),
        ("  Year-end MSA-completion bonus", "=-'Year-End MSA Bonus'!$B$10", _MONEY, False, None),
        ("  Marketing", "=-Assumptions!$B$23", _MONEY, False, None),
        ("  Legal & accounting", "=-Assumptions!$B$24", _MONEY, False, None),
        ("  Software & AI tools", "=-Assumptions!$B$25", _MONEY, False, None),
        ("  Founder / ops comp", "=-Assumptions!$B$26", _MONEY, False, None),
        ("Total operating expenses", "=SUM(B10:B15)", _MONEY, True, None),
        ("EBITDA", "=B9+B16", _MONEY, True, _HILITE),
        ("EBITDA margin", "=IF(B5=0,0,B17/B5)", _PCT, False, None),
        ("Total salesperson compensation", "='Upfront Commission (15%)'!$D$10+'Year-End MSA Bonus'!$B$10", _MONEY, True, None),
    ]
    r = 5
    for label, formula, fmt, bold, fill in rows:
        lc = p.cell(row=r, column=1, value=label)
        cc = p.cell(row=r, column=2, value=formula)
        cc.number_format = fmt
        if bold:
            lc.font = _BOLD
            cc.font = _BOLD
        if fill:
            cc.fill = fill
            lc.fill = fill
        r += 1

    p.cell(row=r + 1, column=1,
           value="Reality check: 1,200 premium prepaid subs in Year 1 from one rep is an aggressive "
                 "'ceiling' scenario. Dial closes/mo down in Assumptions for a base case.").font = _MUTED


def _total_comp(wb: Workbook) -> None:
    t = wb.create_sheet("Salesperson Total Comp")
    for col, w in zip("ABCD", (40, 16, 16, 16)):
        t.column_dimensions[col].width = w
    _title(t, "Salesperson total compensation & mix sensitivity")

    # Headline single-scenario
    rows = [
        ("Upfront commission (15%)", "='Upfront Commission (15%)'!$D$10", _MONEY, True),
        ("Year-end MSA-completion bonus", "='Year-End MSA Bonus'!$B$10", _MONEY, True),
        ("Total salesperson compensation", "=B5+B6", _MONEY, True),
        ("Total comp as % of revenue", "=IF(Assumptions!$B$17=0,0,B7/Assumptions!$B$17)", _PCT, False),
    ]
    r = 5
    for label, formula, fmt, bold in rows:
        lc = t.cell(row=r, column=1, value=label)
        cc = t.cell(row=r, column=2, value=formula)
        cc.number_format = fmt
        if bold:
            lc.font = _BOLD
            cc.font = _BOLD
        if label.startswith("Total salesperson compensation"):
            cc.fill = _HILITE
        r += 1

    # Mix sensitivity: total comp by Tier-1 mix (0 / 10 / 20%)
    st = 11
    t.cell(row=st, column=1, value="Sensitivity — by Tier-1 mix (1,200 subs)").font = _BOLD
    t.cell(row=st, column=1).fill = _SUB_FILL
    hdr = ["Metric", "0% T1", "10% T1", "20% T1"]
    for c, h in enumerate(hdr, start=1):
        cell = t.cell(row=st + 1, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")
    mixes = {"B": 0, "C": 10, "D": 20}
    base = st + 2
    labels = [
        "Avg full-paid price / sub (net)",
        "Prepaid annual revenue (net)",
        "Upfront commission (15%)",
        "Year-end MSA bonus",
        "Total salesperson comp",
    ]
    for i, lab in enumerate(labels):
        t.cell(row=base + i, column=1, value=lab).font = _BOLD if lab.startswith("Total") else Font()
    for col, m in mixes.items():
        price = f"((Assumptions!$B$5*{m}/100)+(Assumptions!$B$6*(1-{m}/100)))*12*(1-Assumptions!$B$9/100)"
        rev = f"Assumptions!$B$15*{price}"
        t[f"{col}{base}"] = f"={price}"
        t[f"{col}{base+1}"] = f"={rev}"
        t[f"{col}{base+2}"] = f"=({rev})*Assumptions!$B$10/100"
        t[f"{col}{base+3}"] = f"=({rev})*Assumptions!$B$11/100*Assumptions!$B$12/100"
        t[f"{col}{base+4}"] = f"={col}{base+2}+{col}{base+3}"
        for k in range(5):
            t[f"{col}{base+k}"].number_format = _MONEY
        t[f"{col}{base+4}"].font = _BOLD
        t[f"{col}{base+4}"].fill = _HILITE

    t.cell(row=base + 6, column=1,
           value="Note: upfront commission is front-loaded cash at signing; the year-end bonus rewards "
                 "12-mo MSA completion (retention). Confirm plan mechanics with counsel/CPA.").font = _MUTED


def _legal(wb: Workbook) -> None:
    lg = wb.create_sheet("Legal & Provenance")
    lg.column_dimensions["A"].width = 100
    lg.cell(row=1, column=1, value=ENTITY).font = Font(bold=True, size=13, color=_NAVY)
    for i, line in enumerate([
        "Internal planning model from user-supplied assumptions — illustrative only, not a forecast, "
        "offer of employment, or compensation guarantee.",
        "Upfront commission is paid on collected, full-paid (prepaid annual) contracts. Consider a "
        "pro-rata clawback if a prepaid contract is refunded/cancelled before the 12-mo MSA completes.",
        "The year-end MSA-completion bonus rewards contracts that complete the full 12-month term.",
        f"Founder signature of provenance: {SIG}",
        f"© {date.today().year} {ENTITY}. All rights reserved. Confidential.",
    ], start=3):
        cell = lg.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def build() -> Workbook:
    wb = Workbook()
    _assumptions(wb)
    _upfront(wb)
    _bonus(wb)
    _pnl(wb)
    _total_comp(wb)
    _legal(wb)
    for ws in wb.worksheets:
        _watermark(ws)
    return wb


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "JHI_Sales_Commission_Prepaid_MSA.xlsx"
    build().save(out)
    print(f"wrote {out}")
