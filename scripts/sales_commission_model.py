# JHI-SIG: 69M2705M | Sales Commission Model | JHI Research & Analytics Firm, Inc. (proprietary)
"""Generate the interactive Sales Commission + EBITDA Excel model.

A 100%-commission ground-floor rep on Tier 1 ($1,500) + Tier 2 ($299), paid a
monthly residual (% of MRR) while subscriptions stay active. Editable assumptions
drive a live 24-month schedule and a Year-1-by-mix table. No macros.

Sheets:
  1. Assumptions              — editable inputs (prices, mix, opex)
  2. Monthly Model            — 24-month commission schedule
  3. Year-1 by Mix            — commission by Tier 1/2 mix
  4. Year-1 EBITDA by Mix     — annual P&L parallel to the commission table
  5. Monthly EBITDA & Opex    — month-by-month P&L with operating-cost columns
  6. Salesperson Bonus (MSA)  — prepaid annual Tier 1&2 book, bonus = 10% of EBITDA
  7. Legal & Provenance

Run:  python scripts/sales_commission_model.py [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SIG = "JHI-SIG: 69M2705M"
ENTITY = "JHI Research & Analytics Firm, Inc."
_ENTITY_HF = ENTITY.replace("&", "&&")

_NAVY_FILL = PatternFill("solid", fgColor="0C1F33")
_SUB_FILL = PatternFill("solid", fgColor="EAEFF6")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)
_MONEY = '"$"#,##0'
_PCT = '0.0%'
_NAVY = "0C1F33"


def _watermark(ws: Worksheet) -> None:
    ws.oddFooter.left.text = f"&8© {_ENTITY_HF}  ·  {SIG}"
    ws.oddFooter.center.text = "&8Confidential — not for redistribution"
    ws.oddFooter.right.text = "&8Page &P of &N"


def _title(ws: Worksheet, subtitle: str) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = ENTITY
    ws["A1"].font = _WHITE
    ws["A1"].fill = _NAVY_FILL
    ws.merge_cells("A2:D2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(bold=True, size=13, color=_NAVY)
    ws.merge_cells("A3:D3")
    ws["A3"] = f"generated {date.today().isoformat()}  ·  {SIG}"
    ws["A3"].font = _MUTED


def _monthly_ebitda(wb: Workbook) -> None:
    """Month-by-month Year-1 P&L: revenue → COGS → opex → EBITDA, months as columns.

    Fixed annual costs (SF1 data license, marketing, legal, software, founder comp)
    are spread evenly across 12 months; variable costs scale with monthly revenue.
    """
    me = wb.create_sheet("Monthly EBITDA & Opex")
    me.column_dimensions["A"].width = 30
    for col in "BCDEFGHIJKLM":
        me.column_dimensions[col].width = 11
    me.column_dimensions["N"].width = 14
    _title(me, "Year-1 monthly EBITDA & operating-cost statement")
    me.cell(
        row=4,
        column=1,
        value="Fixed annual costs spread /12; variable costs scale with revenue. Ramp = "
        "closes/mo from Assumptions; churn per Assumptions.",
    ).font = _MUTED

    # Header: A = line item, B..M = Mo 1..12, N = Year-1 total
    hdr_row = 5
    me.cell(row=hdr_row, column=1, value="$ per month").font = _WHITE
    me.cell(row=hdr_row, column=1).fill = _NAVY_FILL
    for i in range(12):
        c = me.cell(row=hdr_row, column=2 + i, value=f"Mo {i + 1}")
        c.font = _WHITE
        c.fill = _NAVY_FILL
        c.alignment = Alignment(horizontal="right")
    tot = me.cell(row=hdr_row, column=14, value="Year 1")
    tot.font = _WHITE
    tot.fill = _NAVY_FILL
    tot.alignment = Alignment(horizontal="right")

    cols = "BCDEFGHIJKLM"  # months 1..12
    # Row map (label, per-month formula template using {c}=current col & {p}=prev col, number format, bold?)
    R_NEW, R_ACTIVE, R_REV = 6, 7, 8
    R_PROC, R_INFRA, R_SF1, R_COGS, R_GP = 9, 10, 11, 12, 13
    R_COMM, R_MKTG, R_LEGAL, R_SW, R_FOUND, R_OPEX, R_EBITDA, R_MARGIN = 14, 15, 16, 17, 18, 19, 20, 21

    def put(row: int, label: str, fmt: str, bold: bool = False, fill: PatternFill | None = None) -> None:
        lc = me.cell(row=row, column=1, value=label)
        if bold:
            lc.font = _BOLD
        if fill:
            lc.fill = fill

    put(R_NEW, "New subscriptions", "#,##0")
    put(R_ACTIVE, "Active subscriptions", "#,##0")
    put(R_REV, "Billed MRR (revenue)", _MONEY, bold=True)
    put(R_PROC, "  Payment processing", _MONEY)
    put(R_INFRA, "  Infrastructure & support", _MONEY)
    put(R_SF1, "  Data license (SF1)", _MONEY)
    put(R_COGS, "Total COGS", _MONEY)
    put(R_GP, "Gross profit", _MONEY, bold=True)
    put(R_COMM, "  Sales commission (residual)", _MONEY)
    put(R_MKTG, "  Marketing", _MONEY)
    put(R_LEGAL, "  Legal & accounting", _MONEY)
    put(R_SW, "  Software & AI tools", _MONEY)
    put(R_FOUND, "  Founder / ops comp", _MONEY)
    put(R_OPEX, "Total operating expenses", _MONEY, bold=True)
    put(R_EBITDA, "EBITDA", _MONEY, bold=True, fill=_SUB_FILL)
    put(R_MARGIN, "EBITDA margin", _PCT)

    for i, col in enumerate(cols):
        prev = cols[i - 1] if i > 0 else None
        me[f"{col}{R_NEW}"] = "=Assumptions!$B$8"
        me[f"{col}{R_NEW}"].number_format = "#,##0"
        if prev is None:
            me[f"{col}{R_ACTIVE}"] = f"={col}{R_NEW}"
        else:
            me[f"{col}{R_ACTIVE}"] = f"={prev}{R_ACTIVE}*(1-Assumptions!$B$10/100)+{col}{R_NEW}"
        me[f"{col}{R_ACTIVE}"].number_format = "#,##0"
        me[f"{col}{R_REV}"] = f"={col}{R_ACTIVE}*Assumptions!$B$13"
        me[f"{col}{R_PROC}"] = f"={col}{R_REV}*Assumptions!$B$16/100"
        me[f"{col}{R_INFRA}"] = f"={col}{R_REV}*Assumptions!$B$17/100"
        me[f"{col}{R_SF1}"] = "=Assumptions!$B$18/12"
        me[f"{col}{R_COGS}"] = f"=SUM({col}{R_PROC}:{col}{R_SF1})"
        me[f"{col}{R_GP}"] = f"={col}{R_REV}-{col}{R_COGS}"
        me[f"{col}{R_COMM}"] = f"={col}{R_REV}*Assumptions!$B$9/100"
        me[f"{col}{R_MKTG}"] = "=Assumptions!$B$19/12"
        me[f"{col}{R_LEGAL}"] = "=Assumptions!$B$20/12"
        me[f"{col}{R_SW}"] = "=Assumptions!$B$21/12"
        me[f"{col}{R_FOUND}"] = "=Assumptions!$B$22/12"
        me[f"{col}{R_OPEX}"] = f"=SUM({col}{R_COMM}:{col}{R_FOUND})"
        me[f"{col}{R_EBITDA}"] = f"={col}{R_GP}-{col}{R_OPEX}"
        me[f"{col}{R_MARGIN}"] = f"=IF({col}{R_REV}=0,0,{col}{R_EBITDA}/{col}{R_REV})"
        for row in (R_REV, R_PROC, R_INFRA, R_SF1, R_COGS, R_GP, R_COMM, R_MKTG, R_LEGAL, R_SW, R_FOUND, R_OPEX, R_EBITDA):
            me[f"{col}{row}"].number_format = _MONEY
        me[f"{col}{R_MARGIN}"].number_format = _PCT

    # Year-1 total column (N): flows = SUM; stocks/margin handled explicitly.
    n = "N"
    for row in (R_NEW, R_REV, R_PROC, R_INFRA, R_SF1, R_COGS, R_GP, R_COMM, R_MKTG, R_LEGAL, R_SW, R_FOUND, R_OPEX, R_EBITDA):
        me[f"{n}{row}"] = f"=SUM(B{row}:M{row})"
        me[f"{n}{row}"].number_format = "#,##0" if row == R_NEW else _MONEY
    me[f"{n}{R_ACTIVE}"] = f"=M{R_ACTIVE}"  # ending active subs
    me[f"{n}{R_ACTIVE}"].number_format = "#,##0"
    me[f"{n}{R_MARGIN}"] = f"=IF(N{R_REV}=0,0,N{R_EBITDA}/N{R_REV})"
    me[f"{n}{R_MARGIN}"].number_format = _PCT
    for row in (R_REV, R_GP, R_OPEX, R_EBITDA):
        me[f"{n}{row}"].font = _BOLD

    me.cell(
        row=R_MARGIN + 2,
        column=1,
        value="Note: 'Active subscriptions' Year-1 column shows the ending balance (a stock, not a sum).",
    ).font = _MUTED


def _bonus(wb: Workbook) -> None:
    """Salesperson bonus on a prepaid-annual Tier 1 & 2 MSA book.

    Contracts are paid up front for 12 months at a prepay discount; the salesperson
    bonus is a profit share = bonus_rate % of the book's EBITDA (after gross margin
    and operating expenses). Bonus floored at $0.
    """
    b = wb.create_sheet("Salesperson Bonus (MSA)")
    for col, w in zip("ABCDE", (40, 16, 16, 16, 16)):
        b.column_dimensions[col].width = w
    _title(b, "Salesperson bonus — prepaid annual MSA (Tier 1 & 2)")
    b.cell(
        row=4,
        column=1,
        value="Annual MSA paid up front at a prepay discount. Bonus = rate % of EBITDA "
        "(after gross margins & opex), floored at $0. Illustrative internal model.",
    ).font = _MUTED

    # --- Editable inputs ---
    inputs = [
        ("Tier 1 annual contracts (count)", 50, "Enterprise / Family Office, prepaid annual"),
        ("Tier 2 annual contracts (count)", 500, "Professional, prepaid annual"),
        ("Annual prepay discount (%)", 10, "Discount for paying 12 months up front"),
        ("Bonus rate (% of EBITDA)", 10, "Salesperson profit share on this book"),
    ]
    r = 5
    for label, val, note in inputs:
        b.cell(row=r, column=1, value=label).font = _BOLD
        c = b.cell(row=r, column=2, value=val)
        c.fill = _INPUT_FILL
        c.number_format = "#,##0" if val >= 100 else "0.0"
        b.cell(row=r, column=3, value=note).font = _MUTED
        r += 1
    # refs: B5 T1count, B6 T2count, B7 disc, B8 bonusrate

    # --- Derived P&L (single scenario driven by inputs) ---
    rows = [
        ("Tier 1 annual list price", "=Assumptions!$B$5*12", _MONEY, False),
        ("Tier 1 prepaid price (net of discount)", "=Assumptions!$B$5*12*(1-$B$7/100)", _MONEY, False),
        ("Tier 2 annual list price", "=Assumptions!$B$6*12", _MONEY, False),
        ("Tier 2 prepaid price (net of discount)", "=Assumptions!$B$6*12*(1-$B$7/100)", _MONEY, False),
        ("Tier 1 annual revenue", "=$B$5*Assumptions!$B$5*12*(1-$B$7/100)", _MONEY, False),
        ("Tier 2 annual revenue", "=$B$6*Assumptions!$B$6*12*(1-$B$7/100)", _MONEY, False),
        ("Total annual revenue (prepaid, net)", "=B14+B15", _MONEY, True),
        ("COGS (proc+infra % + SF1)", "=B16*(Assumptions!$B$16+Assumptions!$B$17)/100+Assumptions!$B$18", _MONEY, False),
        ("Gross profit", "=B16-B17", _MONEY, True),
        ("Operating expenses (mktg+legal+sw+founder)", "=Assumptions!$B$19+Assumptions!$B$20+Assumptions!$B$21+Assumptions!$B$22", _MONEY, False),
        ("EBITDA (after gross margins & opex)", "=B18-B19", _MONEY, True),
        ("EBITDA margin", "=IF(B16=0,0,B20/B16)", _PCT, False),
        ("Salesperson bonus (rate % of EBITDA)", "=MAX(0,B20*$B$8/100)", _MONEY, True),
        ("EBITDA after salesperson bonus", "=B20-B22", _MONEY, False),
    ]
    rr = 10
    for label, formula, fmt, bold in rows:
        lc = b.cell(row=rr, column=1, value=label)
        cc = b.cell(row=rr, column=2, value=formula)
        cc.number_format = fmt
        if bold:
            lc.font = _BOLD
            cc.font = _BOLD
        if label.startswith("Salesperson bonus"):
            lc.fill = _SUB_FILL
            cc.fill = _SUB_FILL
        rr += 1

    # --- Sensitivity table: bonus by book volume (fixed 10% T1 / 90% T2 mix) ---
    st = rr + 2
    b.cell(row=st, column=1, value="Sensitivity — bonus by book volume (≈10% T1 / 90% T2)").font = _BOLD
    b.cell(row=st, column=1).fill = _SUB_FILL
    hdr = ["Metric", "100 contracts", "300 contracts", "600 contracts"]
    for c, h in enumerate(hdr, start=1):
        cell = b.cell(row=st + 1, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")
    # counts per column: (T1, T2)
    counts = {"B": (10, 90), "C": (30, 270), "D": (60, 540)}
    base = st + 2
    labels = [
        "Contracts (T1 / T2)",
        "Annual revenue (prepaid, net)",
        "Gross profit",
        "EBITDA",
        "Salesperson bonus",
    ]
    for i, lab in enumerate(labels):
        b.cell(row=base + i, column=1, value=lab).font = _BOLD if lab in ("EBITDA", "Salesperson bonus") else Font()
    for col, (t1, t2) in counts.items():
        rev = f"({t1}*Assumptions!$B$5+{t2}*Assumptions!$B$6)*12*(1-$B$7/100)"
        b[f"{col}{base}"] = f"{t1} / {t2}"
        b[f"{col}{base}"].alignment = Alignment(horizontal="right")
        b[f"{col}{base+1}"] = f"={rev}"
        b[f"{col}{base+2}"] = f"={rev}-(({rev})*(Assumptions!$B$16+Assumptions!$B$17)/100+Assumptions!$B$18)"
        # EBITDA = GP - fixed opex
        b[f"{col}{base+3}"] = f"={col}{base+2}-(Assumptions!$B$19+Assumptions!$B$20+Assumptions!$B$21+Assumptions!$B$22)"
        b[f"{col}{base+4}"] = f"=MAX(0,{col}{base+3}*$B$8/100)"
        for k in range(1, 5):
            b[f"{col}{base+k}"].number_format = _MONEY
        b[f"{col}{base+4}"].font = _BOLD

    b.cell(
        row=base + 6,
        column=1,
        value="Note: book-level EBITDA charges the full fixed annual opex to this book (conservative "
        "at low volume). Bonus floored at $0. Confirm plan mechanics with counsel/CPA.",
    ).font = _MUTED


def build() -> Workbook:
    wb = Workbook()

    # --- Assumptions (editable) ---
    a = wb.active
    a.title = "Assumptions"
    a.column_dimensions["A"].width = 36
    a.column_dimensions["B"].width = 14
    a.column_dimensions["C"].width = 40
    _title(a, "Sales Commission Model — assumptions")
    rows = [
        ("Tier 1 price ($/mo)", 1500, "Enterprise / Family Office"),
        ("Tier 2 price ($/mo)", 299, "Professional"),
        ("Tier 1 mix (% of closes)", 0, "0 = all Tier 2; try 10 or 20"),
        ("Closes per month", 100, "New subs signed each month"),
        ("Residual commission (% of MRR)", 10, "Paid monthly while active"),
        ("Monthly churn (%)", 0, "Subs lost per month"),
        ("Blended gross margin (%)", 96, "For JHI contribution"),
    ]
    r = 5
    for label, val, note in rows:
        a.cell(row=r, column=1, value=label).font = _BOLD
        c = a.cell(row=r, column=2, value=val)
        c.fill = _INPUT_FILL
        a.cell(row=r, column=3, value=note).font = _MUTED
        r += 1
    # cell refs
    p1, p2, mix, closes, resid, churn, gm = (f"B{i}" for i in range(5, 12))

    a.cell(row=13, column=1, value="Avg MRR / sub").font = _BOLD
    a.cell(row=13, column=2, value=f"=${p1[0]}${p1[1:]}*${mix[0]}${mix[1:]}/100+${p2[0]}${p2[1:]}*(1-${mix[0]}${mix[1:]}/100)").number_format = _MONEY

    # Operating assumptions (editable) for the EBITDA view.
    a.cell(row=15, column=1, value="OPERATING ASSUMPTIONS (editable)").font = _BOLD
    a.cell(row=15, column=1).fill = _SUB_FILL
    opex = [
        ("Payment processing (% of revenue)", 3),
        ("Infra / support (% of revenue)", 2),
        ("Data license — SF1 ($/yr, fixed)", 18000),
        ("Marketing ($/yr)", 36000),
        ("Legal & accounting ($/yr)", 18000),
        ("Software & AI tools ($/yr)", 12000),
        ("Founder / ops comp ($/yr)", 60000),
    ]
    rr = 16
    for label, val in opex:
        a.cell(row=rr, column=1, value=label).font = _BOLD
        cc = a.cell(row=rr, column=2, value=val)
        cc.fill = _INPUT_FILL
        cc.number_format = _MONEY if val >= 1000 else "0.0"
        rr += 1
    # Assumptions refs: proc=B16, infra=B17, dataSF1=B18, mktg=B19, legal=B20, sw=B21, founder=B22

    # --- Monthly Model (live 24-month schedule) ---
    m = wb.create_sheet("Monthly Model")
    for col, w in zip("ABCDEFG", (8, 12, 12, 12, 14, 14, 18)):
        m.column_dimensions[col].width = w
    _title(m, "24-month commission schedule (live)")
    headers = ["Month", "New subs", "Active subs", "Avg MRR", "Billed MRR", "Commission", "JHI GM after comm."]
    for c, h in enumerate(headers, start=1):
        cell = m.cell(row=5, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
    first = 6
    for i in range(24):
        row = first + i
        m.cell(row=row, column=1, value=i + 1)
        m.cell(row=row, column=2, value=f"=Assumptions!${closes[0]}${closes[1:]}")
        if i == 0:
            active = f"=B{row}"
        else:
            active = f"=C{row-1}*(1-Assumptions!${churn[0]}${churn[1:]}/100)+B{row}"
        m.cell(row=row, column=3, value=active).number_format = "#,##0"
        m.cell(row=row, column=4, value="=Assumptions!$B$13").number_format = _MONEY
        m.cell(row=row, column=5, value=f"=C{row}*D{row}").number_format = _MONEY
        m.cell(row=row, column=6, value=f"=E{row}*Assumptions!${resid[0]}${resid[1:]}/100").number_format = _MONEY
        m.cell(row=row, column=7, value=f"=E{row}*Assumptions!${gm[0]}${gm[1:]}/100-F{row}").number_format = _MONEY
    last = first + 23
    y1 = first + 11
    # Summary
    s = last + 2
    m.cell(row=s, column=1, value="Year-1 commission").font = _BOLD
    m.cell(row=s, column=6, value=f"=SUM(F{first}:F{y1})").number_format = _MONEY
    m.cell(row=s + 1, column=1, value="Year-2 commission").font = _BOLD
    m.cell(row=s + 1, column=6, value=f"=SUM(F{y1+1}:F{last})").number_format = _MONEY
    m.cell(row=s + 2, column=1, value="Exit run-rate (mo 12 × 12)").font = _BOLD
    m.cell(row=s + 2, column=6, value=f"=F{y1}*12").number_format = _MONEY
    m.cell(row=s + 3, column=1, value="Year-1 JHI GM after commission").font = _BOLD
    m.cell(row=s + 3, column=7, value=f"=SUM(G{first}:G{y1})").number_format = _MONEY

    # --- Year-1 by mix (reference; no-churn ramp = residual% × avgMRR × closes × 78) ---
    x = wb.create_sheet("Year-1 by Mix")
    for col, w in zip("ABCD", (18, 16, 22, 22)):
        x.column_dimensions[col].width = w
    _title(x, "Year-1 commission by Tier 1 / Tier 2 mix")
    x.cell(row=5, column=1, value="Ramp basis: 100 closes/mo → sub-months = closes × (1+2+…+12) = closes × 78 (no churn).").font = _MUTED
    hdr = ["Mix (T1 / T2)", "Avg MRR", "Year-1 commission", "Exit run-rate (yr)"]
    for c, h in enumerate(hdr, start=1):
        cell = x.cell(row=7, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
    mixes = [(0, "0% / 100%"), (10, "10% / 90%"), (20, "20% / 80%")]
    rr = 8
    for pct, label in mixes:
        avg = f"=Assumptions!$B$5*{pct}/100+Assumptions!$B$6*(1-{pct}/100)"
        # note: B5/B6 are prices? prices are B5=Tier1,B6=Tier2 -> yes rows 5,6
        x.cell(row=rr, column=1, value=label).font = _BOLD
        x.cell(row=rr, column=2, value=avg).number_format = _MONEY
        x.cell(row=rr, column=3,
               value=f"=Assumptions!$B$9/100*B{rr}*Assumptions!$B$8*78").number_format = _MONEY
        x.cell(row=rr, column=4,
               value=f"=Assumptions!$B$8*12*B{rr}*Assumptions!$B$9/100*12").number_format = _MONEY
        rr += 1
    x.cell(row=rr + 1, column=1,
           value="~All-Tier-2 Year-1 ≈ $233K (a light Tier-1 sprinkle → ~$236K).").font = _MUTED

    # --- Year-1 EBITDA by Mix (live P&L parallel to the commission table) ---
    e = wb.create_sheet("Year-1 EBITDA by Mix")
    for col, w in zip("ABCD", (30, 16, 16, 16)):
        e.column_dimensions[col].width = w
    _title(e, "Year-1 EBITDA by Tier 1 / Tier 2 mix")
    e.cell(row=5, column=1, value="Same 100/mo ramp (7,800 sub-months). Aggressive 'ceiling' scenario — see notes.").font = _MUTED
    e.cell(row=6, column=1, value="Tier-1 mix %").font = _BOLD
    for col, pct in zip("BCD", (0, 10, 20)):
        e[f"{col}6"] = pct
        e[f"{col}6"].font = _BOLD
    lines = [
        ("Avg MRR / sub", "=Assumptions!$B$5*{c}6/100+Assumptions!$B$6*(1-{c}6/100)", _MONEY),
        ("Revenue (Year-1)", "={c}7*Assumptions!$B$8*78", _MONEY),
        ("COGS (proc+infra % + SF1)", "={c}8*(Assumptions!$B$16+Assumptions!$B$17)/100+Assumptions!$B$18", _MONEY),
        ("Gross profit", "={c}8-{c}9", _MONEY),
        ("Sales commission (residual)", "={c}8*Assumptions!$B$9/100", _MONEY),
        ("Marketing", "=Assumptions!$B$19", _MONEY),
        ("Legal & accounting", "=Assumptions!$B$20", _MONEY),
        ("Software & AI tools", "=Assumptions!$B$21", _MONEY),
        ("Founder / ops comp", "=Assumptions!$B$22", _MONEY),
        ("Total operating expenses", "=SUM({c}11:{c}15)", _MONEY),
        ("EBITDA", "={c}10-{c}16", _MONEY),
        ("EBITDA margin", "=IF({c}8=0,0,{c}17/{c}8)", _PCT),
    ]
    for i, (label, formula, fmt) in enumerate(lines):
        row = 7 + i
        lc = e.cell(row=row, column=1, value=label)
        if label in ("Gross profit", "Total operating expenses", "EBITDA", "Revenue (Year-1)"):
            lc.font = _BOLD
        for ci, col in enumerate("BCD", start=2):
            cell = e.cell(row=row, column=ci, value=formula.format(c=col))
            cell.number_format = fmt
            if label in ("EBITDA",):
                cell.font = _BOLD

    # --- Ramp Scenarios (Conservative vs Base vs Aggressive close rate) ---
    rs = wb.create_sheet("Ramp Scenarios")
    for col, w in zip("ABCD", (34, 16, 16, 16)):
        rs.column_dimensions[col].width = w
    _title(rs, "Year-1 by ramp — Conservative vs Base vs Aggressive")
    rs.cell(row=5, column=1, value="Same 10% residual + opex assumptions; only the close rate changes. No-churn ramp (closes × 78 sub-months).").font = _MUTED
    rs.cell(row=6, column=1, value="Scenario").font = _WHITE
    rs.cell(row=6, column=1).fill = _NAVY_FILL
    for col, name in zip("BCD", ("Conservative", "Base", "Aggressive")):
        c = rs.cell(row=6, column="ABCD".index(col) + 1, value=name)
        c.font = _WHITE
        c.fill = _NAVY_FILL
    rs.cell(row=7, column=1, value="Closes per month").font = _BOLD
    for col, val in zip("BCD", (30, 60, 100)):
        cell = rs.cell(row=7, column="ABCD".index(col) + 1, value=val)
        cell.fill = _INPUT_FILL
    # Row map: 8=subs, 9=avgMRR, 10=revenue, 11=commission, 12=COGS, 13=gross,
    # 14=fixed opex, 15=total opex, 16=EBITDA, 17=margin.
    ramp_rows = [
        ("Year-1 subscribers (× 12)", "={c}7*12", "#,##0"),
        ("Avg MRR / sub", "=Assumptions!$B$13", _MONEY),
        ("Year-1 revenue", "={c}7*78*Assumptions!$B$13", _MONEY),
        ("Sales commission (residual)", "={c}10*Assumptions!$B$9/100", _MONEY),
        ("COGS (proc+infra % + SF1)", "={c}10*(Assumptions!$B$16+Assumptions!$B$17)/100+Assumptions!$B$18", _MONEY),
        ("Gross profit", "={c}10-{c}12", _MONEY),
        ("Fixed opex (mktg+legal+sw+founder)", "=Assumptions!$B$19+Assumptions!$B$20+Assumptions!$B$21+Assumptions!$B$22", _MONEY),
        ("Total operating expenses", "={c}11+{c}14", _MONEY),
        ("EBITDA", "={c}13-{c}15", _MONEY),
        ("EBITDA margin", "=IF({c}10=0,0,{c}16/{c}10)", _PCT),
    ]
    for i, (label, formula, fmt) in enumerate(ramp_rows):
        row_i = 8 + i
        lc = rs.cell(row=row_i, column=1, value=label)
        if label in ("Year-1 revenue", "Gross profit", "Total operating expenses", "EBITDA"):
            lc.font = _BOLD
        for ci, col in enumerate("BCD", start=2):
            cell = rs.cell(row=row_i, column=ci, value=formula.format(c=col))
            cell.number_format = fmt
            if label == "EBITDA":
                cell.font = _BOLD
    rs.cell(row=8 + len(ramp_rows) + 1, column=1,
            value="Conservative (30/mo) is the base case for a new, unproven product; Aggressive (100/mo) is the ceiling.").font = _MUTED

    # --- Legal ---
    lg = wb.create_sheet("Legal & Provenance")
    lg.column_dimensions["A"].width = 100
    lg.cell(row=1, column=1, value=ENTITY).font = Font(bold=True, size=13, color=_NAVY)
    for i, line in enumerate([
        "Internal planning model from user-supplied assumptions — illustrative only, not a forecast, "
        "offer of employment, or compensation guarantee.",
        "Residual commission is paid on active, collected subscriptions only. Consider a residual cap / "
        "step-down and a <90-day churn clawback to protect long-term margin.",
        f"Founder signature of provenance: {SIG}",
        f"© {date.today().year} {ENTITY}. All rights reserved. Confidential.",
    ], start=3):
        cell = lg.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        _watermark(ws)
    return wb


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "JHI_Sales_Commission_Model.xlsx"
    build().save(out)
    print(f"wrote {out}")
