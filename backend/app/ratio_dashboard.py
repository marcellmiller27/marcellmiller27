# JHI-SIG: 69M2705M | Unified ratio-dashboard sheet | JHI Research & Analytics Firm, Inc. (proprietary)
"""Aegira unified ratio-dashboard sheet — the canonical 6-section, 8-column
`Ratio_Dashboard` sheet embedded in every institutional workbook
(per-ticker + QoE + Deal X-Ray).

Schema (BOARD_MINUTES_2026-08-26.md §9.4):
  Sections (rows are grouped under these headings):
    1. Profitability
    2. Liquidity
    3. Solvency
    4. Efficiency
    5. Cash Flow
    6. Valuation

  Columns (8):
    A. Ratio            — canonical name
    B. Latest           — most-recent period value
    C. Prior period     — prior-year value
    D. 3-yr trend       — deterministic ▲ / ▼ / →
    E. Peer median      — supplied peer benchmark (None → n/m)
    F. Sector threshold — floor..ceiling text from SectorProfile
    G. Status           — Green / Amber / Red / N/M-sector / N/M-data
    H. Driver / note    — deterministic house-style explanation

Non-negotiables:
  - The N/M guard shows `N/M — sector` when the sector says the ratio is not
    relevant (banks don't get gross-margin flags), and `N/M — data` when
    the input is missing / negative-earnings (e.g. P/E with negative EPS).
  - Every row carries a numeric-format hint; the workbook renderer uses it.
  - Every deliverable is stamped with the SectorProfile version + JHI-SIG.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app import sector_profiles as sp
from app.sector_profiles import (
    ARR_GROWTH,
    ASSET_TURNOVER,
    CFO_TO_CAPEX,
    CFO_TO_NET_INCOME,
    CURRENT_RATIO,
    DEBT_TO_EQUITY,
    DSCR,
    DSO,
    EBITDA_MARGIN,
    EFFICIENCY_RATIO,
    EV_EBITDA,
    FCF_MARGIN,
    FFO_MULTIPLE,
    GROSS_MARGIN,
    INTEREST_COVERAGE,
    INVENTORY_TURNOVER,
    NET_MARGIN,
    NIM,
    OPERATING_MARGIN,
    PB,
    PE,
    PS,
    QUICK_RATIO,
    RD_INTENSITY,
    RECEIVABLES_TURNOVER,
    ROA,
    ROCE,
    ROE,
    Direction,
    SectorProfile,
    Status,
    status_for,
)

logger = logging.getLogger(__name__)

DASHBOARD_SCHEMA_VERSION = "2026.08.27-P1"

_HEADER_FILL = PatternFill("solid", fgColor="0C1F33")
_SECTION_FILL = PatternFill("solid", fgColor="EAEFF6")
_GREEN_FILL = PatternFill("solid", fgColor="D6F2E4")
_AMBER_FILL = PatternFill("solid", fgColor="FFF6D5")
_RED_FILL = PatternFill("solid", fgColor="F8D7D5")
_NM_FILL = PatternFill("solid", fgColor="F1F2F5")
_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)

_FMT_PCT = "0.0%"
_FMT_RATIO = "0.00"
_FMT_MULT = '0.00"x"'
_FMT_DAYS = "0"
_FMT_USD = "#,##0"


# --------------------------------------------------------------------------- #
# Section layout
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class DashboardRow:
    ratio_id: str
    display_name: str
    fmt: str = _FMT_RATIO


PROFITABILITY_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(GROSS_MARGIN, "Gross margin", _FMT_PCT),
    DashboardRow(OPERATING_MARGIN, "Operating margin", _FMT_PCT),
    DashboardRow(EBITDA_MARGIN, "EBITDA margin", _FMT_PCT),
    DashboardRow(NET_MARGIN, "Net margin", _FMT_PCT),
    DashboardRow(ROA, "Return on assets (ROA)", _FMT_PCT),
    DashboardRow(ROE, "Return on equity (ROE)", _FMT_PCT),
    DashboardRow(ROCE, "Return on capital employed (ROCE)", _FMT_PCT),
)

LIQUIDITY_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(CURRENT_RATIO, "Current ratio"),
    DashboardRow(QUICK_RATIO, "Quick ratio"),
)

SOLVENCY_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(DEBT_TO_EQUITY, "Debt / equity"),
    DashboardRow(INTEREST_COVERAGE, "Interest coverage", _FMT_MULT),
    DashboardRow(DSCR, "Debt-service coverage (DSCR)", _FMT_MULT),
)

EFFICIENCY_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(ASSET_TURNOVER, "Asset turnover"),
    DashboardRow(INVENTORY_TURNOVER, "Inventory turnover", _FMT_MULT),
    DashboardRow(RECEIVABLES_TURNOVER, "Receivables turnover", _FMT_MULT),
    DashboardRow(DSO, "Days sales outstanding (DSO)", _FMT_DAYS),
)

CASHFLOW_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(CFO_TO_CAPEX, "CFO / CapEx", _FMT_MULT),
    DashboardRow(CFO_TO_NET_INCOME, "CFO / Net income"),
    DashboardRow(FCF_MARGIN, "Free-cash-flow margin", _FMT_PCT),
)

VALUATION_ROWS: tuple[DashboardRow, ...] = (
    DashboardRow(PE, "Price / Earnings (P/E)", _FMT_MULT),
    DashboardRow(PB, "Price / Book (P/B)", _FMT_MULT),
    DashboardRow(PS, "Price / Sales (P/S)", _FMT_MULT),
    DashboardRow(EV_EBITDA, "EV / EBITDA", _FMT_MULT),
)

SECTOR_KPI_ROWS_BY_ID: dict[str, DashboardRow] = {
    NIM: DashboardRow(NIM, "Net interest margin (NIM)", _FMT_PCT),
    EFFICIENCY_RATIO: DashboardRow(EFFICIENCY_RATIO, "Efficiency ratio", _FMT_PCT),
    ARR_GROWTH: DashboardRow(ARR_GROWTH, "ARR growth", _FMT_PCT),
    RD_INTENSITY: DashboardRow(RD_INTENSITY, "R&D intensity", _FMT_PCT),
    FFO_MULTIPLE: DashboardRow(FFO_MULTIPLE, "Price / FFO", _FMT_MULT),
}


@dataclass(frozen=True)
class SectionSpec:
    title: str
    rows: tuple[DashboardRow, ...]


BASE_SECTIONS: tuple[SectionSpec, ...] = (
    SectionSpec("Profitability", PROFITABILITY_ROWS),
    SectionSpec("Liquidity", LIQUIDITY_ROWS),
    SectionSpec("Solvency", SOLVENCY_ROWS),
    SectionSpec("Efficiency", EFFICIENCY_ROWS),
    SectionSpec("Cash Flow", CASHFLOW_ROWS),
    SectionSpec("Valuation", VALUATION_ROWS),
)


# --------------------------------------------------------------------------- #
# Dashboard input / output types
# --------------------------------------------------------------------------- #
@dataclass
class RatioSeries:
    """One ratio's data feed into the dashboard."""

    latest: float | None = None
    prior: float | None = None
    history: list[float] = field(default_factory=list)  # oldest → newest (≥ 3 pts for trend)
    peer_median: float | None = None
    # Optional data-quality bit — when True, forces N/M-data (e.g. negative EPS
    # for P/E, negative equity for ROE).  Defaults to None (respected only when
    # explicitly set).
    force_nm_data: bool = False


@dataclass
class DashboardInputs:
    """The full data payload for one dashboard sheet."""

    entity_name: str
    period_label: str
    sector: sp.Sector | str | None
    ratios: dict[str, RatioSeries] = field(default_factory=dict)


@dataclass
class DashboardRowResult:
    ratio_id: str
    display_name: str
    section: str
    latest: float | None
    prior: float | None
    trend_arrow: str
    peer_median: float | None
    threshold_text: str
    status: Status
    driver_note: str
    fmt: str
    is_relevant: bool


@dataclass
class DashboardResult:
    entity_name: str
    period_label: str
    sector_name: str
    schema_version: str
    rows: list[DashboardRowResult] = field(default_factory=list)
    summary_flags: list[str] = field(default_factory=list)  # material red / amber cross-section

    def status_counts(self) -> dict[str, int]:
        counts = {"green": 0, "amber": 0, "red": 0, "nm-sector": 0, "nm-data": 0}
        for r in self.rows:
            if r.status == Status.GREEN:
                counts["green"] += 1
            elif r.status == Status.AMBER:
                counts["amber"] += 1
            elif r.status == Status.RED:
                counts["red"] += 1
            elif r.status == Status.NM_SECTOR:
                counts["nm-sector"] += 1
            else:
                counts["nm-data"] += 1
        return counts


# --------------------------------------------------------------------------- #
# Trend arrow + threshold text
# --------------------------------------------------------------------------- #
def _trend_arrow(history: list[float]) -> str:
    """Deterministic trend arrow from a history series (oldest → newest)."""
    clean = [float(v) for v in (history or []) if v is not None]
    if len(clean) < 2:
        return "—"
    lo = clean[0]
    hi = clean[-1]
    if lo == 0 and hi == 0:
        return "→"
    delta = hi - lo
    scale = max(abs(lo), abs(hi), 1e-9)
    pct = abs(delta) / scale
    if pct < 0.02:
        return "→"
    return "▲" if delta > 0 else "▼"


def _threshold_text(threshold: sp.RatioThreshold | None, fmt: str) -> str:
    if threshold is None:
        return "—"
    lo, hi = threshold.good_floor, threshold.good_ceiling
    if fmt == _FMT_PCT:
        left, right = f"{lo * 100:.1f}%", f"{hi * 100:.1f}%"
    elif fmt == _FMT_MULT:
        left, right = f"{lo:.2f}×", f"{hi:.2f}×"
    elif fmt == _FMT_DAYS:
        left, right = f"{lo:.0f}d", f"{hi:.0f}d"
    else:
        left, right = f"{lo:.2f}", f"{hi:.2f}"
    if threshold.direction == Direction.HIGHER:
        return f"≥ {right} good"
    if threshold.direction == Direction.LOWER:
        return f"≤ {left} good"
    return f"{left} – {right}"


def _driver_note(
    row: DashboardRow,
    series: RatioSeries,
    profile: SectorProfile,
    threshold: sp.RatioThreshold | None,
    status: Status,
) -> str:
    """Deterministic house-style driver note."""
    if status == Status.NM_SECTOR:
        return f"Not a defining ratio for {profile.name}."
    if status == Status.NM_DATA:
        if row.ratio_id in {PE, EV_EBITDA} and series.latest is None:
            return "N/M — negative or missing earnings; use peer-relative valuation instead."
        return "N/M — required input missing."
    trend = _trend_arrow(series.history)
    peer = (
        f"peer median {_format_value(series.peer_median, row.fmt)}"
        if series.peer_median is not None else "peer median n/a"
    )
    thresh = _threshold_text(threshold, row.fmt) if threshold else "no sector band"
    return (
        f"Latest {_format_value(series.latest, row.fmt)}; trend {trend}; "
        f"sector band {thresh}; {peer}."
    )


def _format_value(value: float | None, fmt: str) -> str:
    if value is None:
        return "n/a"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "n/a"
    if fmt == _FMT_PCT:
        return f"{v * 100:.1f}%"
    if fmt == _FMT_MULT:
        return f"{v:.2f}×"
    if fmt == _FMT_DAYS:
        return f"{v:.0f}d"
    return f"{v:,.2f}"


# --------------------------------------------------------------------------- #
# Compute
# --------------------------------------------------------------------------- #
def compute_dashboard(inputs: DashboardInputs) -> DashboardResult:
    profile = sp.get_profile(inputs.sector)
    result = DashboardResult(
        entity_name=inputs.entity_name,
        period_label=inputs.period_label,
        sector_name=profile.name,
        schema_version=DASHBOARD_SCHEMA_VERSION,
    )

    sections = list(BASE_SECTIONS)
    # Append sector KPI rows to a synthetic "Sector KPIs" section if any exist
    kpi_rows = tuple(
        SECTOR_KPI_ROWS_BY_ID[k] for k in profile.sector_kpis if k in SECTOR_KPI_ROWS_BY_ID
    )
    if kpi_rows:
        sections.append(SectionSpec(f"Sector KPIs — {profile.name}", kpi_rows))

    material_reds: list[str] = []

    for section in sections:
        for row in section.rows:
            is_relevant = profile.is_relevant(row.ratio_id)
            threshold = profile.threshold_for(row.ratio_id) if is_relevant else None
            series = inputs.ratios.get(row.ratio_id) or RatioSeries()
            if not is_relevant:
                status = Status.NM_SECTOR
            elif series.force_nm_data:
                status = Status.NM_DATA
            elif series.latest is None:
                status = Status.NM_DATA
            else:
                status = status_for(series.latest, threshold)

            row_result = DashboardRowResult(
                ratio_id=row.ratio_id,
                display_name=row.display_name,
                section=section.title,
                latest=series.latest,
                prior=series.prior,
                trend_arrow=_trend_arrow(series.history),
                peer_median=series.peer_median,
                threshold_text=_threshold_text(threshold, row.fmt) if threshold else "—",
                status=status,
                driver_note=_driver_note(row, series, profile, threshold, status),
                fmt=row.fmt,
                is_relevant=is_relevant,
            )
            result.rows.append(row_result)
            if status == Status.RED:
                material_reds.append(f"{row.display_name} — {row_result.driver_note}")

    result.summary_flags = material_reds
    return result


# --------------------------------------------------------------------------- #
# Excel rendering
# --------------------------------------------------------------------------- #
def _status_fill(status: Status) -> PatternFill:
    return {
        Status.GREEN: _GREEN_FILL,
        Status.AMBER: _AMBER_FILL,
        Status.RED: _RED_FILL,
        Status.NM_SECTOR: _NM_FILL,
        Status.NM_DATA: _NM_FILL,
    }.get(status, _NM_FILL)


def _status_label(status: Status) -> str:
    return {
        Status.GREEN: "Green",
        Status.AMBER: "Amber",
        Status.RED: "Red",
        Status.NM_SECTOR: "N/M — sector",
        Status.NM_DATA: "N/M — data",
    }[status]


def render_dashboard_sheet(ws: Worksheet, result: DashboardResult) -> None:
    """Render the unified dashboard on the provided worksheet.

    Assumes the caller has already added a title / brand band above row 1 (the
    dashboard writes starting at row 1 by default so it can be used as either
    a standalone sheet or embedded into a bigger workbook via offset rows).
    """
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 72

    row = 1
    ws.cell(row=row, column=1, value=f"Aegira · Ratio Dashboard — {result.entity_name}").font = Font(bold=True, size=14)
    ws.cell(row=row, column=8, value=f"Sector: {result.sector_name}").font = _MUTED
    row += 1
    ws.cell(row=row, column=1, value=f"Period: {result.period_label}  ·  Schema {result.schema_version}").font = _MUTED
    ws.cell(row=row, column=8, value=f"Prepared {date.today().isoformat()}").font = _MUTED
    row += 2

    # Header
    headers = [
        "Ratio", "Latest", "Prior", "Trend", "Peer median",
        "Sector band", "Status", "Driver / note",
    ]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _WHITE_BOLD
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left" if col in (1, 8) else "center")
    row += 1

    current_section: str | None = None
    for r in result.rows:
        if r.section != current_section:
            current_section = r.section
            for col in range(1, 9):
                c = ws.cell(row=row, column=col, value=r.section if col == 1 else "")
                c.fill = _SECTION_FILL
                if col == 1:
                    c.font = _BOLD
            row += 1

        name_cell = ws.cell(row=row, column=1, value=r.display_name)
        name_cell.alignment = Alignment(wrap_text=True)

        latest_cell = ws.cell(row=row, column=2, value=r.latest if r.latest is not None else "n/a")
        prior_cell = ws.cell(row=row, column=3, value=r.prior if r.prior is not None else "n/a")
        if r.latest is not None:
            latest_cell.number_format = r.fmt
        if r.prior is not None:
            prior_cell.number_format = r.fmt

        trend_cell = ws.cell(row=row, column=4, value=r.trend_arrow)
        trend_cell.alignment = Alignment(horizontal="center")

        peer_cell = ws.cell(row=row, column=5,
                            value=r.peer_median if r.peer_median is not None else "n/a")
        if r.peer_median is not None:
            peer_cell.number_format = r.fmt

        band_cell = ws.cell(row=row, column=6, value=r.threshold_text)
        band_cell.alignment = Alignment(horizontal="center")

        status_cell = ws.cell(row=row, column=7, value=_status_label(r.status))
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.fill = _status_fill(r.status)
        status_cell.font = _BOLD if r.status in (Status.GREEN, Status.RED) else Font()

        note_cell = ws.cell(row=row, column=8, value=r.driver_note)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        note_cell.font = _MUTED

        row += 1

    # Summary block below
    row += 1
    counts = result.status_counts()
    ws.cell(row=row, column=1, value="Status roll-up").font = _BOLD
    row += 1
    for key, label, fill in (
        ("green", "Green", _GREEN_FILL),
        ("amber", "Amber", _AMBER_FILL),
        ("red", "Red", _RED_FILL),
        ("nm-sector", "N/M — sector", _NM_FILL),
        ("nm-data", "N/M — data", _NM_FILL),
    ):
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill
        c.alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2, value=counts[key])
        row += 1

    if result.summary_flags:
        row += 1
        ws.cell(row=row, column=1, value="Red flags").font = _BOLD
        row += 1
        for flag in result.summary_flags:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=flag)
            c.fill = _RED_FILL
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    # Numeric conditional formatting on Latest column for the green/amber/red
    # signal even when the sheet is filtered / sorted downstream.
    _apply_conditional_formatting(ws, first_data_row=5, last_data_row=row - 1)


def _apply_conditional_formatting(ws: Worksheet, first_data_row: int, last_data_row: int) -> None:
    if last_data_row <= first_data_row:
        return
    range_str = f"G{first_data_row}:G{last_data_row}"
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="equal", formula=['"Green"'], fill=_GREEN_FILL),
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="equal", formula=['"Amber"'], fill=_AMBER_FILL),
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="equal", formula=['"Red"'], fill=_RED_FILL),
    )


__all__ = [
    "BASE_SECTIONS",
    "DASHBOARD_SCHEMA_VERSION",
    "DashboardInputs",
    "DashboardResult",
    "DashboardRowResult",
    "RatioSeries",
    "SectionSpec",
    "compute_dashboard",
    "render_dashboard_sheet",
]
