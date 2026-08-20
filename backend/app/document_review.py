# JHI-SIG: 69M2705M | Document Review module | JHI Research & Analytics Firm, Inc. (proprietary)
"""Document Review engine — deterministic extraction + risk analysis of uploaded
acquisition documents (tax returns, P&L, balance sheets, bank statements).

The engine is a pure function of the uploaded bytes: it extracts what it can
(CSV/XLSX tabular figures, PDF text), runs a battery of deterministic risk / fraud
indicators, and returns a 0-100 risk score, a list of flags, and a list of
diligence questions. It NEVER fabricates figures — if a file cannot be parsed, it
returns a graceful ``manual_review_required`` result with no invented numbers, and
it NEVER raises on malformed input.

Risk phrasing and the clamp/threshold style are shared with the Financial
Diligence Suite (``app.financial_diligence``) so the two modules speak the same
QoE language (proof-of-cash, add-back scrutiny, revenue-quality, concentration).
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field

from app.document_review_models import DOC_TYPE_LABELS, DocType, ReviewStatus
from app.financial_diligence import _clamp

# Shared decision-support disclaimer (mirrors the Financial Diligence Suite).
DISCLAIMER = (
    "Software-generated document analysis — decision-support only. This is NOT an "
    "audit, review, compilation, or CPA opinion, and no assurance is expressed. "
    "Extraction is automated and may be incomplete; verify every figure against the "
    "source document before making an offer or extending credit."
)

# --- File-type policy --------------------------------------------------------
ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx"}
ALLOWED_CONTENT_TYPES = {
    "application/pdf",
    "text/csv",
    "application/csv",
    "text/plain",  # some browsers send CSV as text/plain
    "application/vnd.ms-excel",  # legacy CSV/XLS content-type used by some browsers
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",  # generic fallback; extension is the hard gate
}
MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # 25 MB

_NUMBER_RE = re.compile(r"\(?-?\$?\s?\d{1,3}(?:,\d{3})+(?:\.\d+)?\)?|\(?-?\$?\s?\d+(?:\.\d+)?\)?")
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")
_MONTHS = (
    "january february march april may june july august september october "
    "november december jan feb mar apr jun jul aug sep sept oct nov dec"
).split()


@dataclass
class Extraction:
    """What we managed to pull out of a file (best-effort, never fabricated)."""

    text: str = ""
    numbers: list[float] = field(default_factory=list)
    # The longest numeric series found (a candidate trend line, e.g. revenue by period).
    primary_series: list[float] = field(default_factory=list)
    period_years: list[int] = field(default_factory=list)
    parsed: bool = False


def _to_number(token: str) -> float | None:
    """Parse a currency/number token like ``$1,200.50`` or ``(4,000)`` (negative)."""
    t = token.strip()
    if not t:
        return None
    negative = t.startswith("(") and t.endswith(")")
    t = t.strip("()").replace("$", "").replace(",", "").replace(" ", "")
    if t in ("", "-", "."):
        return None
    try:
        value = float(t)
    except ValueError:
        return None
    return -value if negative else value


def _numbers_from_text(text: str) -> list[float]:
    out: list[float] = []
    for match in _NUMBER_RE.findall(text):
        value = _to_number(match)
        # Ignore bare small integers that are almost always row/line numbers or years.
        if value is not None:
            out.append(value)
    return out


def _years_from_text(text: str) -> list[int]:
    return [int(m.group(0)) for m in _YEAR_RE.finditer(text)]


def _pick_primary_series(rows_numbers: list[list[float]]) -> list[float]:
    """Choose the candidate trend line: the numeric row with >=3 points carrying the
    largest total magnitude (financial figures dwarf year-label rows like 2021/2022)."""
    candidates = [r for r in rows_numbers if len(r) >= 3]
    if not candidates:
        return max(rows_numbers, key=len, default=[])
    return max(candidates, key=lambda r: sum(abs(n) for n in r))


def _extract_csv(raw: bytes) -> Extraction:
    ex = Extraction()
    try:
        decoded = raw.decode("utf-8-sig", errors="replace")
    except Exception:  # noqa: BLE001 - decoding must never crash the request
        return ex
    ex.text = decoded
    try:
        rows = list(csv.reader(io.StringIO(decoded)))
    except Exception:  # noqa: BLE001
        rows = []
    all_numbers: list[float] = []
    rows_numbers: list[list[float]] = []
    for row in rows:
        row_numbers = [n for n in (_to_number(c) for c in row) if n is not None]
        all_numbers.extend(row_numbers)
        rows_numbers.append(row_numbers)
    ex.numbers = all_numbers
    ex.primary_series = _pick_primary_series(rows_numbers)
    ex.period_years = sorted(set(_years_from_text(decoded)))
    ex.parsed = bool(all_numbers) or bool(decoded.strip())
    return ex


def _extract_xlsx(raw: bytes) -> Extraction:
    ex = Extraction()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - a corrupt/locked workbook must not crash
        return ex
    all_numbers: list[float] = []
    rows_numbers: list[list[float]] = []
    text_parts: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                row_numbers: list[float] = []
                for cell in row:
                    if isinstance(cell, bool):
                        continue
                    if isinstance(cell, (int, float)):
                        row_numbers.append(float(cell))
                    elif cell is not None:
                        text_parts.append(str(cell))
                all_numbers.extend(row_numbers)
                rows_numbers.append(row_numbers)
    except Exception:  # noqa: BLE001
        pass
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    ex.text = " ".join(text_parts)
    ex.numbers = all_numbers
    ex.primary_series = _pick_primary_series(rows_numbers)
    ex.period_years = sorted(set(_years_from_text(ex.text)))
    ex.parsed = bool(all_numbers) or bool(text_parts)
    return ex


def _extract_pdf(raw: bytes) -> Extraction:
    ex = Extraction()
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        parts = [(page.extract_text() or "") for page in reader.pages]
    except Exception:  # noqa: BLE001 - scanned/encrypted/corrupt PDFs must not crash
        return ex
    text = "\n".join(parts).strip()
    ex.text = text
    ex.numbers = _numbers_from_text(text)
    ex.primary_series = ex.numbers  # PDF text has no reliable row structure
    ex.period_years = sorted(set(_years_from_text(text)))
    ex.parsed = bool(text)
    return ex


def extract(filename: str, content_type: str, raw: bytes) -> Extraction:
    """Dispatch on extension (the trusted signal) to the right extractor."""
    ext = _extension(filename)
    if ext == ".csv":
        return _extract_csv(raw)
    if ext == ".xlsx":
        return _extract_xlsx(raw)
    if ext == ".pdf":
        return _extract_pdf(raw)
    # Unknown extension should have been rejected upstream; degrade gracefully.
    return Extraction()


def _extension(filename: str) -> str:
    name = (filename or "").lower().strip()
    dot = name.rfind(".")
    return name[dot:] if dot != -1 else ""


# --- Risk indicators ---------------------------------------------------------
@dataclass
class Indicator:
    triggered: bool
    weight: int
    flag: str
    questions: list[str] = field(default_factory=list)


def _round_number_ratio(numbers: list[float]) -> tuple[float, int]:
    """Fraction of material figures that are suspiciously round (multiples of 1,000)."""
    material = [n for n in numbers if abs(n) >= 1000]
    if len(material) < 5:
        return 0.0, len(material)
    rounded = sum(1 for n in material if n % 1000 == 0)
    return rounded / len(material), len(material)


def _declining(series: list[float]) -> bool:
    """A material, sustained decline across a numeric series (>=3 points)."""
    trend = [n for n in series if n != 0]
    if len(trend) < 3:
        return False
    first, last = trend[0], trend[-1]
    if first <= 0:
        return False
    monotonic = all(b <= a for a, b in zip(trend, trend[1:]))
    dropped = (last - first) / first <= -0.10
    return monotonic and dropped


def _missing_year_gap(years: list[int]) -> bool:
    if len(years) < 2:
        return False
    ordered = sorted(set(years))
    return (ordered[-1] - ordered[0]) > (len(ordered) - 1)


def _indicators(doc_type: DocType, ex: Extraction) -> list[Indicator]:
    lower = ex.text.lower()
    numbers = ex.numbers
    inds: list[Indicator] = []

    has_negative = any(n < 0 for n in numbers)
    inds.append(
        Indicator(
            triggered=has_negative,
            weight=12,
            flag="Negative figures detected — confirm they are true losses/credits, not sign errors.",
            questions=["Explain each negative line item and whether it recurs."],
        )
    )

    inds.append(
        Indicator(
            triggered=_declining(ex.primary_series),
            weight=25,
            flag="Primary figure series is declining across periods — earnings trend is deteriorating.",
            questions=[
                "What is driving the period-over-period decline, and is it structural or one-time?",
            ],
        )
    )

    inds.append(
        Indicator(
            triggered=_missing_year_gap(ex.period_years),
            weight=20,
            flag="Reporting periods are non-contiguous (a year appears to be missing) — request the full series.",
            questions=["Provide the missing period(s); why was that year omitted?"],
        )
    )

    ratio, material_count = _round_number_ratio(numbers)
    inds.append(
        Indicator(
            triggered=ratio >= 0.6 and material_count >= 5,
            weight=18,
            flag=(
                f"{ratio * 100:.0f}% of material figures are round thousands — "
                "values look estimated rather than actual."
            ),
            questions=["Are these figures actuals tied to source ledgers, or estimates/plugs?"],
        )
    )

    cash_terms = ("cash basis", "cash-basis")
    accrual_terms = ("accrual", "accounts receivable", "accounts payable", "deferred revenue")
    mismatch = any(t in lower for t in cash_terms) and any(t in lower for t in accrual_terms)
    inds.append(
        Indicator(
            triggered=mismatch,
            weight=15,
            flag="Cash-basis and accrual terms both present — confirm the accounting basis and reconcile.",
            questions=["Are the statements cash or accrual basis? Provide a reconciliation between them."],
        )
    )

    addback_terms = ("add-back", "addback", "add back", "discretionary", "owner", "personal", "one-time", "non-recurring", "nonrecurring")
    has_addbacks = any(t in lower for t in addback_terms)
    inds.append(
        Indicator(
            triggered=has_addbacks,
            weight=12,
            flag="Owner add-back / discretionary language present — each add-back needs documentary support.",
            questions=["Itemize every add-back with third-party support; which continue post-close?"],
        )
    )

    # Doc-type-specific proof-of-cash cue for bank statements.
    if doc_type == DocType.BANK_STATEMENTS:
        inds.append(
            Indicator(
                triggered=True,
                weight=6,
                flag="Bank statements uploaded — tie total deposits to reported revenue (proof-of-cash).",
                questions=["Do total deposits reconcile to reported revenue within materiality?"],
            )
        )

    return inds


_BASELINE_QUESTIONS: dict[DocType, list[str]] = {
    DocType.TAX_RETURNS: [
        "Do the tax returns reconcile to the P&L (book-to-tax differences explained)?",
        "Are all schedules and K-1s included for every period?",
    ],
    DocType.PNL: [
        "Is revenue recognized consistently period-over-period?",
        "What is the gross-margin trend and what drives any change?",
    ],
    DocType.BALANCE_SHEET: [
        "Are there debt-like items (deferred revenue, capital leases, unpaid taxes) to treat as debt at close?",
        "Is net working capital sufficient, and what is the target peg?",
    ],
    DocType.BANK_STATEMENTS: [
        "Do deposits tie to invoiced revenue (proof-of-cash)?",
        "Are there unexplained transfers, related-party flows, or commingled personal transactions?",
    ],
}


def _band(score: int) -> str:
    if score >= 67:
        return "High"
    if score >= 34:
        return "Medium"
    return "Low"


@dataclass
class Analysis:
    status: ReviewStatus
    risk_score: int | None
    risk_band: str
    summary: str
    flags: list[str]
    questions: list[str]


def analyze(doc_type: DocType, filename: str, content_type: str, raw: bytes) -> Analysis:
    """Extract + score one uploaded document. Never raises; never fabricates figures."""
    ex = extract(filename, content_type, raw)
    label = DOC_TYPE_LABELS.get(doc_type, str(doc_type))

    if not ex.parsed or (not ex.numbers and len(ex.text.strip()) < 20):
        # Nothing usable was extracted (e.g. scanned/encrypted PDF or empty file).
        return Analysis(
            status=ReviewStatus.MANUAL_REVIEW_REQUIRED,
            risk_score=None,
            risk_band="",
            summary=(
                f"{label} received but could not be parsed automatically "
                "(likely scanned/encrypted or an unsupported layout). Routed for manual review — "
                "no figures were inferred."
            ),
            flags=["Automated extraction failed — analyst must review the source document manually."],
            questions=_BASELINE_QUESTIONS.get(doc_type, []),
        )

    indicators = _indicators(doc_type, ex)
    triggered = [i for i in indicators if i.triggered]

    # Base risk reflects an unverified third-party document; each indicator adds weight.
    score = 8.0 + sum(i.weight for i in triggered)
    risk_score = int(_clamp(score, 0, 100))

    flags = [i.flag for i in triggered]
    questions: list[str] = list(_BASELINE_QUESTIONS.get(doc_type, []))
    for ind in triggered:
        for q in ind.questions:
            if q not in questions:
                questions.append(q)

    summary = (
        f"{label}: extracted {len(ex.numbers)} figures across "
        f"{len(ex.period_years) or 'unlabeled'} period(s); "
        f"{len(triggered)} risk indicator(s) triggered → {_band(risk_score)} risk ({risk_score}/100)."
    )

    return Analysis(
        status=ReviewStatus.ANALYZED,
        risk_score=risk_score,
        risk_band=_band(risk_score),
        summary=summary,
        flags=flags,
        questions=questions,
    )
