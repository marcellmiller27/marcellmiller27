"""Client-ready multi-year Excel workbook of a company's SEC EDGAR financials.

Premium (T1 Enterprise / T2 Professional) deliverable: a multi-year financials
table plus native Excel charts (revenue & net income; margins). Data is public-
domain SEC XBRL (freely redistributable). Pure function -> xlsx bytes.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.edgar_models import EdgarHistory

NAVY = "0C1F33"
GOLD = "9A6B12"
LIGHT = "EAEFF6"
WHITE = "FFFFFF"

_thin = Side(style="thin", color="C9D3DF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_DOLLAR_ROWS = [
    ("Revenue", "revenue"),
    ("Gross profit", "gross_profit"),
    ("Operating income", "operating_income"),
    ("Net income", "net_income"),
    ("Total assets", "total_assets"),
    ("Stockholders' equity", "stockholders_equity"),
]
_MARGIN_ROWS = [
    ("Gross margin %", "gross_margin"),
    ("Operating margin %", "operating_margin"),
    ("Net margin %", "net_margin"),
]


def company_workbook(hist: EdgarHistory, branded: bool = False) -> bytes:
    """Return xlsx bytes: multi-year financials table + revenue/margin charts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{hist.entity_name} ({hist.ticker})"
    ws["A1"].font = Font(bold=True, size=15, color=NAVY)
    ws.merge_cells("A2:H2")
    ws["A2"] = f"CIK {hist.cik} · annual (SEC EDGAR) · {hist.currency} · values in $B unless %"
    ws["A2"].font = Font(italic=True, size=9, color="5A6B7D")
    if branded:
        ws.merge_cells("A3:H3")
        ws["A3"] = "John Henry Investments — prepared for client review"
        ws["A3"].font = Font(bold=True, size=9, color=GOLD)

    years = hist.years
    if not years:
        ws["A5"] = "No annual XBRL data available for this filer."
        ws["A5"].font = Font(italic=True, color="5A6B7D")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    n = len(years)
    hdr = 5

    def style_header_cell(cell) -> None:
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # Header row: Metric | FY.. 
    style_header_cell(ws.cell(hdr, 1, "Metric"))
    ws.cell(hdr, 1).alignment = Alignment(horizontal="left", vertical="center")
    for i, yr in enumerate(years):
        style_header_cell(ws.cell(hdr, 2 + i, f"FY{yr.fiscal_year}"))

    def write_rows(start: int, rows, transform, alt_base: int) -> None:
        for ri, (label, attr) in enumerate(rows):
            r = start + ri
            lc = ws.cell(r, 1, label)
            lc.font = Font(bold=True, size=10, color=NAVY)
            lc.border = _BORDER
            for i, yr in enumerate(years):
                v = getattr(yr, attr)
                cell = ws.cell(r, 2 + i, transform(v) if v is not None else None)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="right")
                cell.border = _BORDER
            if (ri + alt_base) % 2:
                for c in range(1, n + 2):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)

    dollar_start = hdr + 1
    write_rows(dollar_start, _DOLLAR_ROWS, lambda v: round(v / 1e9, 2), 0)
    margin_start = dollar_start + len(_DOLLAR_ROWS) + 1
    style_header_cell(ws.cell(margin_start - 1, 1, "Margins (%)"))
    ws.cell(margin_start - 1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for i in range(n):
        style_header_cell(ws.cell(margin_start - 1, 2 + i, f"FY{years[i].fiscal_year}"))
    write_rows(margin_start, _MARGIN_ROWS, lambda v: round(v * 100, 1), 0)

    # --- Charts ---
    cats = Reference(ws, min_col=2, max_col=1 + n, min_row=hdr, max_row=hdr)
    chart_row = margin_start + len(_MARGIN_ROWS) + 2

    c1 = LineChart()
    c1.title = "Revenue & Net Income ($B)"
    c1.y_axis.title = "$B"
    c1.x_axis.title = "Fiscal Year"
    c1.height, c1.width = 8, 16
    for offset, (_label, _attr) in enumerate(_DOLLAR_ROWS):
        if _attr in ("revenue", "net_income"):
            ref = Reference(ws, min_col=1, max_col=1 + n, min_row=dollar_start + offset,
                            max_row=dollar_start + offset)
            c1.add_data(ref, titles_from_data=True, from_rows=True)
    c1.set_categories(cats)
    ws.add_chart(c1, f"A{chart_row}")

    c2 = LineChart()
    c2.title = "Margins (%)"
    c2.y_axis.title = "%"
    c2.x_axis.title = "Fiscal Year"
    c2.height, c2.width = 8, 16
    mref = Reference(ws, min_col=1, max_col=1 + n, min_row=margin_start,
                     max_row=margin_start + len(_MARGIN_ROWS) - 1)
    c2.add_data(mref, titles_from_data=True, from_rows=True)
    c2.set_categories(cats)
    ws.add_chart(c2, f"A{chart_row + 17}")

    footer = chart_row + 34
    ws.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=1 + n)
    ws.cell(row=footer, column=1, value=(
        f"Source: {hist.source}. {hist.disclaimer} Public SEC data; verify against filings."
    )).font = Font(italic=True, size=8, color="5A6B7D")

    ws.column_dimensions[get_column_letter(1)].width = 22
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12
    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
