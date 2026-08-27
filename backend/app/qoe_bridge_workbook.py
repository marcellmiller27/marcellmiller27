# JHI-SIG: 69M2705M | EBITDA Bridge workbook sheet | JHI Research & Analytics Firm, Inc. (proprietary)
"""EBITDA_Bridge sheet renderer — writes the QoE bridge (§11.4) into a workbook.

Consumed by the QoE workbook, the per-ticker institutional workbook (for
private targets), and the Deal X-Ray workbook. The waterfall is optional
(requires matplotlib + Pillow, already dependencies).
"""

from __future__ import annotations

import logging
from datetime import date
from io import BytesIO

from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.qoe_bridge import EBITDABridge, EvidenceGrade, Recurrence

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="0C1F33")
_SECTION_FILL = PatternFill("solid", fgColor="EAEFF6")
_GREEN_FILL = PatternFill("solid", fgColor="D6F2E4")
_AMBER_FILL = PatternFill("solid", fgColor="FFF6D5")
_RED_FILL = PatternFill("solid", fgColor="F8D7D5")
_MUTED_FILL = PatternFill("solid", fgColor="F1F2F5")
_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)

_USD = "#,##0"


def _grade_fill(grade: EvidenceGrade) -> PatternFill:
    return {
        EvidenceGrade.A: _GREEN_FILL,
        EvidenceGrade.B: _AMBER_FILL,
        EvidenceGrade.C: _RED_FILL,
    }.get(grade, _MUTED_FILL)


def render_ebitda_bridge_sheet(ws: Worksheet, bridge: EBITDABridge) -> None:
    ws.column_dimensions["A"].width = 4                 # #
    ws.column_dimensions["B"].width = 42                # Category
    ws.column_dimensions["C"].width = 16                # Seller-view
    ws.column_dimensions["D"].width = 16                # Buyer-view
    ws.column_dimensions["E"].width = 12                # Recurrence
    ws.column_dimensions["F"].width = 10                # Evidence grade
    ws.column_dimensions["G"].width = 42                # Source citation
    ws.column_dimensions["H"].width = 60                # Driver note

    row = 1
    ws.cell(row=row, column=1, value=f"Aegira · EBITDA Bridge — {bridge.business_name}").font = Font(bold=True, size=14)
    ws.cell(row=row, column=8, value="Decision-support / research — not an audit").font = _MUTED
    row += 1
    ws.cell(row=row, column=1,
            value=f"Period: {bridge.period_label}  ·  Prepared {date.today().isoformat()}").font = _MUTED
    row += 2

    # Reported EBITDA
    ws.cell(row=row, column=2, value="Reported EBITDA").font = _BOLD
    ws.cell(row=row, column=3, value=bridge.reported_ebitda).number_format = _USD
    ws.cell(row=row, column=4, value=bridge.reported_ebitda).number_format = _USD
    row += 2

    # Header row
    headers = ["#", "Adjustment category", "Seller-view $", "Buyer-view $",
               "Recurrence", "Evidence", "Source", "Driver note"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _WHITE_BOLD
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left")
    row += 1

    first_data_row = row

    for i, oc in enumerate(bridge.adjustments, start=1):
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=2, value=oc.category_name).alignment = Alignment(wrap_text=True)
        c_seller = ws.cell(row=row, column=3, value=oc.seller_amount)
        c_buyer = ws.cell(row=row, column=4, value=oc.buyer_amount)
        c_seller.number_format = _USD
        c_buyer.number_format = _USD

        recurrence_cell = ws.cell(row=row, column=5, value=oc.recurrence.value)
        recurrence_cell.alignment = Alignment(horizontal="center")

        grade_cell = ws.cell(row=row, column=6, value=oc.evidence_grade.value)
        grade_cell.alignment = Alignment(horizontal="center")
        grade_cell.fill = _grade_fill(oc.evidence_grade)
        grade_cell.font = _BOLD

        ws.cell(row=row, column=7, value=oc.reference).alignment = Alignment(wrap_text=True)
        note_cell = ws.cell(row=row, column=8, value=oc.driver_note)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        note_cell.font = _MUTED
        row += 1

    last_data_row = row - 1

    # Totals
    row += 1
    ws.cell(row=row, column=2, value="Adjusted EBITDA (Seller view)").font = _BOLD
    c_seller = ws.cell(row=row, column=3, value=bridge.adjusted_ebitda_seller)
    c_seller.number_format = _USD
    c_seller.font = _BOLD
    c_seller.fill = _GREEN_FILL
    row += 1
    ws.cell(row=row, column=2, value="Adjusted EBITDA (Buyer view)").font = _BOLD
    c_buyer = ws.cell(row=row, column=4, value=bridge.adjusted_ebitda_buyer)
    c_buyer.number_format = _USD
    c_buyer.font = _BOLD
    c_buyer.fill = _AMBER_FILL

    # Comparison strip: LTM / Run-rate / 3-yr-avg
    row += 2
    ws.cell(row=row, column=2, value="Period comparison").font = _BOLD
    row += 1
    ws.cell(row=row, column=2, value="LTM EBITDA")
    if bridge.ltm_ebitda is not None:
        ws.cell(row=row, column=3, value=bridge.ltm_ebitda).number_format = _USD
    else:
        ws.cell(row=row, column=3, value="n/a").font = _MUTED
    row += 1
    ws.cell(row=row, column=2, value="Run-rate EBITDA")
    if bridge.run_rate_ebitda is not None:
        ws.cell(row=row, column=3, value=bridge.run_rate_ebitda).number_format = _USD
    else:
        gate_reason = (bridge.run_rate_gate or {}).get("reason", "gate not evaluated")
        ws.cell(row=row, column=3, value="blocked").fill = _RED_FILL
        ws.cell(row=row, column=8, value=f"Run-rate blocked: {gate_reason}").font = _MUTED
    row += 1
    ws.cell(row=row, column=2, value="3-yr avg EBITDA")
    if bridge.three_year_avg_ebitda is not None:
        ws.cell(row=row, column=3, value=bridge.three_year_avg_ebitda).number_format = _USD
    else:
        ws.cell(row=row, column=3, value="n/a").font = _MUTED

    # Materiality flags
    if bridge.material_flags:
        row += 2
        ws.cell(row=row, column=2, value=f"Material flags (≥ ${bridge.materiality_threshold:,.0f})").font = _BOLD
        row += 1
        for flag in bridge.material_flags:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            c = ws.cell(row=row, column=2, value=flag)
            c.fill = _AMBER_FILL
            c.alignment = Alignment(wrap_text=True)
            row += 1

    # Blocked adjustments
    if bridge.blocked_adjustments:
        row += 2
        ws.cell(row=row, column=2, value="Blocked adjustments (require additional evidence)").font = _BOLD
        row += 1
        for oc in bridge.blocked_adjustments:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            c = ws.cell(row=row, column=2, value=f"{oc.category_name} — {oc.block_reason}")
            c.fill = _RED_FILL
            c.alignment = Alignment(wrap_text=True)
            row += 1

    # Warnings (analyst-visible; e.g. reserved-vocab scrub, SBC suppression)
    if bridge.warnings:
        row += 2
        ws.cell(row=row, column=2, value="Warnings").font = _BOLD
        row += 1
        for w in bridge.warnings:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            c = ws.cell(row=row, column=2, value=w)
            c.fill = _MUTED_FILL
            c.alignment = Alignment(wrap_text=True)
            c.font = _MUTED
            row += 1

    # Disclaimer + signature
    row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c = ws.cell(row=row, column=2, value=bridge.disclaimer)
    c.alignment = Alignment(wrap_text=True)
    c.font = _MUTED
    row += 1
    ws.cell(row=row, column=2, value=bridge.sig).font = _MUTED

    # Conditional formatting on Recurrence column to keep it readable
    _apply_recurrence_conditional_formatting(ws, first_data_row, last_data_row)

    # Optional waterfall chart image
    try:
        image = _bridge_waterfall_png(bridge)
        if image is not None:
            ws.add_image(image, "K5")
    except Exception:  # pragma: no cover
        logger.exception("bridge waterfall render failed; sheet still valid")


def _apply_recurrence_conditional_formatting(
    ws: Worksheet, first_data_row: int, last_data_row: int
) -> None:
    if last_data_row < first_data_row:
        return
    rng = f"E{first_data_row}:E{last_data_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal",
        formula=[f'"{Recurrence.RECURRING.value}"'],
        fill=_AMBER_FILL,
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal",
        formula=[f'"{Recurrence.ONE_TIME.value}"'],
        fill=_GREEN_FILL,
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal",
        formula=[f'"{Recurrence.FLAG_ONLY.value}"'],
        fill=_RED_FILL,
    ))


def _bridge_waterfall_png(bridge: EBITDABridge) -> XLImage | None:
    """Best-effort waterfall PNG of the seller-side bridge; None on any error."""
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:  # pragma: no cover
        return None

    labels = ["Reported"]
    values = [bridge.reported_ebitda]
    colors = ["#0C1F33"]

    running = bridge.reported_ebitda
    for oc in bridge.adjustments:
        labels.append(oc.category_name[:24])
        values.append(oc.seller_amount)
        colors.append("#7EB77F" if oc.seller_amount >= 0 else "#C25F5F")
        running += oc.seller_amount

    labels.append("Adjusted (Seller)")
    values.append(bridge.adjusted_ebitda_seller)
    colors.append("#0C1F33")

    fig, ax = plt.subplots(figsize=(7, 3.5))
    xs = list(range(len(labels)))
    # Draw the bars — first and last are absolute, middle are deltas stacked
    # from the running baseline for visual clarity.
    baseline = 0.0
    running = 0.0
    for i, (label, val, color) in enumerate(zip(labels, values, colors, strict=False)):
        if i == 0:
            ax.bar(i, val, color=color)
            running = val
        elif i == len(labels) - 1:
            ax.bar(i, val, color=color)
        else:
            bottom = running if val >= 0 else running + val
            ax.bar(i, abs(val), bottom=bottom, color=color)
            running += val
    ax.set_xticks(xs)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    ax.set_title("EBITDA bridge — seller view", fontsize=10)
    ax.grid(axis="y", linestyle=":", alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=120)
    plt.close(fig)
    buf.seek(0)
    img = XLImage(buf)
    img.width = 620
    img.height = 320
    return img


__all__ = ["render_ebitda_bridge_sheet"]
