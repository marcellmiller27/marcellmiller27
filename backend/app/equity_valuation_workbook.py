# JHI-SIG: 69M2705M | Equity DCF valuation workbook | JHI Research & Analytics Firm, Inc. (proprietary)
"""Render an EquityValuation into a branded, print-ready DCF workbook (openpyxl),
reusing the shared Aegira/JHI styling from excel_export."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app import excel_export as xl
from app.equity_valuation import EquityValuation

_PCT = "0.0%"
_USD = '#,##0'
_USD2 = '#,##0.00'


def _label(ws, row: int, text: str, bold: bool = True) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = xl._BOLD if bold else Font()


def _kv(ws, row: int, label: str, value, fmt: str | None = None, note: str = "") -> None:
    ws.cell(row=row, column=1, value=label).font = xl._BOLD
    c = ws.cell(row=row, column=2, value=value)
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = xl._MUTED


def write_valuation_sheet(ws, v: EquityValuation) -> None:
    """Populate an existing worksheet with the full DCF valuation view.

    Extracted so the standalone valuation workbook and the institutional per-ticker
    workbook render the IDENTICAL DCF/IRR/signal sheet (single source of truth)."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    xl._title_block(
        ws,
        subtitle=f"Equity Valuation — {v.name} ({v.ticker})",
        business="Cross-Asset Valuation & Action Engine · Phase 1",
    )

    row = 5
    xl._subhead(ws, row, "Action")
    row += 1
    _kv(ws, row, "Signal", v.signal, note="Enter ≥ +20% MoS · Sideline ≤ −10% · else Accumulate")
    row += 1
    _kv(ws, row, "Margin of safety (upside)", v.upside_pct, _PCT)
    row += 1
    _kv(ws, row, "Implied expected return (IRR)", v.expected_return, _PCT,
        note="Discount rate that equates DCF value to today's price")
    row += 2

    xl._subhead(ws, row, "Valuation 2.0 (innovator-fair view)")
    row += 1
    _kv(ws, row, "Archetype (calibration)", v.archetype,
        note=f"Growth cap {v.growth_cap_used:.0%} · R&D treated as investment")
    row += 1
    _kv(ws, row, "Innovation & Moat score", v.innovation_moat_score, "0.0",
        note="0–100: R&D intensity + growth, margin & revenue durability, ROIC")
    row += 1
    _kv(ws, row, "Adjusted owner-earnings", v.adjusted_owner_earnings, _USD, note=v.rnd_treatment)
    row += 1
    _kv(ws, row, "ROIC", v.roic if v.roic is not None else "n/a",
        _PCT if v.roic is not None else None,
        note=f"vs cost of capital {v.cost_of_capital:.1%}")
    row += 1
    _kv(ws, row, "High-growth fade (years)", v.high_growth_years,
        note="Longer when ROIC > cost of capital (durable value creation)")
    row += 1
    _kv(ws, row, "Blended margin of safety", v.composite_margin, _PCT,
        note="Intrinsic upside + Innovation/Moat credit")
    row += 1
    _kv(ws, row, "Classic (prior) value / share", v.classic_intrinsic_per_share, _USD2,
        note=f"Conservative net-income DCF · {v.classic_signal} ({v.classic_upside_pct:.1%})")
    row += 2

    xl._subhead(ws, row, "Market")
    row += 1
    _kv(ws, row, "Market price", v.price, _USD2)
    row += 1
    _kv(ws, row, "Shares outstanding", v.shares_outstanding, _USD)
    row += 1
    _kv(ws, row, "Market capitalization", v.market_cap, _USD)
    row += 2

    xl._subhead(ws, row, "Assumptions (disclosed)")
    row += 1
    _kv(ws, row, "Base free cash flow", v.base_fcf, _USD, note=v.fcf_basis)
    row += 1
    _kv(ws, row, "Growth rate (years 1–5)", v.growth_rate, _PCT, note="Revenue CAGR, capped at 12%")
    row += 1
    _kv(ws, row, "Terminal growth", v.terminal_growth, _PCT, note="~ long-run nominal GDP")
    row += 1
    _kv(ws, row, "Risk-free rate (10Y UST)", v.risk_free, _PCT)
    row += 1
    _kv(ws, row, "Equity risk premium", v.equity_risk_premium, _PCT)
    row += 1
    _kv(ws, row, "Beta", v.beta, "0.00")
    row += 1
    _kv(ws, row, "Discount rate (cost of equity)", v.discount_rate, _PCT,
        note="Risk-free + beta × equity risk premium")
    row += 2

    xl._subhead(ws, row, "DCF projection")
    row += 1
    hdr = ["Year", "Projected FCF", "Present value"]
    for col, text in enumerate(hdr, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = xl._WHITE
        c.fill = xl._NAVY_FILL
    row += 1
    for i, (fcf, pv) in enumerate(zip(v.projected_fcf, v.present_values), start=1):
        ws.cell(row=row, column=1, value=f"Year {i}")
        ws.cell(row=row, column=2, value=fcf).number_format = _USD
        ws.cell(row=row, column=3, value=pv).number_format = _USD
        row += 1
    _kv(ws, row, "Terminal value", v.terminal_value, _USD)
    row += 1
    _kv(ws, row, "PV of terminal value", v.pv_terminal_value, _USD)
    row += 1
    _kv(ws, row, "Intrinsic equity value", v.intrinsic_equity_value, _USD)
    row += 1
    ws.cell(row=row, column=1, value="Intrinsic value / share").font = Font(bold=True, color=xl._NAVY)
    c = ws.cell(row=row, column=2, value=v.intrinsic_per_share)
    c.number_format = _USD2
    c.font = Font(bold=True, color=xl._NAVY)
    row += 2

    xl._subhead(ws, row, "Analyst note")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=4)
    note = ws.cell(row=row, column=1, value=v.rationale)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    row += 6

    xl._subhead(ws, row, "Sources")
    row += 1
    for s in v.sources:
        ws.cell(row=row, column=1, value=f"· {s}").font = xl._MUTED
        row += 1
    ws.cell(row=row, column=1, value=v.disclaimer).font = xl._MUTED
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    xl._watermark(ws)


def equity_valuation_workbook(v: EquityValuation) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Valuation"
    write_valuation_sheet(ws, v)
    xl._legal_sheet(wb)
    return xl._finalize(wb)
