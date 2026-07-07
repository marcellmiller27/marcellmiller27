# JHI-SIG: 69M2705M | Excel Dashboard Export | John Henry Investments (proprietary)
"""Interactive Excel workbook exports for Deal X-Ray (Business Quality Assessment)
and the Financial Diligence Suite (Quality of Earnings).

The workbooks are *live models*: the Dashboard holds editable input cells and the
outputs are written as native Excel formulas, so a subscriber can change assumptions
(price, down payment, rate, multiples, add-backs, …) and watch the outputs — DSCR,
valuation, working capital — recompute inside Excel with no macros. Detail tabs hold
the underlying data; a Legal & Provenance tab carries the disclaimer, the founder
signature (JHI-SIG: 69M2705M), and the JHI Research & Analytics Firm, Inc. attribution.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.deal_xray_models import DealInput, DealXRayReport
from app.financial_diligence_models import DiligenceInput, DiligenceReport

SIG = "JHI-SIG: 69M2705M"
ENTITY = "JHI Research & Analytics Firm, Inc."
LEGAL_LINES = [
    "Decision-support analysis generated from user-supplied figures.",
    "This document is NOT investment advice, a valuation/appraisal, a fairness opinion, "
    "an audit, a review, a CPA opinion, or brokerage.",
    "Formal assurance opinions (Unqualified / Qualified / Adverse / Disclaimer) are issued "
    "only by a licensed partner CPA firm that engages the target entity.",
    "Live-formula cells are a transparent single-model view for you to stress-test; the JHI "
    "engine snapshot may use additional multi-year logic. Verify all figures with a "
    "quality-of-earnings review and licensed professionals before making an offer.",
    f"© {date.today().year} {ENTITY}. All rights reserved. Proprietary & confidential.",
]

# --- palette (matches the platform's institutional light theme) ---
_NAVY = "0C1F33"
_NAVY_FILL = PatternFill("solid", fgColor="0C1F33")
_SUBHEAD_FILL = PatternFill("solid", fgColor="EAEFF6")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")  # soft gold = "you may edit"
_GREEN_FILL = PatternFill("solid", fgColor="D6F2E4")
_RED_FILL = PatternFill("solid", fgColor="F8D7D5")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_MUTED = Font(color="5A6B7D", italic=True, size=9)

_MONEY = '"$"#,##0'
_PCT = '0.0%'
_MULT = '0.00"x"'
_RATIO = '0.00'


def _title_block(ws: Worksheet, subtitle: str, business: str) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = ENTITY
    ws["A1"].font = _WHITE
    ws["A1"].fill = _NAVY_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:D2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(bold=True, size=13, color=_NAVY)
    ws.merge_cells("A3:D3")
    ws["A3"] = f"{business}  ·  generated {date.today().isoformat()}  ·  {SIG}"
    ws["A3"].font = _MUTED


def _subhead(ws: Worksheet, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=_NAVY)
    cell.fill = _SUBHEAD_FILL


def _input_row(ws: Worksheet, row: int, label: str, value: float, fmt: str, note: str = "") -> str:
    ws.cell(row=row, column=1, value=label).font = _BOLD
    c = ws.cell(row=row, column=2, value=value)
    c.fill = _INPUT_FILL
    c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = _MUTED
    return f"B{row}"


def _out_row(ws: Worksheet, row: int, label: str, formula: str, fmt: str, note: str = "") -> None:
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=2, value=formula)
    c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = _MUTED


def _legal_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Legal & Provenance")
    ws.column_dimensions["A"].width = 110
    ws.cell(row=1, column=1, value=ENTITY).font = Font(bold=True, size=13, color=_NAVY)
    ws.cell(row=2, column=1, value="Legal terms & provenance").font = _BOLD
    r = 4
    for line in LEGAL_LINES:
        cell = ws.cell(row=r, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    ws.cell(row=r + 1, column=1, value=f"Founder signature of provenance: {SIG}").font = _BOLD
    ws.cell(row=r + 2, column=1, value=f"Attribution: {ENTITY}").font = _BOLD


def _finalize(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deal_xray_workbook(deal: DealInput, report: DealXRayReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 46
    _title_block(ws, "Business Quality Assessment & Deal Model", deal.business_name)

    v = report.valuation
    _subhead(ws, 5, "INPUTS — edit the highlighted (gold) cells")
    rev = _input_row(ws, 7, "Revenue ($)", deal.revenue, _MONEY)
    ebitda = _input_row(ws, 8, "Reported EBITDA / SDE ($)", deal.reported_ebitda, _MONEY)
    addbacks = _input_row(ws, 9, "Add-backs — total ($)", deal.addbacks, _MONEY)
    basis = _input_row(ws, 10, "Valuation EBITDA basis ($)", v.normalized_ebitda, _MONEY, "Default = engine basis")
    capex = _input_row(ws, 11, "Annual capex ($)", float(v.dcf_assumptions.get("capex", 0)), _MONEY)
    price = _input_row(ws, 12, "Asking price ($)", deal.asking_price, _MONEY)
    down = _input_row(ws, 13, "Down payment (%)", deal.down_payment_pct, "0.0")
    note_pct = _input_row(ws, 14, "Seller note (%)", deal.seller_note_pct, "0.0")
    apr = _input_row(ws, 15, "Loan APR (%)", deal.loan_rate_pct, "0.0")
    term = _input_row(ws, 16, "Loan term (years)", deal.loan_term_years, "0")
    m_lo = _input_row(ws, 17, "Multiple — low", v.industry_multiple_low, _MULT)
    m_base = _input_row(ws, 18, "Multiple — base", v.industry_multiple_base, _MULT)
    m_hi = _input_row(ws, 19, "Multiple — high", v.industry_multiple_high, _MULT)

    _subhead(ws, 21, "LIVE OUTPUTS — recompute when you edit the inputs above")
    _out_row(ws, 23, "Adjusted EBITDA (current yr)", f"=({ebitda})-0.5*MAX(0,({addbacks})-0.25*({ebitda}))", _MONEY)
    _out_row(ws, 24, "EBITDA margin", f"=IF({rev}=0,0,{ebitda}/{rev})", _PCT)
    _out_row(ws, 25, "Asking multiple (on basis)", f"=IF({basis}=0,0,{price}/{basis})", _MULT)
    _out_row(ws, 26, "Valuation — low", f"={basis}*{m_lo}", _MONEY)
    _out_row(ws, 27, "Valuation — base", f"={basis}*{m_base}", _MONEY)
    _out_row(ws, 28, "Valuation — high", f"={basis}*{m_hi}", _MONEY)
    _out_row(ws, 29, "Equity required", f"={price}*{down}/100", _MONEY)
    _out_row(ws, 30, "Seller note ($)", f"={price}*{note_pct}/100", _MONEY)
    _out_row(ws, 31, "Loan amount", f"=MAX(0,{price}-B29-B30)", _MONEY)
    _out_row(ws, 32, "Annual debt service", f"=-PMT({apr}/100,{term},{price}-B29)", _MONEY, "Loan + seller note amortized")
    _out_row(ws, 33, "CFADS (basis − capex)", f"={basis}-{capex}", _MONEY)
    _out_row(ws, 34, "DSCR", "=IF(B32=0,0,B33/B32)", _RATIO, "Target ≥ 1.25")
    _out_row(ws, 35, "SBA fit", f'=IF(AND(B34>=1.25,{down}>=10),"Yes","Review")', "General")

    ws.conditional_formatting.add(
        "B34",
        CellIsRule(operator="greaterThanOrEqual", formula=["1.25"], fill=_GREEN_FILL),
    )
    ws.conditional_formatting.add(
        "B34",
        CellIsRule(operator="lessThan", formula=["1.25"], fill=_RED_FILL),
    )

    _subhead(ws, 37, "SCENARIO — DSCR by down payment (live)")
    ws.cell(row=38, column=2, value="10% down").font = _BOLD
    ws.cell(row=38, column=3, value="15% down").font = _BOLD
    ws.cell(row=38, column=4, value="20% down").font = _BOLD
    ws.cell(row=39, column=1, value="DSCR").font = _BOLD
    ws.cell(row=39, column=2, value=f"=B33/(-PMT({apr}/100,{term},{price}*0.9))").number_format = _RATIO
    ws.cell(row=39, column=3, value=f"=B33/(-PMT({apr}/100,{term},{price}*0.85))").number_format = _RATIO
    ws.cell(row=39, column=4, value=f"=B33/(-PMT({apr}/100,{term},{price}*0.8))").number_format = _RATIO

    _subhead(ws, 41, "JHI ASSESSMENT SNAPSHOT (Business Quality Assessment engine)")
    snap = [
        ("Deal Score (0–100)", report.deal_score),
        ("Recommendation", report.recommendation),
        ("Honest Ethic Rating (0–100)", report.ethic_rating),
        ("Valuation verdict", v.verdict),
    ]
    for i, (label, value) in enumerate(snap):
        ws.cell(row=42 + i, column=1, value=label).font = _BOLD
        ws.cell(row=42 + i, column=2, value=value)
    ws.cell(row=46, column=1, value=report.ethic_note).font = _MUTED
    ws.merge_cells("A46:D46")
    ws.cell(row=48, column=1, value="See the 'Legal & Provenance' tab for terms and disclosures.").font = _MUTED

    # Input guards
    dv_pct = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)
    ws.add_data_validation(dv_pct)
    dv_pct.add("B13")
    dv_pct.add("B14")

    _bqa_sheet(wb, report)
    _deal_detail_sheet(wb, deal, report)
    _legal_sheet(wb)
    return _finalize(wb)


def _bqa_sheet(wb: Workbook, report: DealXRayReport) -> None:
    ws = wb.create_sheet("Business Quality Assessment")
    ws.column_dimensions["A"].width = 26
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 70
    _title_block(ws, "Business Quality Assessment — six-segment scorecard", report.business_name)
    headers = ["Segment", "Weight", "Score", "Weighted", "Findings"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = _WHITE
        cell.fill = _NAVY_FILL
    r = 6
    for seg in report.segments:
        ws.cell(row=r, column=1, value=seg.segment)
        wcell = ws.cell(row=r, column=2, value=round(seg.weight, 4))
        wcell.number_format = _PCT
        wcell.fill = _INPUT_FILL
        ws.cell(row=r, column=3, value=seg.score)
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = _RATIO
        ws.cell(row=r, column=5, value=" ".join(seg.findings)).alignment = Alignment(wrap_text=True)
        r += 1
    ws.cell(row=r, column=1, value="Weighted total (your weights)").font = _BOLD
    ws.cell(row=r, column=4, value=f"=SUM(D6:D{r - 1})").number_format = _RATIO
    ws.cell(row=r + 2, column=1,
            value=f"Official Deal Score (engine): {report.deal_score} · {report.recommendation}").font = _MUTED


def _deal_detail_sheet(wb: Workbook, deal: DealInput, report: DealXRayReport) -> None:
    ws = wb.create_sheet("Detail")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    _title_block(ws, "Deal detail & assumptions", deal.business_name)
    rows = [
        ("Industry", deal.industry),
        ("Year founded", deal.year_founded or "n/a"),
        ("Employees", deal.employees),
        ("Owner involvement", deal.owner_involvement),
        ("Customer concentration (%)", deal.customer_concentration_pct),
        ("Recurring revenue (%)", deal.recurring_revenue_pct),
        ("Revenue — prior year ($)", deal.revenue_prior or "n/a"),
        ("Valuation basis note", report.valuation.basis_note),
    ]
    r = 5
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = _BOLD
        ws.cell(row=r, column=2, value=value)
        r += 1
    if deal.earnings_history:
        ws.cell(row=r + 1, column=1, value="Earnings history (most recent first)").font = _BOLD
        for i, e in enumerate(deal.earnings_history):
            ws.cell(row=r + 2 + i, column=1, value=f"Year -{i}")
            ws.cell(row=r + 2 + i, column=2, value=e).number_format = _MONEY
    kr = r + 4 + len(deal.earnings_history or [])
    ws.cell(row=kr, column=1, value="Diligence questions").font = _BOLD
    for i, q in enumerate(report.diligence_questions):
        ws.cell(row=kr + 1 + i, column=1, value=f"• {q}").alignment = Alignment(wrap_text=True)


def diligence_workbook(deal: DiligenceInput, report: DiligenceReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 44
    _title_block(ws, "Quality of Earnings — dashboard", deal.business_name)

    _subhead(ws, 5, "INPUTS — edit the highlighted (gold) cells")
    rev = _input_row(ws, 7, "Revenue ($)", deal.revenue, _MONEY)
    ebitda = _input_row(ws, 8, "Reported EBITDA / SDE ($)", deal.reported_ebitda, _MONEY)
    q_add = _input_row(ws, 9, "Questionable add-backs ($)", deal.questionable_addbacks, _MONEY)
    onetime = _input_row(ws, 10, "One-time items ($)", deal.one_time_items, _MONEY)
    deposits = _input_row(ws, 11, "Bank deposits ($)", deal.bank_deposits or 0.0, _MONEY)
    ar = _input_row(ws, 12, "Accounts receivable ($)", deal.accounts_receivable, _MONEY)
    inv = _input_row(ws, 13, "Inventory ($)", deal.inventory, _MONEY)
    ap = _input_row(ws, 14, "Accounts payable ($)", deal.accounts_payable, _MONEY)

    _subhead(ws, 16, "LIVE OUTPUTS — recompute when you edit the inputs above")
    _out_row(ws, 18, "Adjusted EBITDA", f"={ebitda}-0.5*{q_add}-{onetime}", _MONEY)
    _out_row(ws, 19, "Proof-of-cash variance", f"=IF({rev}=0,0,({deposits}-{rev})/{rev})", _PCT, "Deposits vs. revenue")
    _out_row(ws, 20, "Net working capital", f"={ar}+{inv}-{ap}", _MONEY)
    _out_row(ws, 21, "NWC % of revenue", f"=IF({rev}=0,0,({ar}+{inv}-{ap})/{rev})", _PCT)

    _subhead(ws, 23, "JHI ASSESSMENT SNAPSHOT")
    snap = [
        ("Financial Integrity Score (0–100)", report.financial_integrity_score),
        ("Recommended tier", report.recommended_tier),
        ("Add-on QoE price (platform)", f"${report.add_on_pricing.platform_low:,.0f}–${report.add_on_pricing.platform_high:,.0f}"),
    ]
    for i, (label, value) in enumerate(snap):
        ws.cell(row=24 + i, column=1, value=label).font = _BOLD
        ws.cell(row=24 + i, column=2, value=value)
    ws.cell(row=27, column=1, value=report.recommended_action).font = _MUTED
    ws.merge_cells("A27:C27")

    if report.red_flags:
        _subhead(ws, 29, "RED FLAGS")
        for i, flag in enumerate(report.red_flags):
            ws.cell(row=30 + i, column=1, value=f"• {flag}").alignment = Alignment(wrap_text=True)

    _legal_sheet(wb)
    return _finalize(wb)
