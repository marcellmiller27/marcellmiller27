# JHI-SIG: 69M2705M | Institutional per-ticker Excel workbook | JHI Research & Analytics Firm, Inc. (proprietary)
"""Build the comprehensive, branded, multi-sheet per-ticker workbook produced from the
Portfolio / ticker view — institutional analysis that goes beyond a single quote:

  1. Cover / Summary        — headline technical + fundamental + opportunity read.
  2. Technicals — Daily     — structure, indicators, S/R, 52-wk, daily trade setups.
  3. Technicals — Weekly    — higher-timeframe structure + position/swing read.
  4. Options context        — realized-vol, ATR expected move, strategy archetypes.
  5. Fundamental Ratios     — full ratio table + multi-period trend (SF1/EDGAR).
  6. DCF Valuation          — reuses the existing valuation sheet (IRR / signal).
  7. Legal & Provenance     — shared disclaimer + founder signature + attribution.

Everything is deterministic and as-of dated; missing data degrades to "n/a" rather
than fabricating numbers (Data Foundation doctrine). OHLC bars come from the shared
Yahoo chart adapter (no new vendor); fundamentals from the SF1-primary/EDGAR provider.
"""

from __future__ import annotations

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment

from app import edgar_services, equity_ratios, equity_technicals, excel_export as xl, fundamentals
from app import market_services
from app.equity_ratios import FundamentalRatios, RatioMetric
from app.equity_technicals import OptionsContext, TechnicalsRead
from app.equity_valuation import EquityValuation, value_equity
from app.equity_valuation_workbook import write_valuation_sheet

logger = logging.getLogger(__name__)

_PCT = "0.0%"
_PCT2 = "0.00%"
_USD = "#,##0"
_USD2 = "#,##0.00"
_RATIO = "0.00"
_MULT = '0.00"x"'
_NA = "n/a"

_FMT_MAP = {
    equity_ratios.FMT_PCT: _PCT,
    equity_ratios.FMT_RATIO: _RATIO,
    equity_ratios.FMT_USD: _USD,
    equity_ratios.FMT_MULT: _MULT,
    equity_ratios.FMT_EPS: _USD2,
}


# ── Small rendering helpers ──────────────────────────────────────────────────
def _kv(ws, row: int, label: str, value, fmt: str | None = None, note: str = "") -> None:
    ws.cell(row=row, column=1, value=label).font = xl._BOLD
    if value is None:
        c = ws.cell(row=row, column=2, value=_NA)
        c.font = xl._MUTED
    else:
        c = ws.cell(row=row, column=2, value=value)
        if fmt:
            c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = xl._MUTED


def _wrap(ws, row: int, text: str, span: int = 4, height: int | None = None) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if height:
        ws.row_dimensions[row].height = height


def _hdr(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = xl._WHITE
        c.fill = xl._NAVY_FILL


def _fmt_pct(x: float | None) -> str:
    return f"{x * 100:.1f}%" if x is not None else _NA


def _fmt_usd(x: float | None) -> str:
    if x is None:
        return _NA
    if abs(x) >= 1e12:
        return f"${x / 1e12:.2f}T"
    if abs(x) >= 1e9:
        return f"${x / 1e9:.2f}B"
    if abs(x) >= 1e6:
        return f"${x / 1e6:.1f}M"
    return f"${x:,.2f}"


# ── Data assembly ────────────────────────────────────────────────────────────
class TickerWorkbookData:
    """Assembled inputs for the workbook (fetched once, reused across sheets)."""

    def __init__(
        self,
        ticker: str,
        *,
        daily: TechnicalsRead,
        weekly: TechnicalsRead,
        options: OptionsContext,
        ratios: FundamentalRatios,
        valuation: EquityValuation | None,
        valuation_error: str | None,
        price: float | None,
        market_cap: float | None,
        name: str,
    ) -> None:
        self.ticker = ticker
        self.daily = daily
        self.weekly = weekly
        self.options = options
        self.ratios = ratios
        self.valuation = valuation
        self.valuation_error = valuation_error
        self.price = price
        self.market_cap = market_cap
        self.name = name


def _resolve_symbol(ticker: str) -> str:
    """Map a ticker to the Yahoo chart symbol via the shared symbol registry."""
    try:
        spec = market_services.resolve_symbol(ticker)
        return spec.provider_symbol
    except Exception:  # noqa: BLE001 - resolve_symbol falls back internally; guard anyway
        return ticker.upper()


def assemble(ticker: str, price: float | None = None) -> TickerWorkbookData:
    """Fetch and compute every input the workbook needs. Resilient: technicals and
    fundamentals/valuation degrade independently so a workbook always builds."""
    ticker = ticker.strip().upper()

    # 1) OHLC → daily + weekly technicals (reuses shared Yahoo adapter, no new vendor).
    provider_symbol = _resolve_symbol(ticker)
    rows = market_services.yahoo_chart_ohlc(provider_symbol, range_="2y", interval="1d")
    daily_bars = equity_technicals.to_bars(rows)
    daily = equity_technicals.compute_technicals(daily_bars, ticker, "Daily", window=252)
    weekly_bars = equity_technicals.aggregate_weekly(daily_bars)
    weekly = equity_technicals.compute_technicals(weekly_bars, ticker, "Weekly", window=52)
    options = equity_technicals.options_context(daily, weekly)

    live_price = price if price is not None else daily.price

    # 2) Fundamentals → ratios (SF1 primary, EDGAR fallback). Governance: derived only.
    fin = fundamentals.equity_fundamentals(ticker, max_years=6)
    ratios = equity_ratios.compute_ratios(fin, price=live_price)
    name = fin.entity_name
    market_cap = (live_price * fin.shares_outstanding) if (
        live_price and fin.shares_outstanding
    ) else None

    # 3) DCF valuation (reuses the existing engine). Degrades gracefully.
    valuation: EquityValuation | None = None
    valuation_error: str | None = None
    try:
        valuation = value_equity(ticker, price=live_price)
    except (edgar_services.ProviderError, ValueError) as exc:
        valuation_error = str(exc)
    except Exception as exc:  # noqa: BLE001 - never let a valuation edge-case break the workbook
        logger.debug("valuation failed for %s in ticker workbook", ticker, exc_info=True)
        valuation_error = str(exc)

    return TickerWorkbookData(
        ticker=ticker,
        daily=daily,
        weekly=weekly,
        options=options,
        ratios=ratios,
        valuation=valuation,
        valuation_error=valuation_error,
        price=live_price,
        market_cap=market_cap,
        name=name,
    )


# ── Sheet builders ───────────────────────────────────────────────────────────
def _headline(data: TickerWorkbookData) -> tuple[str, str, str]:
    """One-line technical read, fundamental read, and opportunity summary."""
    d, w = data.daily, data.weekly
    align = "aligned" if d.trend == w.trend else "diverging"
    tech = (
        f"{d.trend} on the daily and {w.trend} on the weekly ({align}); "
        f"RSI {d.rsi14:.0f}" if d.rsi14 is not None else f"{d.trend} on the daily and {w.trend} on the weekly ({align})"
    )
    if d.rsi14 is not None:
        tech += f", last structure event {d.last_structure_event}."
    else:
        tech += f", last structure event {d.last_structure_event}."

    r = data.ratios
    gm = next((m.value for s in r.sections for m in s.metrics if m.label == "Gross margin"), None)
    roe = next(
        (m.value for s in r.sections for m in s.metrics if m.label == "Return on equity (ROE)"), None
    )
    fund = (
        f"Gross margin {_fmt_pct(gm)}, ROE {_fmt_pct(roe)}"
        + (
            f"; DCF signal {data.valuation.signal} ({_fmt_pct(data.valuation.upside_pct)} margin of safety)."
            if data.valuation
            else " (DCF valuation unavailable for this name)."
        )
    )

    if data.valuation:
        posture = data.valuation.signal
        val_read = f"valuation posture {posture} ({_fmt_pct(data.valuation.upside_pct)} MoS)"
    else:
        val_read = "valuation posture n/a"
    opp = (
        f"Trend is {d.trend.lower()} (daily) / {w.trend.lower()} (weekly) with {val_read}. "
        f"Vol regime {data.options.vol_regime.lower()}; directional bias {data.options.directional_bias}."
    )
    return tech, fund, opp


def _cover_sheet(wb: Workbook, data: TickerWorkbookData) -> None:
    ws = wb.active
    ws.title = "Cover & Summary"
    ws.column_dimensions["A"].width = 30
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 24
    xl._title_block(
        ws,
        subtitle=f"Institutional Ticker Workbook — {data.name} ({data.ticker})",
        business="Portfolio · per-ticker technicals + fundamentals",
    )

    row = 5
    xl._subhead(ws, row, "Snapshot")
    row += 1
    _kv(ws, row, "Ticker", data.ticker)
    row += 1
    _kv(ws, row, "As of (price/technicals)", data.daily.as_of)
    row += 1
    _kv(ws, row, "Price", data.price, _USD2)
    row += 1
    _kv(ws, row, "Market capitalization", data.market_cap, _USD)
    row += 1
    _kv(ws, row, "52-week range", None)
    ws.cell(row=row, column=2, value=f"{_fmt_usd(data.daily.range_low)} – {_fmt_usd(data.daily.range_high)}")
    ws.cell(row=row, column=3, value=f"position {_fmt_pct((data.daily.range_position_pct or 0) / 100)}").font = xl._MUTED
    row += 2

    tech, fund, opp = _headline(data)
    xl._subhead(ws, row, "Technical read")
    row += 1
    _wrap(ws, row, tech, height=30)
    row += 2
    xl._subhead(ws, row, "Fundamental read")
    row += 1
    _wrap(ws, row, fund, height=30)
    row += 2
    xl._subhead(ws, row, "Opportunity summary")
    row += 1
    _wrap(ws, row, opp, height=45)
    row += 3

    xl._subhead(ws, row, "Contents")
    row += 1
    for name in (
        "Technicals — Daily",
        "Technicals — Weekly",
        "Options context",
        "Fundamental Ratios",
        "DCF Valuation" if data.valuation else "DCF Valuation (unavailable — see note)",
        "Legal & Provenance",
    ):
        ws.cell(row=row, column=1, value=f"· {name}").font = xl._MUTED
        row += 1
    row += 1
    _wrap(
        ws,
        row,
        "Research and educational output only — not investment advice, not a recommendation, "
        "and not an audit or CPA opinion. All figures are deterministic and as-of dated.",
        height=30,
    )


def _setups_block(ws, row: int, read: TechnicalsRead) -> int:
    xl._subhead(ws, row, f"Trade setups ({read.timeframe}) — derived from computed levels")
    row += 1
    _hdr(ws, row, ["Bias", "Style", "Trigger", "Stop", "Target", "R:R"])
    row += 1
    for s in read.setups:
        ws.cell(row=row, column=1, value=s.bias)
        ws.cell(row=row, column=2, value=s.style)
        ws.cell(row=row, column=3, value=s.trigger if s.trigger is not None else _NA).number_format = _USD2
        ws.cell(row=row, column=4, value=s.stop if s.stop is not None else _NA).number_format = _USD2
        ws.cell(row=row, column=5, value=s.target if s.target is not None else _NA).number_format = _USD2
        ws.cell(row=row, column=6, value=s.risk_reward if s.risk_reward is not None else _NA)
        row += 1
        _wrap(ws, row, s.rationale, height=28)
        row += 1
    return row + 1


def _technicals_sheet(wb: Workbook, read: TechnicalsRead, title: str, extra_note: str = "") -> None:
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 30
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 18
    xl._title_block(
        ws,
        subtitle=f"Technicals — {read.timeframe} — {read.ticker}",
        business=f"{read.bars_used} {read.timeframe.lower()} bars · as of {read.as_of}",
    )

    row = 5
    xl._subhead(ws, row, "Market structure")
    row += 1
    _kv(ws, row, "Trend", read.trend)
    row += 1
    _kv(ws, row, "Swing sequence", read.structure_sequence, note="Recent HH/HL/LH/LL labels")
    row += 1
    _kv(ws, row, "Last structure event", read.last_structure_event, note=read.last_structure_event_date or "")
    row += 1
    _kv(ws, row, "52-wk / window high", read.range_high, _USD2)
    row += 1
    _kv(ws, row, "52-wk / window low", read.range_low, _USD2)
    row += 1
    _kv(ws, row, "Position in range", (read.range_position_pct or 0) / 100 if read.range_position_pct is not None else None, _PCT)
    row += 2

    xl._subhead(ws, row, "Moving averages")
    row += 1
    _kv(ws, row, "SMA 20 / 50 / 200", None)
    ws.cell(row=row, column=2, value=read.sma20).number_format = _USD2
    ws.cell(row=row, column=3, value=read.sma50 if read.sma50 is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=4, value=read.sma200 if read.sma200 is not None else _NA).number_format = _USD2
    row += 1
    _kv(ws, row, "EMA 12 / 26", None)
    ws.cell(row=row, column=2, value=read.ema12 if read.ema12 is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=3, value=read.ema26 if read.ema26 is not None else _NA).number_format = _USD2
    row += 1
    _kv(ws, row, "EMA 20 / 50 / 200", None)
    ws.cell(row=row, column=2, value=read.ema20 if read.ema20 is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=3, value=read.ema50 if read.ema50 is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=4, value=read.ema200 if read.ema200 is not None else _NA).number_format = _USD2
    row += 2

    xl._subhead(ws, row, "Momentum & volatility")
    row += 1
    _kv(ws, row, "RSI (14)", read.rsi14, _RATIO)
    row += 1
    _kv(ws, row, "MACD line / signal / hist", None)
    ws.cell(row=row, column=2, value=read.macd_line if read.macd_line is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=3, value=read.macd_signal if read.macd_signal is not None else _NA).number_format = _USD2
    ws.cell(row=row, column=4, value=read.macd_hist if read.macd_hist is not None else _NA).number_format = _USD2
    row += 1
    _kv(ws, row, "ATR (14)", read.atr14, _USD2, note="Average True Range")
    row += 1
    _kv(ws, row, "ATR % of price", read.atr_pct, _PCT2)
    row += 1
    _kv(ws, row, "Annualized realized vol", read.realized_vol, _PCT)
    row += 2

    xl._subhead(ws, row, "Support / resistance")
    row += 1
    _kv(ws, row, "Resistance (nearest → far)", None)
    ws.cell(row=row, column=2, value=", ".join(f"{x:,.2f}" for x in read.resistances) or _NA)
    row += 1
    _kv(ws, row, "Support (nearest → far)", None)
    ws.cell(row=row, column=2, value=", ".join(f"{x:,.2f}" for x in read.supports) or _NA)
    row += 2

    row = _setups_block(ws, row, read)

    if extra_note:
        xl._subhead(ws, row, "Timeframe alignment")
        row += 1
        _wrap(ws, row, extra_note, height=30)
        row += 2

    xl._subhead(ws, row, "Notes")
    row += 1
    for n in read.notes:
        _wrap(ws, row, n, height=24)
        row += 1


def _options_sheet(wb: Workbook, oc: OptionsContext, ticker: str) -> None:
    ws = wb.create_sheet("Options context")
    ws.column_dimensions["A"].width = 32
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 22
    xl._title_block(
        ws,
        subtitle=f"Options context (deterministic framework) — {ticker}",
        business=f"as of {oc.as_of}",
    )
    row = 5
    xl._subhead(ws, row, "Volatility & expected move")
    row += 1
    _kv(ws, row, "Directional bias", oc.directional_bias)
    row += 1
    _kv(ws, row, "Volatility regime", oc.vol_regime)
    row += 1
    _kv(ws, row, "Annualized realized vol", oc.annualized_vol, _PCT)
    row += 1
    _kv(ws, row, "Daily ATR % of price", oc.daily_atr_pct, _PCT2)
    row += 1
    _kv(ws, row, "Expected move ~1 week (±)", oc.expected_move_1w, _USD2,
        note=f"± {_fmt_pct(oc.expected_move_1w_pct)}")
    row += 1
    _kv(ws, row, "Expected move ~1 month (±)", oc.expected_move_1m, _USD2,
        note=f"± {_fmt_pct(oc.expected_move_1m_pct)}")
    row += 2

    xl._subhead(ws, row, "Strategy archetypes (framework — not a recommendation)")
    row += 1
    for s in oc.strategies:
        _wrap(ws, row, f"· {s}", height=30)
        row += 1
    row += 1

    xl._subhead(ws, row, "Notes")
    row += 1
    for n in oc.notes:
        _wrap(ws, row, n, height=30)
        row += 1


def _ratio_value_cell(ws, row: int, metric: RatioMetric) -> None:
    ws.cell(row=row, column=1, value=metric.label).font = xl._BOLD
    if metric.value is None:
        ws.cell(row=row, column=2, value=_NA).font = xl._MUTED
    else:
        c = ws.cell(row=row, column=2, value=metric.value)
        c.number_format = _FMT_MAP.get(metric.fmt, _RATIO)
    ws.cell(row=row, column=3, value=metric.note).font = xl._MUTED


def _ratios_sheet(wb: Workbook, r: FundamentalRatios) -> None:
    ws = wb.create_sheet("Fundamental Ratios")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    xl._title_block(
        ws,
        subtitle=f"Fundamental Ratios — {r.name} ({r.ticker})",
        business=f"{r.source_label} · fiscal {r.fiscal_period}",
    )
    row = 5
    for section in r.sections:
        xl._subhead(ws, row, section.title)
        row += 1
        for m in section.metrics:
            _ratio_value_cell(ws, row, m)
            row += 1
        row += 1

    if r.trend:
        xl._subhead(ws, row, "Multi-period trend")
        row += 1
        _hdr(ws, row, ["Fiscal year", "Revenue", "Rev YoY", "Net income", "EPS", "Gross %", "Op %", "Net %"])
        row += 1
        for t in r.trend:
            ws.cell(row=row, column=1, value=t.fiscal_year)
            ws.cell(row=row, column=2, value=t.revenue if t.revenue is not None else _NA).number_format = _USD
            ws.cell(row=row, column=3, value=t.revenue_yoy if t.revenue_yoy is not None else _NA).number_format = _PCT
            ws.cell(row=row, column=4, value=t.net_income if t.net_income is not None else _NA).number_format = _USD
            ws.cell(row=row, column=5, value=t.eps if t.eps is not None else _NA).number_format = _USD2
            ws.cell(row=row, column=6, value=t.gross_margin if t.gross_margin is not None else _NA).number_format = _PCT
            ws.cell(row=row, column=7, value=t.operating_margin if t.operating_margin is not None else _NA).number_format = _PCT
            ws.cell(row=row, column=8, value=t.net_margin if t.net_margin is not None else _NA).number_format = _PCT
            row += 1
        row += 1

    xl._subhead(ws, row, "Sources & notes")
    row += 1
    for n in r.notes:
        _wrap(ws, row, n, height=30)
        row += 1


def _valuation_sheet(wb: Workbook, data: TickerWorkbookData) -> None:
    ws = wb.create_sheet("DCF Valuation")
    if data.valuation is not None:
        write_valuation_sheet(ws, data.valuation)
        return
    # Graceful degradation — a labeled placeholder, never a fabricated valuation.
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    xl._title_block(
        ws,
        subtitle=f"DCF Valuation — {data.name} ({data.ticker})",
        business="Cross-Asset Valuation & Action Engine · Phase 1",
    )
    xl._subhead(ws, 5, "Valuation unavailable")
    _wrap(
        ws,
        6,
        "A disclosed-assumption DCF could not be computed for this name "
        f"({data.valuation_error or 'insufficient fundamentals'}). Per the Data Foundation "
        "doctrine we show no fabricated intrinsic value; the technical and ratio sheets remain "
        "fully populated.",
        span=4,
        height=60,
    )
    xl._watermark(ws)


def build_ticker_workbook(ticker: str, price: float | None = None) -> tuple[bytes, TickerWorkbookData]:
    """Assemble data and render the full institutional per-ticker workbook.

    Returns (xlsx_bytes, data) so callers can reuse the resolved name/as-of for the
    download filename. Raises ``market_services.ProviderError`` only when OHLC history
    cannot be fetched at all (no chart = no workbook)."""
    data = assemble(ticker, price=price)
    return render_workbook(data), data


def render_workbook(data: TickerWorkbookData) -> bytes:
    """Render the workbook from already-assembled data (network-free — testable)."""
    wb = Workbook()
    _cover_sheet(wb, data)
    _technicals_sheet(
        wb,
        data.daily,
        "Technicals — Daily",
        extra_note=(
            f"Daily trend {data.daily.trend}; weekly trend {data.weekly.trend} — "
            + ("aligned (higher conviction)." if data.daily.trend == data.weekly.trend
               else "diverging (lower conviction; defer to the higher timeframe).")
        ),
    )
    _technicals_sheet(
        wb,
        data.weekly,
        "Technicals — Weekly",
        extra_note=(
            f"Weekly trend {data.weekly.trend}; daily trend {data.daily.trend} — "
            + ("aligned." if data.daily.trend == data.weekly.trend else "diverging.")
        ),
    )
    _options_sheet(wb, data.options, data.ticker)
    _ratios_sheet(wb, data.ratios)
    _valuation_sheet(wb, data)
    xl._legal_sheet(wb)
    for sheet in wb.worksheets:
        xl._watermark(sheet)
    return xl._finalize(wb)
