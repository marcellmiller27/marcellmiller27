# JHI-SIG: 69M2705M | Excel Dashboard Export | John Henry Investments (proprietary)
"""Tests for the interactive Excel workbook exports (Deal X-Ray / QoE)."""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _all_text(ws) -> str:
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                out.append(v)
    return " ".join(out)


def test_deal_xray_export_is_interactive_workbook() -> None:
    payload = {
        "business_name": "Carrollton Design Build",
        "industry": "construction_management",
        "revenue": 12_962_195,
        "revenue_prior": 11_701_091,
        "reported_ebitda": 2_381_009,
        "addbacks": 58_745,
        "annual_depreciation": 56_362,
        "earnings_history": [2_381_009, 1_612_599, 827_662],
        "employees": 14,
        "owner_involvement": "owner_operated",
        "customer_concentration_pct": 64,
        "recurring_revenue_pct": 15,
        "asking_price": 6_200_000,
        "down_payment_pct": 10,
    }
    resp = client.post("/api/v1/deal-xray/export.xlsx", json=payload)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == _XLSX_MEDIA
    assert "JHI_BQA_Carrollton_Design_Build.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))  # formulas preserved (data_only=False)
    assert {"Dashboard", "Business Quality Assessment", "Detail", "Legal & Provenance"} <= set(wb.sheetnames)

    ws = wb["Dashboard"]
    # DSCR must be a live formula, not a precomputed number.
    assert str(ws["B34"].value).startswith("=")
    # Valuation + debt-service cells are formulas too.
    assert str(ws["B26"].value).startswith("=")
    assert "PMT" in str(ws["B32"].value)

    legal = _all_text(wb["Legal & Provenance"])
    assert "JHI-SIG: 69M2705M" in legal
    assert "JHI Research & Analytics Firm, Inc." in legal
    assert "NOT investment advice" in legal


def test_deal_xray_export_bqa_sheet_has_weighted_formula() -> None:
    payload = {"business_name": "X Co", "revenue": 2_000_000, "reported_ebitda": 400_000,
               "asking_price": 1_200_000}
    resp = client.post("/api/v1/deal-xray/export.xlsx", json=payload)
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    ws = wb["Business Quality Assessment"]
    # Weighted column = weight * score as a live formula.
    assert str(ws["D6"].value).startswith("=")


def test_diligence_export_is_interactive_workbook() -> None:
    payload = {
        "business_name": "Messy Co",
        "revenue": 2_000_000,
        "reported_ebitda": 600_000,
        "questionable_addbacks": 200_000,
        "one_time_items": 100_000,
        "bank_deposits": 1_700_000,
        "accounts_receivable": 500_000,
        "inventory": 400_000,
        "accounts_payable": 100_000,
        "customer_concentration_pct": 70,
    }
    resp = client.post("/api/v1/financial-diligence/export.xlsx", json=payload)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == _XLSX_MEDIA
    wb = load_workbook(BytesIO(resp.content))
    assert {"Dashboard", "Legal & Provenance"} <= set(wb.sheetnames)
    ws = wb["Dashboard"]
    # Net working capital is a live formula (AR + inventory - AP).
    assert str(ws["B20"].value).startswith("=")
    legal = _all_text(wb["Legal & Provenance"])
    assert "JHI-SIG: 69M2705M" in legal
    assert "not investment advice" in legal.lower()
    assert "an audit" in legal.lower()


def test_export_rejects_invalid_input() -> None:
    resp = client.post("/api/v1/deal-xray/export.xlsx", json={"business_name": "X", "revenue": -1})
    assert resp.status_code == 422


def test_deal_xray_pdf_export() -> None:
    resp = client.post("/api/v1/deal-xray/export.pdf", json={
        "business_name": "Carrollton Design Build", "revenue": 12_962_195,
        "reported_ebitda": 2_381_009, "asking_price": 6_200_000,
    })
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"
    assert "JHI_BQA_Carrollton_Design_Build.pdf" in resp.headers["content-disposition"]


def test_diligence_pdf_export() -> None:
    resp = client.post("/api/v1/financial-diligence/export.pdf", json={
        "business_name": "Messy Co", "revenue": 2_000_000, "reported_ebitda": 600_000,
        "bank_deposits": 1_700_000, "accounts_receivable": 500_000, "accounts_payable": 100_000,
        "customer_concentration_pct": 70,
    })
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_excel_has_provenance_footer() -> None:
    resp = client.post("/api/v1/deal-xray/export.xlsx", json={
        "business_name": "Footer Co", "revenue": 1_000_000, "reported_ebitda": 200_000,
        "asking_price": 800_000,
    })
    wb = load_workbook(BytesIO(resp.content))
    footer = wb["Dashboard"].oddFooter.left.text or ""
    assert "69M2705M" in footer
