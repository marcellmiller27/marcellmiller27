# JHI-SIG: 69M2705M | Ratio dashboard tests | JHI Research & Analytics Firm, Inc. (proprietary)
"""Tests for the unified 6-section / 8-column ratio dashboard."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.ratio_dashboard import (
    BASE_SECTIONS,
    DASHBOARD_SCHEMA_VERSION,
    DashboardInputs,
    RatioSeries,
    compute_dashboard,
    render_dashboard_sheet,
)
from app.sector_profiles import (
    CURRENT_RATIO,
    DEBT_TO_EQUITY,
    EBITDA_MARGIN,
    GROSS_MARGIN,
    INVENTORY_TURNOVER,
    NIM,
    PE,
    ROA,
    Sector,
    Status,
)


def test_six_base_sections_present():
    titles = [s.title for s in BASE_SECTIONS]
    assert titles == [
        "Profitability",
        "Liquidity",
        "Solvency",
        "Efficiency",
        "Cash Flow",
        "Valuation",
    ]


def test_schema_version_stamped():
    assert DASHBOARD_SCHEMA_VERSION.startswith("2026.")


def test_dashboard_computes_status_green_for_healthy_tech():
    inputs = DashboardInputs(
        entity_name="Acme SaaS Corp.",
        period_label="FY2026",
        sector=Sector.TECHNOLOGY,
        ratios={
            GROSS_MARGIN: RatioSeries(latest=0.80, prior=0.78,
                                      history=[0.72, 0.75, 0.78, 0.80]),
        },
    )
    result = compute_dashboard(inputs)
    gm_row = next(r for r in result.rows if r.ratio_id == GROSS_MARGIN)
    assert gm_row.status == Status.GREEN
    assert gm_row.trend_arrow == "▲"


def test_dashboard_nm_sector_when_ratio_not_relevant():
    """Bank must render inventory-turnover as N/M — sector, not red."""
    inputs = DashboardInputs(
        entity_name="Regional Bank Corp.",
        period_label="FY2026",
        sector=Sector.BANK,
        ratios={INVENTORY_TURNOVER: RatioSeries(latest=6.0)},
    )
    result = compute_dashboard(inputs)
    row = next(r for r in result.rows if r.ratio_id == INVENTORY_TURNOVER)
    assert row.status == Status.NM_SECTOR
    assert not row.is_relevant
    assert "not a defining ratio" in row.driver_note.lower()


def test_dashboard_nm_data_when_input_missing():
    inputs = DashboardInputs(
        entity_name="TestCo",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={},
    )
    result = compute_dashboard(inputs)
    ebitda_row = next(r for r in result.rows if r.ratio_id == EBITDA_MARGIN)
    assert ebitda_row.status == Status.NM_DATA
    assert "N/M" in ebitda_row.driver_note or "missing" in ebitda_row.driver_note.lower()


def test_dashboard_pe_nm_data_for_negative_earnings():
    """Force-N/M-data flag simulates the P/E-with-negative-EPS case."""
    inputs = DashboardInputs(
        entity_name="LossCo",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={PE: RatioSeries(latest=None, force_nm_data=True)},
    )
    result = compute_dashboard(inputs)
    pe_row = next(r for r in result.rows if r.ratio_id == PE)
    assert pe_row.status == Status.NM_DATA
    assert "negative" in pe_row.driver_note.lower() or "n/m" in pe_row.driver_note.lower()


def test_dashboard_trend_arrows():
    inputs = DashboardInputs(
        entity_name="TestCo",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={
            EBITDA_MARGIN: RatioSeries(latest=0.18, history=[0.12, 0.15, 0.18]),
            CURRENT_RATIO: RatioSeries(latest=2.0, history=[3.0, 2.5, 2.0]),
            DEBT_TO_EQUITY: RatioSeries(latest=0.75, history=[0.75, 0.75, 0.75]),
        },
    )
    result = compute_dashboard(inputs)
    trends = {r.ratio_id: r.trend_arrow for r in result.rows}
    assert trends[EBITDA_MARGIN] == "▲"
    assert trends[CURRENT_RATIO] == "▼"
    assert trends[DEBT_TO_EQUITY] == "→"


def test_dashboard_status_counts_totals():
    inputs = DashboardInputs(
        entity_name="TestCo",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={
            EBITDA_MARGIN: RatioSeries(latest=0.30),      # green
            DEBT_TO_EQUITY: RatioSeries(latest=3.0),      # red
        },
    )
    result = compute_dashboard(inputs)
    counts = result.status_counts()
    assert counts["green"] >= 1
    assert counts["red"] >= 1
    total = sum(counts.values())
    assert total == len(result.rows)


def test_dashboard_summary_flags_include_red_reds():
    inputs = DashboardInputs(
        entity_name="TestCo",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={
            DEBT_TO_EQUITY: RatioSeries(latest=3.0),
        },
    )
    result = compute_dashboard(inputs)
    assert any("Debt" in flag for flag in result.summary_flags)


def test_dashboard_sector_kpis_row_appended_for_bank():
    # NIM band is HIGHER 2.5%..4.0% — 4.5% should be green, 3.2% amber.
    inputs = DashboardInputs(
        entity_name="Regional Bank Corp.",
        period_label="FY2026",
        sector=Sector.BANK,
        ratios={ROA: RatioSeries(latest=0.012), NIM: RatioSeries(latest=0.045)},
    )
    result = compute_dashboard(inputs)
    kpi_rows = [r for r in result.rows if r.section.startswith("Sector KPIs")]
    assert kpi_rows, "bank sector KPI section should be present"
    nim_row = next((r for r in kpi_rows if r.ratio_id == NIM), None)
    assert nim_row is not None
    assert nim_row.status == Status.GREEN


# --------------------------------------------------------------------------- #
# Excel rendering — round-trip
# --------------------------------------------------------------------------- #
def test_render_dashboard_sheet_writes_all_sections():
    inputs = DashboardInputs(
        entity_name="RoundTrip Corp.",
        period_label="FY2026",
        sector=Sector.INDUSTRIAL,
        ratios={
            GROSS_MARGIN: RatioSeries(latest=0.30, prior=0.28,
                                      history=[0.25, 0.28, 0.30]),
            EBITDA_MARGIN: RatioSeries(latest=0.20),
            DEBT_TO_EQUITY: RatioSeries(latest=0.60, peer_median=0.80),
        },
    )
    result = compute_dashboard(inputs)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Dashboard"
    render_dashboard_sheet(ws, result)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    reopened = load_workbook(buf)
    ws2 = reopened["Ratio Dashboard"]
    all_text = [
        str(c.value) for row in ws2.iter_rows(values_only=False)
        for c in row if c.value is not None
    ]
    joined = "\n".join(all_text)
    assert "Profitability" in joined
    assert "Liquidity" in joined
    assert "Solvency" in joined
    assert "Efficiency" in joined
    assert "Cash Flow" in joined
    assert "Valuation" in joined
    assert "RoundTrip Corp." in joined
    assert DASHBOARD_SCHEMA_VERSION in joined
