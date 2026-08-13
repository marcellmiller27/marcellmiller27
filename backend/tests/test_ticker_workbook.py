# JHI-SIG: 69M2705M | Institutional per-ticker workbook tests | JHI Research & Analytics Firm, Inc. (proprietary)
"""Tests for the fundamental-ratio breakdown and the multi-sheet per-ticker workbook.
All inputs are synthetic — no network."""

import datetime as dt
import math
from io import BytesIO

from openpyxl import load_workbook

from app import equity_ratios, equity_technicals as t, ticker_workbook as tw
from app.fundamentals import SOURCE_EDGAR, SOURCE_SF1, EquityFundamentals, FundamentalsYear


def _synthetic_bars(n: int = 400, drift: float = 0.0012) -> list[dict]:
    bars: list[dict] = []
    d = dt.date(2023, 1, 2)
    p = 100.0
    i = 0
    while len(bars) < n:
        if d.weekday() < 5:
            p = p * (1 + drift) + 1.5 * math.sin(i / 11.0)
            bars.append(
                {
                    "date": d.isoformat(),
                    "open": p * 0.999,
                    "high": p * 1.008 + 0.3,
                    "low": p * 0.992 - 0.3,
                    "close": p,
                    "volume": 1_000_000,
                }
            )
            i += 1
        d += dt.timedelta(days=1)
    return bars


def _rich_fundamentals(source: str = SOURCE_SF1) -> EquityFundamentals:
    years = [
        FundamentalsYear(
            fiscal_year=y,
            revenue=1e9 * (1.1 ** (y - 2020)),
            net_income=1.5e8 * (1.1 ** (y - 2020)),
            operating_income=2e8 * (1.1 ** (y - 2020)),
            stockholders_equity=8e8,
            gross_margin=0.42,
            operating_margin=0.2,
            net_margin=0.15,
            eps=1.2 * (1.1 ** (y - 2020)),
        )
        for y in range(2020, 2025)
    ]
    return EquityFundamentals(
        ticker="SYN",
        entity_name="Synthetic Corp.",
        source=source,
        fiscal_year=2024,
        period_end="2024-12-31",
        revenue=1.46e9,
        operating_income=2.9e8,
        net_income=2.2e8,
        stockholders_equity=8e8,
        gross_margin=0.42,
        operating_margin=0.2,
        net_margin=0.15,
        shares_outstanding=1.8e8,
        ebitda=3.5e8,
        free_cash_flow=1.9e8,
        total_assets=2.5e9,
        total_liabilities=1.7e9,
        total_debt=6e8,
        current_assets=9e8,
        current_liabilities=5e8,
        inventory=2e8,
        cash_and_equivalents=3e8,
        ebit=2.9e8,
        interest_expense=3e7,
        eps=1.22,
        eps_basic=1.25,
        dividends_per_share=0.4,
        roic=0.18,
        roe=0.275,
        years=years,
    )


# ── Ratio breakdown ──────────────────────────────────────────────────────────
def test_ratios_build_with_synthetic_fundamentals() -> None:
    fin = _rich_fundamentals()
    r = equity_ratios.compute_ratios(fin, price=50.0)
    flat = {m.label: m.value for s in r.sections for m in s.metrics}
    # Core ratios computed from the synthetic inputs.
    assert flat["Gross margin"] == 0.42
    assert flat["Debt-to-equity"] == fin.total_debt / fin.stockholders_equity
    assert flat["Current ratio"] == fin.current_assets / fin.current_liabilities
    # Quick ratio excludes inventory.
    assert flat["Quick ratio"] == (fin.current_assets - fin.inventory) / fin.current_liabilities
    # Interest coverage = EBIT / |interest expense|.
    assert flat["Interest coverage"] == fin.ebit / abs(fin.interest_expense)
    # FCF margin = FCF / revenue.
    assert flat["FCF margin"] == fin.free_cash_flow / fin.revenue
    # P/E = price / diluted EPS.
    assert flat["Price / earnings (P/E)"] == 50.0 / fin.eps
    # Multi-period trend with YoY populated after the first year.
    assert len(r.trend) == 5
    assert r.trend[0].revenue_yoy is None
    assert r.trend[-1].revenue_yoy is not None and r.trend[-1].revenue_yoy > 0
    assert "Sharadar SF1" in r.source_label


def test_ratios_degrade_gracefully_on_sparse_edgar() -> None:
    # EDGAR-style sparse bundle: no current assets/debt/interest → those ratios are None,
    # never fabricated.
    fin = EquityFundamentals(
        ticker="SPARSE",
        entity_name="Sparse Inc.",
        source=SOURCE_EDGAR,
        fiscal_year=2024,
        period_end="2024-12-31",
        revenue=5e8,
        net_income=6e7,
        stockholders_equity=4e8,
        gross_margin=0.3,
        shares_outstanding=1e8,
        total_assets=1e9,
        years=[FundamentalsYear(fiscal_year=2024, revenue=5e8, net_income=6e7)],
    )
    r = equity_ratios.compute_ratios(fin, price=None)
    flat = {m.label: m.value for s in r.sections for m in s.metrics}
    assert flat["Current ratio"] is None
    assert flat["Interest coverage"] is None
    assert flat["Price / earnings (P/E)"] is None  # no price
    # But ROA is derivable from net income / total assets.
    assert flat["Return on assets (ROA)"] == fin.net_income / fin.total_assets
    assert "SEC EDGAR" in r.source_label


# ── Full workbook ────────────────────────────────────────────────────────────
def _assemble_data(with_valuation: bool = True) -> tw.TickerWorkbookData:
    bars = _synthetic_bars(450, drift=0.0012)
    db = t.to_bars(bars)
    daily = t.compute_technicals(db, "SYN", "Daily", window=252)
    weekly = t.compute_technicals(t.aggregate_weekly(db), "SYN", "Weekly", window=52)
    oc = t.options_context(daily, weekly)
    fin = _rich_fundamentals()
    ratios = equity_ratios.compute_ratios(fin, price=daily.price)
    valuation = None
    valuation_error = None
    if with_valuation:
        import app.equity_valuation as ev

        ev.fundamentals.equity_fundamentals = lambda tk, max_years=5: fin  # type: ignore[assignment]
        valuation = ev.value_equity("SYN", price=daily.price, risk_free=0.045)
    else:
        valuation_error = "positive earnings required"
    return tw.TickerWorkbookData(
        ticker="SYN",
        daily=daily,
        weekly=weekly,
        options=oc,
        ratios=ratios,
        valuation=valuation,
        valuation_error=valuation_error,
        price=daily.price,
        market_cap=daily.price * fin.shares_outstanding,
        name=fin.entity_name,
    )


def test_full_workbook_opens_with_all_sheets() -> None:
    data = _assemble_data(with_valuation=True)
    content = tw.render_workbook(data)
    assert content[:2] == b"PK"  # xlsx is a zip archive
    assert len(content) > 8000
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == [
        "Cover & Summary",
        "Technicals — Daily",
        "Technicals — Weekly",
        "Options context",
        "Fundamental Ratios",
        "DCF Valuation",
        "Legal & Provenance",
    ]
    # Cover carries branding and the ticker.
    cover = wb["Cover & Summary"]
    assert cover["A1"].value == "Aegira"
    joined = " ".join(
        str(c.value) for col in cover.iter_cols() for c in col if c.value is not None
    )
    assert "SYN" in joined and "Opportunity summary" in joined


def test_workbook_degrades_when_valuation_unavailable() -> None:
    data = _assemble_data(with_valuation=False)
    content = tw.render_workbook(data)
    wb = load_workbook(BytesIO(content))
    assert "DCF Valuation" in wb.sheetnames
    val = wb["DCF Valuation"]
    text = " ".join(str(c.value) for col in val.iter_cols() for c in col if c.value is not None)
    assert "unavailable" in text.lower()  # labeled, not fabricated


def test_technicals_sheets_include_setups_and_indicators() -> None:
    data = _assemble_data(with_valuation=True)
    content = tw.render_workbook(data)
    wb = load_workbook(BytesIO(content))
    for name in ("Technicals — Daily", "Technicals — Weekly"):
        ws = wb[name]
        text = " ".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
        assert "Trade setups" in text
        assert "RSI (14)" in text
        assert "Market structure" in text


def test_ratios_sheet_has_trend_table() -> None:
    data = _assemble_data(with_valuation=True)
    content = tw.render_workbook(data)
    wb = load_workbook(BytesIO(content))
    ws = wb["Fundamental Ratios"]
    text = " ".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
    assert "Multi-period trend" in text
    assert "Debt-to-equity" in text
    assert "Return on equity (ROE)" in text
