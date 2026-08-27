# JHI-SIG: 69M2705M | QoE bridge workbook tests | JHI Research & Analytics Firm, Inc. (proprietary)
"""Smoke tests for the EBITDA_Bridge sheet renderer + its integration into the
diligence workbook."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.excel_export import diligence_workbook
from app.financial_diligence import analyze
from app.financial_diligence_models import DiligenceInput
from app.qoe_bridge import (
    AdjustmentInput,
    BridgeInput,
    EvidenceGrade,
    Side,
    build_bridge,
)
from app.qoe_bridge_workbook import render_ebitda_bridge_sheet


def test_bridge_sheet_renders_headers_and_totals():
    bridge = build_bridge(BridgeInput(
        business_name="Carrollton Design Build",
        period_label="FY2025",
        reported_ebitda=1_200_000,
        ltm_ebitda=1_200_000,
        three_year_avg_ebitda=1_050_000,
        adjustments=[
            AdjustmentInput(
                category_id="owner_comp_market",
                amount=150_000, sign=+1, side=Side.BOTH,
                evidence_grade=EvidenceGrade.A,
                evidence_source="BLS OEWS 11-1021 · Dallas · 2025",
            ),
            AdjustmentInput(
                category_id="one_time_legal",
                amount=80_000, sign=+1, side=Side.SELLER,
                evidence_grade=EvidenceGrade.B,
                evidence_source="Matter #45231 closed 2026-Q1",
            ),
        ],
    ))
    wb = Workbook()
    ws = wb.active
    ws.title = "EBITDA Bridge"
    render_ebitda_bridge_sheet(ws, bridge)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    reopened = load_workbook(buf)
    ws2 = reopened["EBITDA Bridge"]
    text = "\n".join(
        str(c.value) for row in ws2.iter_rows(values_only=False)
        for c in row if c.value is not None
    )
    assert "Aegira · EBITDA Bridge" in text
    assert "Reported EBITDA" in text
    assert "Adjusted EBITDA (Seller view)" in text
    assert "Adjusted EBITDA (Buyer view)" in text
    assert "JHI-SIG: 69M2705M" in text


def test_diligence_workbook_now_carries_bridge_and_dashboard():
    deal = DiligenceInput(
        business_name="Carrollton Design Build",
        industry="construction",
        period_label="FY2025",
        revenue=8_500_000,
        reported_ebitda=1_400_000,
        addbacks_claimed=180_000,
        questionable_addbacks=100_000,
        one_time_items=60_000,
        bank_deposits=8_600_000,
        accounts_receivable=1_100_000,
        inventory=250_000,
        accounts_payable=520_000,
        recurring_revenue_pct=35,
        customer_concentration_pct=28,
        debt_like_items=180_000,
        asking_price=6_500_000,
        post_loi=True,
    )
    report = analyze(deal)
    xlsx = diligence_workbook(deal, report)
    buf = BytesIO(xlsx)
    wb = load_workbook(buf)
    sheet_names = wb.sheetnames
    assert "Dashboard" in sheet_names
    assert "EBITDA Bridge" in sheet_names
    assert "Ratio Dashboard" in sheet_names
